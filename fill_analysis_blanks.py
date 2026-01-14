#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
분석표의 공란을 기초자료 수집표에서 추론하여 채우는 스크립트
- 집계표: 기초자료에서 직접 값 복사
- 분석표: 증감률, 기여도 계산 (집계표 기반)
- 비공개 자료: 가중치를 이용한 계산
"""

from openpyxl import load_workbook
import pathlib
import re
from typing import Optional, Dict, Tuple, List

def norm(s):
    """헤더 텍스트 정규화"""
    if s is None:
        return ''
    s = str(s).replace('\n', '')
    s = re.sub(r'\s+', '', s)
    return s

def safe_float(value, default=None):
    """안전한 float 변환"""
    if value is None:
        return default
    try:
        if isinstance(value, str):
            value = value.strip()
            if value in ('-', '', '없음', 'nan', 'none', 'n/a'):
                return default
        result = float(value)
        return result if not (result != result) else default  # NaN 체크
    except (ValueError, TypeError):
        return default

def get_cell_value(ws, row, col):
    """셀 값 가져오기 (수식이면 계산된 값)"""
    cell = ws.cell(row, col)
    if cell.data_type == 'f':
        # 수식인 경우 계산된 값 사용
        return cell.value if hasattr(cell, 'value') else None
    return cell.value

# 시트 매핑
SHEET_MAPPING = {
    'A(광공업생산)집계': '광공업생산',
    'B(서비스업생산)집계': '서비스업생산',
    'C(소비)집계': '소비(소매, 추가)',
    'D(고용률)집계': '연령별고용률',
    'D(취업자 수)집계': '고용',
    'D(실업)집계': '실업자 수',
    'E(지출목적물가)집계': '지출목적별 물가',
    'E(품목성질물가)집계': '품목성질별 물가',
    "F'(건설)집계": '건설 (공표자료)',
    "F''(지방용) 집계": '건설 (비공표자료)',
    'G(수출)집계': '수출',
    'H(수입)집계': '수입',
    'I(순인구이동)집계': '시도 간 이동',
    'J(시군구순인구이동)': '시군구인구이동',
}

ANALYSIS_TO_AGGREGATE = {
    'A 분석': 'A(광공업생산)집계',
    'B 분석': 'B(서비스업생산)집계',
    'C 분석': 'C(소비)집계',
    'D(고용률)분석': 'D(고용률)집계',
    'D(취업자 수)분석': 'D(취업자 수)집계',
    'D(실업)분석': 'D(실업)집계',
    'E(지출목적물가) 분석': 'E(지출목적물가)집계',
    'E(품목성질물가)분석': 'E(품목성질물가)집계',
    "F'분석": "F'(건설)집계",
    "F'' 분석": "F''(지방용) 집계",
    'G 분석': 'G(수출)집계',
    'H 분석': 'H(수입)집계',
}

def find_header_row(ws, max_scan=10):
    """헤더 행 찾기"""
    for r in range(1, min(max_scan, ws.max_row + 1)):
        headers = [norm(ws.cell(r, c).value) for c in range(1, min(30, ws.max_column + 1))]
        if any(h in ('지역코드', '지역이름', '산업코드', '산업이름', '분류단계') for h in headers):
            return r
    return None

def build_header_map(ws, header_row):
    """헤더 맵 생성"""
    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(header_row, c).value)
        if v:
            header_map[v] = c
    return header_map

def fill_aggregation_from_base(wb_a, wb_b, sheet_a_name, sheet_b_name):
    """집계표를 기초자료에서 채우기"""
    if sheet_a_name not in wb_a.sheetnames or sheet_b_name not in wb_b.sheetnames:
        return 0
    
    ws_a = wb_a[sheet_a_name]
    ws_b = wb_b[sheet_b_name]
    
    hdr_a = find_header_row(ws_a)
    hdr_b = find_header_row(ws_b)
    if hdr_a is None or hdr_b is None:
        return 0
    
    ha = build_header_map(ws_a, hdr_a)
    hb = build_header_map(ws_b, hdr_b)
    
    # 가중치 컬럼 확인 (먼저 확인)
    weight_col_a = ha.get('가중치')
    weight_col_b = hb.get('가중치')
    
    # 키 컬럼
    key_cols = []
    for kc in ['지역코드', '지역이름', '분류단계', '산업코드', '산업이름', '업태종류', '분류코드']:
        if kc in ha and kc in hb:
            key_cols.append(kc)
    
    if not key_cols:
        return 0
    
    # 연도/분기 컬럼 찾기
    value_cols = []
    for name, col_a in ha.items():
        if re.match(r'^\d{4}$', name) and name in hb:
            value_cols.append((name, col_a, hb[name]))
        elif re.search(r'\d{4}[.\s]*\d/4', name) and name in hb:
            value_cols.append((name, col_a, hb[name]))
    
    # 건설 시트 특별 처리: 헤더가 없는 컬럼도 채우기
    if "F'" in sheet_a_name or "건설" in sheet_a_name:
        # 기초자료에서 2025 3/4p, 2024 3/4 찾기
        target_quarters = []
        for name, col_b in hb.items():
            if '2025' in str(name) and ('3/4' in str(name) or '3분기' in str(name)):
                target_quarters.append(('2025_3q', None, col_b))  # 분석표 컬럼은 나중에 결정
            elif '2024' in str(name) and ('3/4' in str(name) or '3분기' in str(name)):
                target_quarters.append(('2024_3q', None, col_b))
        
        # 분석표에서 헤더가 없는 컬럼 찾기 (열 24, 25 등)
        for c in range(ws_a.max_column, 0, -1):
            header_val = norm(ws_a.cell(hdr_a, c).value)
            if not header_val and c not in [col for _, col, _ in value_cols]:
                # 빈 헤더 컬럼에 분기 데이터 매핑
                if len(target_quarters) > 0:
                    quarter_name, _, base_col = target_quarters.pop(0)
                    value_cols.append((quarter_name, c, base_col))
    
    # 기초자료 인덱스
    base_index = {}
    for r in range(hdr_b + 1, ws_b.max_row + 1):
        key = []
        empty = True
        for kc in key_cols:
            val = get_cell_value(ws_b, r, hb[kc])
            if val not in (None, ''):
                empty = False
            key.append(str(val).strip() if val is not None else '')
        if not empty:
            base_index[tuple(key)] = r
    
    # 채우기
    filled = 0
    for r in range(hdr_a + 1, ws_a.max_row + 1):
        key = []
        empty = True
        for kc in key_cols:
            val = get_cell_value(ws_a, r, ha[kc])
            if val not in (None, ''):
                empty = False
            key.append(str(val).strip() if val is not None else '')
        if empty:
            continue
        
        br = base_index.get(tuple(key))
        if br is None:
            continue
        
        # 가중치 처리: 결측치는 N/A로 표시 (기본값으로 채우지 않음)
        if weight_col_a and weight_col_b:
            cell_weight = ws_a.cell(r, weight_col_a)
            if cell_weight.value in (None, ''):
                weight_val = get_cell_value(ws_b, br, weight_col_b)
                if weight_val not in (None, '') and weight_val != 0:
                    cell_weight.value = weight_val
                    filled += 1
                else:
                    # 결측치나 0인 경우 N/A로 표시
                    cell_weight.value = 'N/A'
        
        # 다른 값 채우기
        for name, col_a, col_b in value_cols:
            if col_a is None or name == '가중치':  # 가중치는 이미 처리
                continue
            cell_a = ws_a.cell(r, col_a)
            if cell_a.value in (None, ''):
                val_b = get_cell_value(ws_b, br, col_b)
                if val_b not in (None, ''):
                    cell_a.value = val_b
                    filled += 1
    
    return filled

def fill_analysis_from_aggregate(wb_a, sheet_a_name):
    """분석표를 집계표에서 계산하여 채우기"""
    if sheet_a_name not in wb_a.sheetnames:
        return 0
    
    agg_sheet_name = ANALYSIS_TO_AGGREGATE.get(sheet_a_name)
    if not agg_sheet_name or agg_sheet_name not in wb_a.sheetnames:
        return 0
    
    ws_analysis = wb_a[sheet_a_name]
    ws_aggregate = wb_a[agg_sheet_name]
    
    hdr_analysis = find_header_row(ws_analysis)
    hdr_agg = find_header_row(ws_aggregate)
    if hdr_analysis is None or hdr_agg is None:
        return 0
    
    ha_analysis = build_header_map(ws_analysis, hdr_analysis)
    ha_agg = build_header_map(ws_aggregate, hdr_agg)
    
    # 증감률, 기여도 컬럼 찾기
    growth_col = None
    contrib_col = None
    for name, col in ha_analysis.items():
        if ('증감률' in name or '전년동기비' in name or '증감' in name) and 'X' not in name:
            growth_col = col
        if '기여도' in name or '기여율' in name:
            contrib_col = col
    
    # 건설 시트는 증감률 컬럼이 없을 수 있으므로 계산 필요
    is_construction = "F'" in sheet_a_name or "건설" in sheet_a_name
    
    # 연도 컬럼 찾기
    year_cols_agg = {}
    for name, col in ha_agg.items():
        if re.match(r'^\d{4}$', name):
            year_cols_agg[name] = col
    
    # 최신 연도와 전년도
    years = sorted([int(y) for y in year_cols_agg.keys() if y.isdigit()], reverse=True)
    if len(years) < 2:
        return 0
    
    curr_year_col = year_cols_agg.get(str(years[0]))
    prev_year_col = year_cols_agg.get(str(years[1]))
    
    # 키 컬럼 (건설 시트는 공정이름도 포함)
    key_cols = []
    for kc in ['지역이름', '분류단계', '산업코드', '분류코드', '공정이름']:
        if kc in ha_analysis and kc in ha_agg:
            key_cols.append(kc)
    # 공정이름이 한쪽에만 있어도 키로 사용
    if '공정이름' in ha_analysis and '공정이름' not in ha_agg:
        # 집계표에 공정이름이 없으면 분류코드로 매칭
        if '분류코드' in ha_analysis and '분류코드' in ha_agg:
            key_cols.append('분류코드')
    if not key_cols:
        return 0
    
    # 집계표 인덱스
    agg_index = {}
    for r in range(hdr_agg + 1, ws_aggregate.max_row + 1):
        key = []
        for kc in key_cols:
            val = get_cell_value(ws_aggregate, r, ha_agg[kc])
            key.append(str(val).strip() if val is not None else '')
        agg_index[tuple(key)] = r
    
    # 가중치 컬럼
    weight_col = ha_agg.get('가중치')
    
    filled = 0
    
    for r in range(hdr_analysis + 1, ws_analysis.max_row + 1):
        # 키 추출 (수식 참조 처리)
        key = []
        for kc in key_cols:
            val = get_cell_value(ws_analysis, r, ha_analysis[kc])
            # 수식 참조인 경우 집계표에서 직접 가져오기
            if isinstance(val, str) and '!' in val:
                # 수식 참조 파싱 (예: 'A(광공업생산)집계'!E5)
                match = re.search(r"!([A-Z]+)(\d+)", val)
                if match:
                    ref_col = match.group(1)
                    ref_row = int(match.group(2))
                    # 열 번호로 변환
                    col_num = 0
                    for char in ref_col:
                        col_num = col_num * 26 + (ord(char) - ord('A') + 1)
                    if ref_row <= ws_aggregate.max_row and col_num <= ws_aggregate.max_column:
                        val = get_cell_value(ws_aggregate, ref_row, col_num)
            key.append(str(val).strip() if val is not None else '')
        
        agg_row = agg_index.get(tuple(key))
        if agg_row is None:
            continue
        
        # 증감률 계산
        growth_rate = None
        if growth_col and curr_year_col and prev_year_col:
            cell_growth = ws_analysis.cell(r, growth_col)
            if cell_growth.value in (None, ''):
                curr_val = safe_float(get_cell_value(ws_aggregate, agg_row, curr_year_col))
                prev_val = safe_float(get_cell_value(ws_aggregate, agg_row, prev_year_col))
                if curr_val is not None and prev_val is not None and prev_val != 0:
                    growth_rate = ((curr_val - prev_val) / prev_val) * 100
                    cell_growth.value = round(growth_rate, 1)
                    filled += 1
        elif is_construction and curr_year_col and prev_year_col:
            # 건설 시트는 증감률 컬럼이 없어도 계산만 함
            curr_val = safe_float(get_cell_value(ws_aggregate, agg_row, curr_year_col))
            prev_val = safe_float(get_cell_value(ws_aggregate, agg_row, prev_year_col))
            if curr_val is not None and prev_val is not None and prev_val != 0:
                growth_rate = ((curr_val - prev_val) / prev_val) * 100
        
        # 기여도 계산
        if contrib_col:
            cell_contrib = ws_analysis.cell(r, contrib_col)
            if cell_contrib.value in (None, ''):
                # 증감률 가져오기
                growth_val = growth_rate
                if growth_val is None:
                    if growth_col:
                        growth_val = safe_float(ws_analysis.cell(r, growth_col).value)
                    if growth_val is None and curr_year_col and prev_year_col:
                        curr_val = safe_float(get_cell_value(ws_aggregate, agg_row, curr_year_col))
                        prev_val = safe_float(get_cell_value(ws_aggregate, agg_row, prev_year_col))
                        if curr_val is not None and prev_val is not None and prev_val != 0:
                            growth_val = ((curr_val - prev_val) / prev_val) * 100
                
                # 가중치 가져오기 (건설은 가중치가 없을 수 있으므로 다른 방식 사용)
                weight = None
                if weight_col:
                    weight_val = get_cell_value(ws_aggregate, agg_row, weight_col)
                    # N/A나 결측치는 None으로 처리
                    if weight_val not in (None, '', 'N/A'):
                        try:
                            weight = float(weight_val)
                            if weight == 0:
                                weight = None  # 0도 None으로 처리
                        except (ValueError, TypeError):
                            weight = None
                
                # 건설 시트는 가중치 대신 다른 방식으로 기여도 계산할 수 있음
                # 여기서는 기본적으로 가중치가 있으면 사용
                if growth_val is not None and weight is not None:
                    contribution = growth_val * weight / 100.0
                    cell_contrib.value = round(contribution, 6)
                    filled += 1
                elif is_construction and growth_val is not None:
                    # 건설은 가중치 없이도 증감률을 기여도로 사용할 수 있음
                    cell_contrib.value = round(growth_val, 6)
                    filled += 1
                # 가중치가 N/A인 경우 기여도는 계산하지 않음 (결측치로 유지)
    
    return filled

def main():
    analysis_path = pathlib.Path('/Users/topos/Downloads/분석표_2025년_3분기_자동생성 (2).xlsx')
    base_path = pathlib.Path('기초자료 수집표_2025년 3분기(캡스톤).xlsx')
    if not base_path.exists():
        base_path = pathlib.Path('uploads/기초자료_수집표_2025년_3분기캡스톤_2b806a5b.xlsx')
    
    if not analysis_path.exists():
        print(f"분석표 파일을 찾을 수 없습니다: {analysis_path}")
        return
    
    if not base_path.exists():
        print(f"기초자료 파일을 찾을 수 없습니다: {base_path}")
        return
    
    print(f"분석표: {analysis_path}")
    print(f"기초자료: {base_path}")
    
    wb_a = load_workbook(analysis_path, data_only=False)
    wb_b = load_workbook(base_path, data_only=True)
    
    total_filled = 0
    
    # 1. 집계표 채우기 (가중치 포함)
    print("\n=== 집계표 채우기 ===")
    for sheet_a_name, sheet_b_name in SHEET_MAPPING.items():
        filled = fill_aggregation_from_base(wb_a, wb_b, sheet_a_name, sheet_b_name)
        if filled > 0:
            print(f"  {sheet_a_name}: {filled}개 셀 채움")
            total_filled += filled
    
    # 2. 가중치 채우기 (명시적으로)
    print("\n=== 가중치 채우기 ===")
    weight_sheets = {
        'A(광공업생산)집계': '광공업생산',
        'B(서비스업생산)집계': '서비스업생산',
        'C(소비)집계': '소비(소매, 추가)',
        'E(지출목적물가)집계': '지출목적별 물가',
        'E(품목성질물가)집계': '품목성질별 물가',
    }
    
    for sheet_a_name, sheet_b_name in weight_sheets.items():
        if sheet_a_name not in wb_a.sheetnames or sheet_b_name not in wb_b.sheetnames:
            continue
        
        ws_a = wb_a[sheet_a_name]
        ws_b = wb_b[sheet_b_name]
        
        hdr_a = find_header_row(ws_a)
        hdr_b = find_header_row(ws_b)
        if hdr_a is None or hdr_b is None:
            continue
        
        ha = build_header_map(ws_a, hdr_a)
        hb = build_header_map(ws_b, hdr_b)
        
        weight_col_a = ha.get('가중치')
        weight_col_b = hb.get('가중치')
        
        if not weight_col_a or not weight_col_b:
            continue
        
        # 키 컬럼
        key_cols = []
        for kc in ['지역코드', '지역이름', '분류단계', '산업코드', '산업이름', '업태종류', '분류코드']:
            if kc in ha and kc in hb:
                key_cols.append(kc)
        
        if not key_cols:
            continue
        
        # 기초자료 인덱스
        base_index = {}
        for r in range(hdr_b + 1, ws_b.max_row + 1):
            key = []
            empty = True
            for kc in key_cols:
                val = get_cell_value(ws_b, r, hb[kc])
                if val not in (None, ''):
                    empty = False
                key.append(str(val).strip() if val is not None else '')
            if not empty:
                base_index[tuple(key)] = r
        
        # 가중치 채우기
        filled = 0
        for r in range(hdr_a + 1, ws_a.max_row + 1):
            key = []
            empty = True
            for kc in key_cols:
                val = get_cell_value(ws_a, r, ha[kc])
                if val not in (None, ''):
                    empty = False
                key.append(str(val).strip() if val is not None else '')
            if empty:
                continue
            
            cell_weight = ws_a.cell(r, weight_col_a)
            if cell_weight.value in (None, ''):
                br = base_index.get(tuple(key))
                if br:
                    weight_val = get_cell_value(ws_b, br, weight_col_b)
                    if weight_val not in (None, '') and weight_val != 0:
                        cell_weight.value = weight_val
                        filled += 1
                    else:
                        # 결측치나 0인 경우 N/A로 표시
                        cell_weight.value = 'N/A'
        
        if filled > 0:
            print(f"  {sheet_a_name}: {filled}개 가중치 채움")
            total_filled += filled
    
    # 3. 분석표 채우기 (가중치가 채워진 후 기여도 등 계산)
    print("\n=== 분석표 채우기 ===")
    for sheet_a_name in ANALYSIS_TO_AGGREGATE.keys():
        filled = fill_analysis_from_aggregate(wb_a, sheet_a_name)
        if filled > 0:
            print(f"  {sheet_a_name}: {filled}개 셀 채움")
            total_filled += filled
    
    # 저장
    out_path = pathlib.Path('exports/분석표_2025년_3분기_자동생성_공란채움.xlsx')
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb_a.save(out_path)
    
    print(f"\n총 {total_filled}개 셀을 채웠습니다.")
    print(f"저장 위치: {out_path}")

if __name__ == '__main__':
    main()
