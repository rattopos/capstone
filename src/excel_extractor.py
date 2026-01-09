"""
엑셀 데이터 추출 모듈
엑셀 파일에서 시트별 셀 값 추출 기능 제공
"""

import re
from pathlib import Path
from typing import Any, List, Tuple, Union
import openpyxl
from openpyxl import load_workbook


class ExcelExtractor:
    """엑셀 파일에서 데이터를 추출하는 클래스"""
    
    def __init__(self, excel_path: str):
        """
        엑셀 추출기 초기화
        
        Args:
            excel_path: 엑셀 파일 경로
        """
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.sheets = {}
        
    def load_workbook(self) -> None:
        """
        엑셀 파일을 로드합니다.
        
        Raises:
            FileNotFoundError: 엑셀 파일이 존재하지 않을 때
            IOError: 파일 읽기 실패 시
        """
        if not self.excel_path.exists():
            raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {self.excel_path}")
        
        try:
            self.workbook = load_workbook(self.excel_path, data_only=True)
        except IOError as e:
            raise IOError(f"엑셀 파일 읽기 실패: {e}")
    
    def get_sheet_names(self) -> List[str]:
        """
        엑셀 파일의 모든 시트 이름을 반환합니다.
        "완료체크" 시트는 제외됩니다.
        
        Returns:
            시트 이름 리스트 (완료체크 제외)
        """
        if self.workbook is None:
            self.load_workbook()
        
        # "완료체크" 시트 제외
        excluded_sheets = ['완료체크']
        return [name for name in self.workbook.sheetnames if name not in excluded_sheets]
    
    def get_sheet(self, sheet_name: str):
        """
        특정 시트 객체를 가져옵니다.
        
        Args:
            sheet_name: 시트 이름
            
        Returns:
            openpyxl Worksheet 객체
            
        Raises:
            ValueError: 시트가 존재하지 않을 때
        """
        if self.workbook is None:
            self.load_workbook()
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"시트 '{sheet_name}'을 찾을 수 없습니다. "
                           f"사용 가능한 시트: {', '.join(self.workbook.sheetnames)}")
        
        if sheet_name not in self.sheets:
            self.sheets[sheet_name] = self.workbook[sheet_name]
        
        return self.sheets[sheet_name]
    
    def parse_cell_address(self, cell_address: str) -> Tuple[int, int]:
        """
        셀 주소 문자열을 행과 열 인덱스로 변환합니다.
        
        Args:
            cell_address: 셀 주소 (예: 'A1', 'B5')
            
        Returns:
            (행 번호, 열 번호) 튜플 (1-based 인덱스)
            
        Raises:
            ValueError: 잘못된 셀 주소 형식일 때
        """
        # 정규식으로 셀 주소 파싱 (예: 'A1', 'AA100')
        match = re.match(r'([A-Z]+)(\d+)', cell_address.upper())
        if not match:
            raise ValueError(f"잘못된 셀 주소 형식: {cell_address}")
        
        col_str = match.group(1)
        row_num = int(match.group(2))
        
        # 열 문자열을 숫자로 변환 (A=1, B=2, ..., Z=26, AA=27, ...)
        col_num = 0
        for char in col_str:
            col_num = col_num * 26 + (ord(char) - ord('A') + 1)
        
        return row_num, col_num
    
    def get_cell_value(self, sheet_name: str, cell_address: str) -> Any:
        """
        특정 셀의 값을 가져옵니다.
        
        Args:
            sheet_name: 시트 이름
            cell_address: 셀 주소 (예: 'A1')
            
        Returns:
            셀 값 (숫자, 문자열, 날짜 등)
            
        Raises:
            ValueError: 잘못된 셀 주소 또는 시트 이름일 때
        """
        sheet = self.get_sheet(sheet_name)
        row_num, col_num = self.parse_cell_address(cell_address)
        
        cell = sheet.cell(row=row_num, column=col_num)
        value = cell.value
        
        return value
    
    def get_cell_range(self, sheet_name: str, range_address: str) -> List[Any]:
        """
        셀 범위의 모든 값을 가져옵니다.
        
        Args:
            sheet_name: 시트 이름
            range_address: 셀 범위 (예: 'A1:A5' 또는 'A1:B2')
            
        Returns:
            셀 값 리스트 (행 우선 순서)
            
        Raises:
            ValueError: 잘못된 범위 형식일 때
        """
        # 범위 형식 파싱 (예: 'A1:A5' 또는 'A1:B2')
        if ':' not in range_address:
            # 단일 셀인 경우
            return [self.get_cell_value(sheet_name, range_address)]
        
        start_cell, end_cell = range_address.split(':')
        start_row, start_col = self.parse_cell_address(start_cell.strip())
        end_row, end_col = self.parse_cell_address(end_cell.strip())
        
        values = []
        sheet = self.get_sheet(sheet_name)
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                values.append(cell.value)
        
        return values
    
    def parse_cell_reference(self, cell_reference: str) -> Union[str, List[str]]:
        """
        셀 참조를 파싱합니다. 단일 셀이나 범위를 처리합니다.
        
        Args:
            cell_reference: 셀 참조 문자열 (예: 'A1' 또는 'A1:A5')
            
        Returns:
            단일 셀인 경우 셀 주소 문자열, 범위인 경우 셀 주소 리스트
        """
        if ':' in cell_reference:
            # 범위인 경우
            start_cell, end_cell = cell_reference.split(':')
            return [start_cell.strip(), end_cell.strip()]
        else:
            # 단일 셀인 경우
            return cell_reference.strip()
    
    def extract_value(self, sheet_name: str, cell_reference: str) -> Union[Any, List[Any]]:
        """
        시트에서 셀 참조에 해당하는 값을 추출합니다.
        
        Args:
            sheet_name: 시트 이름
            cell_reference: 셀 참조 (단일 셀이나 범위)
            
        Returns:
            단일 셀인 경우 값, 범위인 경우 값 리스트
        """
        parsed = self.parse_cell_reference(cell_reference)
        
        if isinstance(parsed, list):
            return self.get_cell_range(sheet_name, cell_reference)
        else:
            return self.get_cell_value(sheet_name, parsed)
    
    def close(self) -> None:
        """워크북을 닫습니다."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.sheets = {}
    
    def __enter__(self):
        """Context manager 진입"""
        self.load_workbook()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager 종료"""
        self.close()

