# -*- coding: utf-8 -*-
"""
Excel 파일 전처리 서비스

업로드된 엑셀 파일의 수식을 계산하여 분석 시트의 값을 채웁니다.
"""

import os
import shutil
from pathlib import Path
from typing import Optional, Tuple
import platform


def preprocess_excel(
    excel_path: str, 
    output_path: Optional[str] = None,
    use_xlwings: bool = False
) -> Tuple[str, bool, str]:
    """
    엑셀 파일의 수식을 계산하여 새 파일로 저장 (최적화 버전)
    
    openpyxl의 data_only=True를 최우선으로 사용하여 빠른 처리를 목표로 합니다.
    xlwings는 기본적으로 사용하지 않으며, 명시적으로 요청할 때만 실행됩니다.
    
    Args:
        excel_path: 원본 엑셀 파일 경로
        output_path: 출력 파일 경로 (None이면 원본 덮어쓰기)
        use_xlwings: xlwings 사용 여부 (기본값: False, Excel 앱 필요하고 느림)
    
    Returns:
        Tuple[str, bool, str]: (처리된 파일 경로, 성공 여부, 메시지)
    """
    if output_path is None:
        output_path = excel_path
    
    # 1. openpyxl data_only=True 사용 (가장 빠름 - 수식 계산된 값 직접 읽기)
    result_path, success, message = _try_openpyxl_data_only(excel_path, output_path)
    if success:
        return result_path, True, message
    
    # 2. openpyxl로 직접 계산 시도 (백엔드 직접 계산 - 시트 간 참조 매핑)
    result_path, success, message = _try_openpyxl_calculation(excel_path, output_path)
    if success:
        return result_path, True, message
    
    # 3. formulas 라이브러리 시도 (선택적 - 복잡한 수식 지원, 하지만 느릴 수 있음)
    # 주의: 대부분의 경우 위 두 방법으로 충분하므로, 필요시에만 사용
    try:
        import formulas
        result_path, success, message = _try_formulas(excel_path, output_path)
        if success:
            return result_path, True, message
    except ImportError:
        pass  # formulas가 없으면 건너뛰기
    
    # 4. xlwings 시도 (명시적으로 요청한 경우만 - Excel 앱 필요, 매우 느림)
    if use_xlwings:
        result_path, success, message = _try_xlwings(excel_path, output_path)
    if success:
        return result_path, True, message
    
    # 5. 모두 실패 시 원본 반환 (fallback 로직 사용)
    print(f"[전처리] 수식 계산 실패 - 원본 파일 사용, generator fallback 로직 활성화")
    return excel_path, False, "수식 계산 실패 - fallback 모드"


def _try_xlwings(excel_path: str, output_path: str) -> Tuple[str, bool, str]:
    """
    xlwings를 사용하여 Excel 앱으로 수식 계산 (명시적 요청 시에만 사용)
    
    주의: 
    - Excel 앱 실행이 필요하므로 매우 느립니다.
    - 기본 로직에서 제외되며, use_xlwings=True로 명시적으로 요청할 때만 실행됩니다.
    - 대부분의 경우 openpyxl data_only=True로 충분합니다.
    """
    try:
        import xlwings as xw
        
        print(f"[xlwings] Excel 앱으로 수식 계산 시작 (명시적 요청)...")
        
        # Excel 앱을 백그라운드에서 실행
        app = xw.App(visible=False)
        
        try:
            # 파일 열기
            wb = app.books.open(excel_path)
            
            # 모든 시트의 수식 강제 재계산
            wb.app.calculate()
            
            # 저장 (수식 결과가 캐시됨)
            wb.save(output_path)
            wb.close()
            
            print(f"[xlwings] 수식 계산 완료: {output_path}")
            return output_path, True, "xlwings로 수식 계산 완료"
            
        finally:
            app.quit()
            
    except ImportError:
        return excel_path, False, "xlwings 미설치"
    except Exception as e:
        print(f"[xlwings] 오류: {e}")
        return excel_path, False, f"xlwings 오류: {str(e)}"


def _try_formulas(excel_path: str, output_path: str) -> Tuple[str, bool, str]:
    """formulas 라이브러리를 사용하여 순수 Python으로 수식 계산"""
    try:
        import formulas
        
        print(f"[formulas] 순수 Python으로 수식 계산 시작...")
        
        # 수식 모델 생성
        xl_model = formulas.ExcelModel().loads(excel_path).finish()
        
        # 모든 수식 계산
        xl_model.calculate()
        
        # 결과 저장
        xl_model.write(output_path)
        
        print(f"[formulas] 수식 계산 완료: {output_path}")
        return output_path, True, "formulas로 수식 계산 완료"
        
    except ImportError:
        return excel_path, False, "formulas 미설치"
    except Exception as e:
        print(f"[formulas] 오류: {e}")
        return excel_path, False, f"formulas 오류: {str(e)}"


def _try_openpyxl_data_only(excel_path: str, output_path: str) -> Tuple[str, bool, str]:
    """
    openpyxl의 data_only=True를 사용하여 수식 계산된 값 직접 읽기 (가장 빠름)
    
    엑셀 파일이 이미 수식이 계산된 상태라면, data_only=True로 읽으면
    계산된 값이 바로 나옵니다. 별도 계산 과정이 필요 없어 가장 빠릅니다.
    """
    try:
        import openpyxl
        
        print(f"[openpyxl] data_only=True로 값 읽기 시도...")
        
        # 원본 파일 복사 (원본 보존)
        if excel_path != output_path:
            shutil.copy2(excel_path, output_path)
        
        # data_only=True로 열어서 계산된 값 확인
        wb = openpyxl.load_workbook(output_path, data_only=True)
        
        # 수식이 계산된 값이 있는지 확인 (첫 번째 시트의 일부 셀 확인)
        has_calculated_values = False
        for sheet_name in wb.sheetnames[:3]:  # 처음 3개 시트만 확인
            ws = wb[sheet_name]
            # 첫 10행, 첫 10열에서 값이 있는지 확인
            for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), 
                                   min_col=1, max_col=min(10, ws.max_column), 
                                   values_only=True):
                for cell_value in row:
                    if cell_value is not None and not isinstance(cell_value, str):
                        has_calculated_values = True
                        break
                if has_calculated_values:
                    break
            if has_calculated_values:
                break
        
        wb.close()
        
        if has_calculated_values:
            print(f"[openpyxl] data_only=True로 값 읽기 성공 (수식이 이미 계산된 상태)")
            return output_path, True, "openpyxl data_only=True로 값 읽기 완료"
        else:
            # 계산된 값이 없으면 다음 방법 시도
            return excel_path, False, "수식이 계산되지 않은 상태"
        
    except ImportError:
        return excel_path, False, "openpyxl 미설치"
    except Exception as e:
        print(f"[openpyxl data_only] 오류: {e}")
        return excel_path, False, f"openpyxl data_only 오류: {str(e)}"


def _try_openpyxl_calculation(excel_path: str, output_path: str) -> Tuple[str, bool, str]:
    """
    openpyxl을 사용하여 시트 간 참조 수식 계산 (백엔드 직접 계산)
    
    분석 시트의 수식이 집계 시트를 참조하는 경우:
    ='시트이름'!셀주소 형태의 간단한 참조를 백엔드에서 직접 계산하여 매핑
    
    주의: 이 방법은 data_only=True로 읽을 수 없는 경우에만 사용됩니다.
    대부분의 경우 _try_openpyxl_data_only가 더 빠르고 효율적입니다.
    """
    try:
        import openpyxl
        import re
        
        print(f"[openpyxl] 백엔드 직접 계산 시작 (시트 간 참조 매핑)...")
        
        # 분석 시트 → 집계 시트 매핑
        analysis_aggregate_mapping = {
            'A 분석': 'A(광공업생산)집계',
            'B 분석': 'B(서비스업생산)집계',
            'C 분석': 'C(소비)집계',
            'D(고용률)분석': 'D(고용률)집계',
            'D(실업)분석': 'D(실업)집계',
            "F'분석": "F'(건설)집계",
            'G 분석': 'G(수출)집계',
            'H 분석': 'H(수입)집계',
        }
        
        # 원본 파일 복사 (원본 보존)
        if excel_path != output_path:
            shutil.copy2(excel_path, output_path)
        
        # 수식 포함 모드로 열기
        wb = openpyxl.load_workbook(output_path, data_only=False)
        
        # 집계 시트 데이터 캐싱 (값 모드로 읽기)
        wb_data = openpyxl.load_workbook(output_path, data_only=True, read_only=True)
        aggregate_cache = {}
        
        # 필요한 집계 시트만 미리 캐싱
        required_aggregate_sheets = set(analysis_aggregate_mapping.values())
        for aggregate_sheet in required_aggregate_sheets:
            if aggregate_sheet in wb_data.sheetnames:
                ws_agg = wb_data[aggregate_sheet]
                sheet_data = {}
                # 값이 있는 셀만 저장 (성능 최적화)
                for row in ws_agg.iter_rows(min_row=1, max_row=min(ws_agg.max_row, 1000), values_only=False):
                    for cell in row:
                        if cell.value is not None:
                            sheet_data[(cell.row, cell.column)] = cell.value
                aggregate_cache[aggregate_sheet] = sheet_data
        
        wb_data.close()
        
        # 분석 시트의 수식을 값으로 대체
        calculated_count = 0
        formula_count = 0
        
        # 열 문자를 숫자로 변환
        def col_letter_to_number(col_letter: str) -> int:
            col = 0
            for i, c in enumerate(reversed(col_letter)):
                col += (ord(c) - ord('A') + 1) * (26 ** i)
            return col
        
        for analysis_sheet, aggregate_sheet in analysis_aggregate_mapping.items():
            if analysis_sheet not in wb.sheetnames or aggregate_sheet not in aggregate_cache:
                continue
            
            ws_analysis = wb[analysis_sheet]
            agg_data = aggregate_cache[aggregate_sheet]
            
            # 수식이 있는 셀만 처리
            for row in ws_analysis.iter_rows(min_row=1, max_row=min(ws_analysis.max_row, 1000)):
                for cell in row:
                    if cell.value is None:
                        continue
                    
                    val = str(cell.value)
                    if val.startswith('='):
                        formula_count += 1
                        # 집계 시트 참조 파싱: ='시트이름'!셀주소
                        match = re.match(r"^='?([^'!]+)'?!([A-Z]+)(\d+)$", val)
                        if match:
                            ref_sheet = match.group(1)
                            ref_col_letter = match.group(2)
                            ref_row = int(match.group(3))
                            ref_col = col_letter_to_number(ref_col_letter)
                            
                            if ref_sheet in aggregate_cache:
                                ref_value = aggregate_cache[ref_sheet].get((ref_row, ref_col))
                                if ref_value is not None:
                                    cell.value = ref_value
                                    calculated_count += 1
        
        wb.save(output_path)
        wb.close()
        
        if calculated_count > 0:
            print(f"[openpyxl] 백엔드 직접 계산 완료: {calculated_count}개 셀")
            return output_path, True, f"백엔드 직접 계산 완료 ({calculated_count}개 수식)"
        else:
            return excel_path, False, "계산할 수식 없음"
        
    except ImportError:
        return excel_path, False, "openpyxl 미설치"
    except Exception as e:
        print(f"[openpyxl 계산] 오류: {e}")
        return excel_path, False, f"openpyxl 계산 오류: {str(e)}"


def check_available_methods() -> dict:
    """사용 가능한 수식 계산 방법 확인"""
    methods = {
        'xlwings': False,
        'formulas': False,
        'openpyxl': False,
        'excel_installed': False,
    }
    
    # xlwings 확인
    try:
        import xlwings as xw
        methods['xlwings'] = True
        
        # Excel 설치 여부 확인 (Mac/Windows)
        if platform.system() == 'Darwin':  # macOS
            methods['excel_installed'] = os.path.exists('/Applications/Microsoft Excel.app')
        elif platform.system() == 'Windows':
            try:
                app = xw.App(visible=False)
                app.quit()
                methods['excel_installed'] = True
            except:
                pass
    except ImportError:
        pass
    
    # formulas 확인
    try:
        import formulas
        methods['formulas'] = True
    except ImportError:
        pass
    
    # openpyxl 확인 (항상 있어야 함)
    try:
        import openpyxl
        methods['openpyxl'] = True
    except ImportError:
        pass
    
    return methods


def get_recommended_method() -> str:
    """권장 수식 계산 방법 반환"""
    methods = check_available_methods()
    
    # 백엔드 직접 계산이 가장 빠르므로 우선 권장
    if methods['openpyxl']:
        return 'openpyxl (백엔드 직접 계산 - 가장 빠름)'
    elif methods['formulas']:
        return 'formulas (순수 Python - 복잡한 수식 지원)'
    elif methods['xlwings'] and methods['excel_installed']:
        return 'xlwings (Excel 앱 사용 - fallback, 느림)'
    else:
        return 'fallback (generator에서 집계 시트 직접 사용)'


if __name__ == '__main__':
    # 테스트
    print("=== 사용 가능한 수식 계산 방법 ===")
    methods = check_available_methods()
    for method, available in methods.items():
        status = "✓ 사용 가능" if available else "✗ 사용 불가"
        print(f"  {method}: {status}")
    
    print(f"\n권장 방법: {get_recommended_method()}")

