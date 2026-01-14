#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
비공표자료 시트 데이터를 분석표 집계표에 삽입하는 스크립트

기초자료 수집표의 "건설 (비공표자료)" 시트 데이터를
분석표의 "F'(건설)집계" 시트에 삽입하고 결측치를 보완합니다.
"""

import openpyxl
from openpyxl.cell.cell import MergedCell
from pathlib import Path
import sys


def fill_nonpublic_data(raw_excel_path: str, analysis_excel_path: str, output_path: str = None):
    """
    비공표자료 데이터를 집계표에 삽입
    
    Args:
        raw_excel_path: 기초자료 수집표 경로
        analysis_excel_path: 분석표 경로
        output_path: 출력 파일 경로 (None이면 분석표 파일 덮어쓰기)
    """
    if output_path is None:
        output_path = analysis_excel_path
    
    # 파일 열기
    print(f"[읽기] 기초자료 수집표: {raw_excel_path}")
    raw_wb = openpyxl.load_workbook(raw_excel_path, data_only=True)
    
    if '건설 (비공표자료)' not in raw_wb.sheetnames:
        print("[오류] '건설 (비공표자료)' 시트를 찾을 수 없습니다.")
        raw_wb.close()
        return False
    
    raw_ws = raw_wb['건설 (비공표자료)']
    
    print(f"[읽기] 분석표: {analysis_excel_path}")
    analysis_wb = openpyxl.load_workbook(analysis_excel_path, data_only=False)
    
    if "F'(건설)집계" not in analysis_wb.sheetnames:
        print("[오류] \"F'(건설)집계\" 시트를 찾을 수 없습니다.")
        raw_wb.close()
        analysis_wb.close()
        return False
    
    analysis_ws = analysis_wb["F'(건설)집계"]
    
    # 헤더 행 확인 (보통 3행)
    header_row = 3
    
    # 비공표자료에서 데이터가 있는 행 찾기 및 매핑 생성
    print("\n[1단계] 비공표자료 데이터 스캔 중...")
    raw_data_map = {}  # {(region_code, division_level, division_code): {row: raw_row_idx, data: {...}}}
    
    for row_idx in range(header_row + 1, raw_ws.max_row + 1):
        region_code = raw_ws.cell(row=row_idx, column=1).value
        region_name = raw_ws.cell(row=row_idx, column=2).value
        division_level = raw_ws.cell(row=row_idx, column=3).value
        division_code = raw_ws.cell(row=row_idx, column=4).value
        item_name = raw_ws.cell(row=row_idx, column=5).value
        
        # None 값 처리
        region_code = str(region_code) if region_code is not None else ''
        division_level = str(division_level) if division_level is not None else ''
        division_code = str(division_code) if division_code is not None else ''
        
        # 데이터가 있는지 확인 (연도 열부터 확인)
        # 건설 시트는 열 6부터 연도 데이터 (2020, 2021, 2022, 2023, 2024)
        # 그 다음 분기 데이터
        has_data = False
        data_values = {}
        
        # 연도 데이터 열 (6~10: 2020~2024)
        for col_offset in range(5):  # 0~4
            col = 6 + col_offset
            if col <= raw_ws.max_column:
                val = raw_ws.cell(row=row_idx, column=col).value
                if val is not None and val != '':
                    has_data = True
                    data_values[f'year_{col_offset}'] = val
        
        # 분기 데이터 열 확인 (11번째 열부터, 최대 13개 분기)
        # 실제로는 헤더를 확인해야 하지만, 여기서는 열 11부터 확인
        for col_offset in range(13):  # 0~12
            col = 11 + col_offset
            if col <= raw_ws.max_column:
                val = raw_ws.cell(row=row_idx, column=col).value
                if val is not None and val != '':
                    has_data = True
                    data_values[f'quarter_{col_offset}'] = val
        
        if has_data:
            key = (region_code, division_level, division_code)
            raw_data_map[key] = {
                'row': row_idx,
                'region_code': region_code,
                'region_name': region_name,
                'division_level': division_level,
                'division_code': division_code,
                'item_name': item_name,
                'data': data_values
            }
    
    print(f"  → 비공표자료에서 {len(raw_data_map)}개 행의 데이터 발견")
    
    # 집계표에서 매칭되는 행 찾기 및 데이터 삽입
    print("\n[2단계] 집계표에 데이터 삽입 중...")
    filled_count = 0
    updated_count = 0
    
    # 집계표의 열 구조 확인
    # 건설 집계표는 열 6부터 연도 데이터, 그 다음 분기 데이터
    # SHEET_STRUCTURE에 따르면:
    # 'F'(건설)집계': {'meta_start': 2, 'raw_meta_cols': 4, 'year_start': 6, 'quarter_start': 11, ...}
    year_start_col = 6  # 1-based
    quarter_start_col = 11  # 1-based
    
    for row_idx in range(header_row + 1, analysis_ws.max_row + 1):
        region_code_agg = analysis_ws.cell(row=row_idx, column=1).value
        division_level_agg = analysis_ws.cell(row=row_idx, column=3).value
        division_code_agg = analysis_ws.cell(row=row_idx, column=4).value
        
        # None 값 처리
        region_code_agg = str(region_code_agg) if region_code_agg is not None else ''
        division_level_agg = str(division_level_agg) if division_level_agg is not None else ''
        division_code_agg = str(division_code_agg) if division_code_agg is not None else ''
        
        key = (region_code_agg, division_level_agg, division_code_agg)
        
        if key in raw_data_map:
            raw_data = raw_data_map[key]
            row_updated = False
            
            # 공정 이름 복사 (열 5)
            item_name_col = 5
            item_name_cell = analysis_ws.cell(row=row_idx, column=item_name_col)
            if not isinstance(item_name_cell, MergedCell):
                existing_item_name = item_name_cell.value
                new_item_name = raw_data['item_name']
                # 공정 이름이 비어있거나 다르면 업데이트
                if existing_item_name is None or existing_item_name == '' or str(existing_item_name).strip() != str(new_item_name).strip():
                    item_name_cell.value = new_item_name
                    row_updated = True
            
            # 연도 데이터 복사 (열 6~10: 2020~2024)
            for col_offset in range(5):  # 0~4
                year_key = f'year_{col_offset}'
                if year_key in raw_data['data']:
                    target_col = year_start_col + col_offset
                    target_cell = analysis_ws.cell(row=row_idx, column=target_col)
                    
                    # 병합된 셀은 건너뛰기
                    if isinstance(target_cell, MergedCell):
                        continue
                    
                    # 기존 값이 비어있거나 0인 경우에만 채우기 (결측치 보완)
                    existing_value = target_cell.value
                    new_value = raw_data['data'][year_key]
                    
                    if existing_value is None or existing_value == '' or existing_value == 0:
                        target_cell.value = new_value
                        row_updated = True
                        filled_count += 1
                    elif existing_value != new_value:
                        # 값이 다르면 업데이트 (비공표자료가 더 정확할 수 있음)
                        target_cell.value = new_value
                        row_updated = True
                        updated_count += 1
            
            # 분기 데이터 복사 (열 11부터, 최대 13개 분기)
            for col_offset in range(13):  # 0~12
                quarter_key = f'quarter_{col_offset}'
                if quarter_key in raw_data['data']:
                    target_col = quarter_start_col + col_offset
                    if target_col > analysis_ws.max_column:
                        break
                    
                    target_cell = analysis_ws.cell(row=row_idx, column=target_col)
                    
                    # 병합된 셀은 건너뛰기
                    if isinstance(target_cell, MergedCell):
                        continue
                    
                    # 기존 값이 비어있거나 0인 경우에만 채우기 (결측치 보완)
                    existing_value = target_cell.value
                    new_value = raw_data['data'][quarter_key]
                    
                    if existing_value is None or existing_value == '' or existing_value == 0:
                        target_cell.value = new_value
                        row_updated = True
                        filled_count += 1
                    elif existing_value != new_value:
                        # 값이 다르면 업데이트
                        target_cell.value = new_value
                        row_updated = True
                        updated_count += 1
            
            if row_updated:
                print(f"  → 행 {row_idx} 업데이트: {raw_data['item_name']}")
    
    # 파일 저장
    print(f"\n[3단계] 파일 저장 중...")
    analysis_wb.save(output_path)
    
    raw_wb.close()
    analysis_wb.close()
    
    print(f"\n[완료] 비공표자료 데이터 삽입 완료")
    print(f"  → 새로 채운 셀: {filled_count}개")
    print(f"  → 업데이트된 셀: {updated_count}개")
    print(f"  → 출력 파일: {output_path}")
    
    return True


def main():
    """메인 함수"""
    import argparse
    
    parser = argparse.ArgumentParser(description='비공표자료 데이터를 집계표에 삽입')
    parser.add_argument('--raw', type=str, 
                       default='기초자료 수집표_2025년 3분기(캡스톤).xlsx',
                       help='기초자료 수집표 파일 경로')
    parser.add_argument('--analysis', type=str,
                       default='분석표_25년 3분기★_캡스톤.xlsx',
                       help='분석표 파일 경로')
    parser.add_argument('--output', type=str, default=None,
                       help='출력 파일 경로 (None이면 분석표 파일 덮어쓰기)')
    
    args = parser.parse_args()
    
    # 파일 존재 확인
    if not Path(args.raw).exists():
        print(f"[오류] 기초자료 수집표 파일을 찾을 수 없습니다: {args.raw}")
        return 1
    
    if not Path(args.analysis).exists():
        print(f"[오류] 분석표 파일을 찾을 수 없습니다: {args.analysis}")
        return 1
    
    # 실행
    success = fill_nonpublic_data(args.raw, args.analysis, args.output)
    
    return 0 if success else 1


if __name__ == '__main__':
    sys.exit(main())
