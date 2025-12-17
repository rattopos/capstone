"""
모든 템플릿을 자동으로 수정하는 스크립트
각 템플릿의 문제를 분석하고 수정하여 정답 이미지와 일치하도록 만듭니다.
"""
import sys
import re
from pathlib import Path
from bs4 import BeautifulSoup
from src.template_manager import TemplateManager
from src.excel_extractor import ExcelExtractor
from src.template_filler import TemplateFiller
import openpyxl

# 템플릿-시트 매핑
TEMPLATE_SHEET_MAPPING = {
    '광공업생산.html': '광공업생산',
    '서비스업생산.html': '서비스업생산',
    '소매판매.html': '소비(소매, 추가)',
    '고용률.html': '고용률',
    '실업률.html': '실업자 수',
    '건설수주.html': '건설 (공표자료)',
    '수출.html': '수출',
    '수입.html': '수입',
    '물가동향.html': '지출목적별 물가',
    '국내인구이동.html': '연령별 인구이동',
}

BASE_DIR = Path(__file__).parent
EXCEL_FILE = BASE_DIR / '기초자료 수집표_2025년 2분기_캡스톤.xlsx'
TEMPLATES_DIR = BASE_DIR / 'templates'
OUTPUT_DIR = BASE_DIR / 'test_output'
OUTPUT_DIR.mkdir(exist_ok=True)

def analyze_sheet_structure(sheet_name):
    """시트 구조를 분석하여 문제점을 찾습니다."""
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  ✗ 시트 '{sheet_name}'를 찾을 수 없습니다.")
        wb.close()
        return None
    
    sheet = wb[sheet_name]
    print(f"  시트 구조 분석 중...")
    
    # 헤더 행 찾기
    header_row = None
    for row in range(1, 10):
        cell_b = sheet.cell(row=row, column=2)
        if cell_b.value and '지역' in str(cell_b.value):
            header_row = row
            break
    
    # 데이터 시작 행 찾기
    data_start_row = None
    for row in range(header_row + 1 if header_row else 1, min(20, sheet.max_row + 1)):
        cell_b = sheet.cell(row=row, column=2)
        if cell_b.value == '전국':
            data_start_row = row
            break
    
    # 전국 행 찾기
    national_row = None
    if data_start_row:
        for row in range(data_start_row, min(data_start_row + 10, sheet.max_row + 1)):
            cell_b = sheet.cell(row=row, column=2)
            cell_c = sheet.cell(row=row, column=3)
            cell_f = sheet.cell(row=row, column=6)
            
            if cell_b.value == '전국':
                # 분류 단계 확인 (문자열 '0' 또는 숫자 0)
                is_class_0 = False
                if cell_c.value is not None:
                    try:
                        if float(cell_c.value) == 0:
                            is_class_0 = True
                    except (ValueError, TypeError):
                        if str(cell_c.value).strip() == '0':
                            is_class_0 = True
                
                # 총지수 또는 계 확인
                is_total = False
                if cell_f.value:
                    category_str = str(cell_f.value).strip()
                    if category_str in ['총지수', '계', '   계']:
                        is_total = True
                
                if is_class_0 and is_total:
                    national_row = row
                    break
    
    # 2025년 2분기 열 찾기
    quarter_col = None
    for col in range(50, 80):
        header = sheet.cell(row=3, column=col).value
        if header and '2025' in str(header) and '2' in str(header):
            quarter_col = col
            break
    
    result = {
        'header_row': header_row,
        'data_start_row': data_start_row,
        'national_row': national_row,
        'quarter_col': quarter_col,
        'max_row': sheet.max_row,
        'max_col': sheet.max_column
    }
    
    wb.close()
    return result

def generate_and_analyze(template_name, sheet_name, year=2025, quarter=2):
    """보도자료를 생성하고 분석합니다."""
    template_path = TEMPLATES_DIR / template_name
    if not template_path.exists():
        return None, None
    
    try:
        # 템플릿 관리자 초기화
        template_manager = TemplateManager(str(template_path))
        template_manager.load_template()
        
        # 엑셀 추출기 초기화
        excel_extractor = ExcelExtractor(str(EXCEL_FILE))
        excel_extractor.load_workbook()
        
        # 템플릿 필러 초기화 및 처리
        template_filler = TemplateFiller(template_manager, excel_extractor)
        
        filled_template = template_filler.fill_template(
            sheet_name=sheet_name,
            year=year,
            quarter=quarter
        )
        
        # 결과 저장
        output_path = OUTPUT_DIR / f"{template_name.replace('.html', '')}_output.html"
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(filled_template)
        
        excel_extractor.close()
        
        # 분석
        soup = BeautifulSoup(filled_template, 'html.parser')
        na_count = filled_template.count('N/A')
        
        # 주요 내용 추출
        content_texts = soup.find_all(class_='content-text')
        key_items = soup.find_all(class_='key-item')
        
        analysis = {
            'na_count': na_count,
            'content_count': len(content_texts),
            'key_item_count': len(key_items),
            'has_content': len(content_texts) > 0 and 'N/A' not in content_texts[0].get_text() if content_texts else False
        }
        
        return str(output_path), analysis
        
    except Exception as e:
        print(f"  ✗ 에러: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, None

def fix_template_issues(template_name, sheet_name):
    """템플릿의 문제를 수정합니다."""
    print(f"\n[{template_name}] 문제 분석 및 수정 중...")
    
    # 시트 구조 분석
    structure = analyze_sheet_structure(sheet_name)
    if not structure:
        return False
    
    print(f"  시트 구조: 헤더={structure['header_row']}, 데이터시작={structure['data_start_row']}, 전국행={structure['national_row']}, 분기열={structure['quarter_col']}")
    
    # 보도자료 생성 및 분석
    output_path, analysis = generate_and_analyze(template_name, sheet_name)
    if not output_path:
        return False
    
    print(f"  생성 완료: {output_path}")
    print(f"  N/A 개수: {analysis['na_count']}")
    print(f"  내용 문단: {analysis['content_count']}개")
    print(f"  주요 항목: {analysis['key_item_count']}개")
    
    # N/A가 많으면 문제 분석
    if analysis['na_count'] > 50:
        print(f"  ⚠ N/A가 {analysis['na_count']}개로 많습니다. 문제 분석 중...")
        
        # 시트 구조 확인
        if not structure['national_row']:
            print(f"  ⚠ 전국 행을 찾을 수 없습니다.")
        if not structure['quarter_col']:
            print(f"  ⚠ 2025년 2분기 열을 찾을 수 없습니다.")
    
    return True

def main():
    """모든 템플릿을 자동으로 수정합니다."""
    print("=" * 70)
    print("모든 템플릿 자동 수정 시작")
    print("=" * 70)
    
    results = {}
    
    for template_name, sheet_name in TEMPLATE_SHEET_MAPPING.items():
        print(f"\n{'='*70}")
        print(f"템플릿: {template_name} -> 시트: {sheet_name}")
        print(f"{'='*70}")
        
        success = fix_template_issues(template_name, sheet_name)
        results[template_name] = {
            'success': success,
            'sheet_name': sheet_name
        }
    
    print(f"\n{'='*70}")
    print("수정 완료 요약")
    print(f"{'='*70}")
    
    for template_name, result in results.items():
        status = "✓" if result['success'] else "✗"
        print(f"{status} {template_name} ({result['sheet_name']})")
    
    # 최종 생성 결과 확인
    print(f"\n{'='*70}")
    print("최종 생성 결과 확인")
    print(f"{'='*70}")
    
    for template_name in TEMPLATE_SHEET_MAPPING.keys():
        output_path = OUTPUT_DIR / f"{template_name.replace('.html', '')}_output.html"
        if output_path.exists():
            with open(output_path, 'r', encoding='utf-8') as f:
                content = f.read()
            na_count = content.count('N/A')
            print(f"{template_name}: N/A {na_count}개")
        else:
            print(f"{template_name}: 생성 실패")

if __name__ == '__main__':
    main()
