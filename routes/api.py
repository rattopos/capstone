# -*- coding: utf-8 -*-
"""
API 라우트
"""

import json
import base64
import re
from pathlib import Path
from urllib.parse import quote

from flask import Blueprint, request, jsonify, session, send_file, make_response
import unicodedata
import uuid

from config.settings import (
    BASE_DIR,
    TEMPLATES_DIR,
    UPLOAD_FOLDER,
    EXPORT_FOLDER,
    TEMP_DIR,
    TEMP_OUTPUT_DIR,
    TEMP_REGIONAL_OUTPUT_DIR,
    TEMP_CALCULATED_DIR
)
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


def safe_filename(filename):
    """한글을 보존하면서 안전한 파일명 생성
    
    - 한글, 영문, 숫자, 언더스코어, 하이픈, 점 허용
    - 위험한 문자 제거
    - 파일명 충돌 방지를 위해 UUID 추가
    """
    # 파일명과 확장자 분리
    if '.' in filename:
        name, ext = filename.rsplit('.', 1)
        ext = '.' + ext.lower()
    else:
        name = filename
        ext = ''
    
    # 유니코드 정규화
    name = unicodedata.normalize('NFC', name)
    
    # 허용할 문자만 유지 (한글, 영문, 숫자, 언더스코어, 하이픈, 공백)
    safe_chars = []
    for char in name:
        if char.isalnum() or char in ('_', '-', ' ', '년', '분기'):
            safe_chars.append(char)
        elif '\uAC00' <= char <= '\uD7A3':  # 한글 완성형
            safe_chars.append(char)
        elif '\u3131' <= char <= '\u3163':  # 한글 자모
            safe_chars.append(char)
    
    name = ''.join(safe_chars).strip()
    
    # 공백을 언더스코어로
    name = name.replace(' ', '_')
    
    # 빈 파일명 방지
    if not name:
        name = 'upload'
    
    # 파일명 충돌 방지를 위해 짧은 UUID 추가
    short_uuid = str(uuid.uuid4())[:8]
    
    return f"{name}_{short_uuid}{ext}"


def send_file_with_korean_filename(filepath, filename, mimetype):
    """한글 파일명을 지원하는 파일 다운로드 응답 생성 (RFC 5987)"""
    response = make_response(send_file(filepath, mimetype=mimetype))
    
    # RFC 5987 방식으로 한글 파일명 인코딩
    encoded_filename = quote(filename, safe='')
    
    # Content-Disposition 헤더 설정 (ASCII fallback + UTF-8 filename)
    ascii_filename = filename.encode('ascii', 'ignore').decode('ascii') or 'download'
    response.headers['Content-Disposition'] = (
        f"attachment; filename=\"{ascii_filename}\"; "
        f"filename*=UTF-8''{encoded_filename}"
    )
    
    return response
from config.reports import REPORT_ORDER, SECTOR_REPORTS, REGIONAL_REPORTS, SUMMARY_REPORTS, STATISTICS_REPORTS
from utils.excel_utils import extract_year_quarter_from_excel
from services.report_generator import (
    generate_report_html,
    generate_regional_report_html,
    generate_statistics_report_html,
    generate_individual_statistics_html
)
from services.excel_processor import preprocess_excel, check_available_methods, get_recommended_method
from services.excel_cache import set_cached_calculated_path
# from data_converter import DataConverter  # 레거시 모듈 - 더 이상 사용하지 않음
import openpyxl

api_bp = Blueprint('api', __name__, url_prefix='/api')


def _resolve_year_quarter(excel_path: str, year=None, quarter=None):
    """연도/분기 해석 (하드코딩 없이 엑셀에서 추출)"""
    if year is not None and quarter is not None:
        return year, quarter, None
    if excel_path and Path(excel_path).exists():
        try:
            y, q = extract_year_quarter_from_excel(excel_path)
            if y is not None and q is not None:
                return y, q, None
        except Exception as e:
            return None, None, f"연도/분기 추출 실패: {e}"
    return None, None, "연도/분기 정보가 없습니다"


def cleanup_upload_folder(keep_current_files=True, cleanup_excel_only=True):
    """업로드 폴더 정리 (현재 세션 파일 제외)
    
    Args:
        keep_current_files: True면 현재 세션에서 사용 중인 파일은 보존
        cleanup_excel_only: True면 엑셀 파일만 정리 (HTML 등은 보존)
    """
    try:
        # 현재 세션에서 사용 중인 파일 목록
        protected_files = set()
        if keep_current_files:
            excel_path = session.get('excel_path')
            # 기초자료 수집표는 사용하지 않으므로 보호 목록에서 제외
            
            if excel_path:
                protected_files.add(Path(excel_path).name)
        
        # 업로드 폴더의 모든 파일 확인
        deleted_count = 0
        for file_path in UPLOAD_FOLDER.glob('*'):
            if file_path.is_file():
                # 정리 대상인지 확인
                should_delete = False
                
                # keep_current_files=False면 보호 목록 체크 없이 바로 삭제 대상
                if not keep_current_files:
                    # 엑셀 파일만 정리하는 경우
                    if cleanup_excel_only:
                        if file_path.suffix.lower() in ['.xlsx', '.xls']:
                            should_delete = True
                    else:
                        # 디버그 파일이 아닌 경우 모두 삭제
                        if '디버그' not in file_path.name:
                            should_delete = True
                # 현재 세션 파일이 아닌 경우
                elif file_path.name not in protected_files:
                    # 엑셀 파일만 정리하는 경우
                    if cleanup_excel_only:
                        if file_path.suffix.lower() in ['.xlsx', '.xls']:
                            should_delete = True
                    else:
                        # 디버그 파일이 아닌 경우 모두 삭제
                        if '디버그' not in file_path.name:
                            should_delete = True
                
                if should_delete:
                    try:
                        file_path.unlink()
                        deleted_count += 1
                        print(f"[정리] 파일 삭제: {file_path.name}")
                    except Exception as e:
                        print(f"[경고] 파일 삭제 실패 ({file_path.name}): {e}")
        
        if deleted_count > 0:
            print(f"[정리] 업로드 폴더 정리 완료: {deleted_count}개 파일 삭제")
        
        return deleted_count
        
    except Exception as e:
        print(f"[경고] 업로드 폴더 정리 중 오류: {e}")
        return 0


def cleanup_temp_artifacts(excel_path: str | None = None) -> None:
    """임시 파일 폴더 정리 (calculated/output 등)"""
    try:
        import shutil
        from services.excel_cache import clear_excel_cache

        # 임시 파일 삭제 로직 비활성화
        # for temp_dir in (TEMP_OUTPUT_DIR, TEMP_REGIONAL_OUTPUT_DIR, TEMP_CALCULATED_DIR):
        #     try:
        #         if temp_dir.exists():
        #             shutil.rmtree(temp_dir)
        #     except Exception as e:
        #         print(f"[경고] 임시 폴더 삭제 실패 ({temp_dir}): {e}")

        # 상위 TEMP_DIR가 비어있으면 정리 (비활성화)
        # try:
        #     if TEMP_DIR.exists() and not any(TEMP_DIR.iterdir()):
        #         TEMP_DIR.rmdir()
        # except Exception:
        #     pass

        # 캐시에서 계산 경로 제거만 수행 (파일 삭제는 안함)
        if excel_path:
            # clear_excel_cache는 이미 services/excel_cache.py에서 비활성화됨
            clear_excel_cache(excel_path, preserve_calculated_path=False)
    except Exception as e:
        print(f"[경고] 임시 파일 정리 중 오류: {e}")


def _calculate_analysis_sheets(excel_path: str, preserve_formulas: bool = True):
    """분석 시트의 수식 계산 및 검증 (수식 보존 옵션)
    
    분석 시트의 수식은 집계 시트를 참조합니다.
    수식을 유지하면서 계산 결과를 로그로 출력합니다.
    
    Args:
        excel_path: 분석표 파일 경로
        preserve_formulas: True면 수식 유지 (엑셀에서 계산), False면 값으로 대체
    """
    # 분석 시트 → 집계 시트 매핑
    analysis_aggregate_mapping = {
        'A 분석': 'A(광공업생산)집계',
        'B 분석': 'B(서비스업생산)집계',
        'C 분석': 'C(소비)집계',
        'D(고용률)분석': 'D(고용률)집계',
        'D(실업)분석': 'D(실업)집계',
        'E(지출목적물가) 분석': 'E(지출목적물가)집계',
        'E(품목성질물가)분석': 'E(품목성질물가)집계',
        "F'분석": "F'(건설)집계",
        'G 분석': 'G(수출)집계',
        'H 분석': 'H(수입)집계',
    }
    
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    calculated_count = 0
    formula_count = 0
    
    for analysis_sheet, aggregate_sheet in analysis_aggregate_mapping.items():
        if analysis_sheet not in wb.sheetnames:
            continue
        if aggregate_sheet not in wb.sheetnames:
            continue
        
        ws_analysis = wb[analysis_sheet]
        ws_aggregate = wb[aggregate_sheet]
        
        # 집계 시트를 dict로 캐싱 (빠른 조회용)
        aggregate_data = {}
        for row in ws_aggregate.iter_rows(min_row=1, max_row=ws_aggregate.max_row):
            for cell in row:
                if cell.value is not None:
                    aggregate_data[(cell.row, cell.column)] = cell.value
        
        # 분석 시트의 수식 셀 처리
        for row in ws_analysis.iter_rows(min_row=1, max_row=ws_analysis.max_row):
            for cell in row:
                if cell.value is None:
                    continue
                    
                val = str(cell.value)
                
                # 수식인 경우 (=로 시작)
                if val.startswith('='):
                    formula_count += 1
                    
                    # 집계 시트 참조 파싱: ='시트이름'!셀주소
                    import re
                    match = re.match(r"^='?([^'!]+)'?!([A-Z]+)(\d+)$", val)
                    if match:
                        ref_sheet = match.group(1)
                        ref_col_letter = match.group(2)
                        ref_row = int(match.group(3))
                        
                        # 열 문자를 숫자로 변환 (A=1, B=2, ...)
                        ref_col = 0
                        for i, c in enumerate(reversed(ref_col_letter)):
                            ref_col += (ord(c) - ord('A') + 1) * (26 ** i)
                        
                        # 집계 시트에서 값 가져오기
                        ref_value = aggregate_data.get((ref_row, ref_col))
                        if ref_value is not None:
                            calculated_count += 1
                            
                            if preserve_formulas:
                                # 수식 유지 (엑셀에서 열면 자동 계산)
                                # 계산 결과는 로그로만 출력
                                pass
                            else:
                                # 수식을 계산된 값으로 대체
                                cell.value = ref_value
    
    wb.save(excel_path)
    wb.close()
    
    if preserve_formulas:
        print(f"[분석표] 수식 보존 완료: {excel_path}")
        print(f"  → 총 {formula_count}개 수식 유지, {calculated_count}개 참조 값 확인")
        print(f"  → 엑셀에서 열면 수식이 자동으로 계산됩니다.")
    else:
        print(f"[분석표] 수식 계산 완료: {excel_path}")
        print(f"  → {calculated_count}개 셀 값으로 변환")


@api_bp.route('/upload', methods=['POST'])
def upload_excel():
    """분석표 파일 업로드
    
    분석표 → 지역경제동향 생성
    """
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '파일이 없습니다'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '파일이 선택되지 않았습니다'})
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': '엑셀 파일만 업로드 가능합니다'})
    
    # 새 파일 업로드 전 이전 파일 정리 (모든 이전 파일 삭제)
    # 현재 세션 파일도 포함하여 모두 정리 (새 파일로 교체하므로)
    old_excel_path = session.get('excel_path')
    cleanup_upload_folder(keep_current_files=False, cleanup_excel_only=True)
    
    # 한글 파일명 보존하면서 안전한 파일명 생성
    filename = safe_filename(file.filename)
    filepath = Path(UPLOAD_FOLDER) / filename
    file.save(str(filepath))
    
    # 저장된 파일 크기 확인 (데이터 유실 방지)
    saved_size = filepath.stat().st_size
    print(f"[업로드] 분석표 파일 저장 완료: {filename} ({saved_size:,} bytes)")
    
        # ===== 분석표 업로드 → 지역경제동향 생성 =====
    try:
        print(f"\n{'='*50}")
        print(f"[업로드] 분석표 업로드: {filename}")
        print(f"{'='*50}")
        
        # 수식 계산 전처리 (분석 시트의 수식을 계산)
        print(f"[전처리] 엑셀 수식 계산 시작...")
        processed_path, preprocess_success, preprocess_msg = preprocess_excel(str(filepath))
        
        if preprocess_success:
            print(f"[전처리] 성공: {preprocess_msg}")
            # 전처리된 파일 경로 사용
            filepath = Path(processed_path)
            # 전처리된 결과를 전역 캐시에 등록 (분석 시트 재계산 방지)
            set_cached_calculated_path(str(filepath), str(filepath))
        else:
            print(f"[전처리] {preprocess_msg} - generator fallback 로직 사용")
        
        # 연도/분기 추출 (선택적 - 실패해도 계속 진행, 실제 데이터 처리 시 추출)
        year, quarter = None, None
        try:
            year, quarter = extract_year_quarter_from_excel(str(filepath))
            print(f"[업로드] 연도/분기 추출 성공: {year}년 {quarter}분기")
        except (ValueError, Exception) as e:
            # 연도/분기 추출 실패해도 계속 진행 (보도자료 생성 시 데이터에서 추출)
            print(f"[업로드] 연도/분기 추출 실패 (계속 진행): {str(e)}")
            print(f"[업로드] 보도자료 생성 시 데이터에서 연도/분기를 추출합니다.")
        
        # 세션에 저장 (파일 수정 시간 포함)
        session['excel_path'] = str(filepath)
        session['year'] = year
        session['quarter'] = quarter
        session['file_type'] = 'analysis'
        try:
            session['excel_file_mtime'] = Path(filepath).stat().st_mtime
        except OSError:
            pass  # 파일 시간 확인 실패는 무시

        # 업로드 직후 미리보기 없이 자동 생성/내보내기
        auto_year, auto_quarter, auto_err = _resolve_year_quarter(str(filepath), year, quarter)
        if auto_year is not None and auto_quarter is not None:
            auto_generate = _generate_all_reports_core(auto_year, auto_quarter, cleanup_after=False)
            auto_export = _export_hwp_ready_core([], auto_year, auto_quarter, output_folder=EXPORT_FOLDER)
            cleanup_temp_artifacts(str(filepath))
        else:
            auto_generate = {'success': False, 'error': auto_err or '연도/분기 정보 없음', 'generated': [], 'errors': []}
            auto_export = {'success': False, 'error': auto_err or '연도/분기 정보 없음'}
        
        return jsonify({
            'success': True,
            'filename': filename,
            'file_type': 'analysis',
            'year': year,
            'quarter': quarter,
            'reports': REPORT_ORDER,
            'regional_reports': REGIONAL_REPORTS,
            'conversion_info': None,
            'preprocessing': {
                'success': preprocess_success,
                'message': preprocess_msg,
                'method': get_recommended_method()
            },
            'auto_generate': auto_generate,
            'auto_export': auto_export
        })
    
    except Exception as e:
        import traceback
        print(f"[오류] 분석표 처리 실패: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'분석표 처리 중 오류가 발생했습니다: {str(e)}'
        })


# 레거시 엔드포인트 - data_converter 모듈이 제거되어 비활성화됨
# @api_bp.route('/download-analysis', methods=['GET'])
# def download_analysis():
#     """분석표 다운로드 (다운로드 시점에 생성 + 수식 계산)"""
#     import time
#     import zipfile
#     
#     raw_excel_path = session.get('raw_excel_path')
#     
#     if not raw_excel_path or not Path(raw_excel_path).exists():
#         return jsonify({'success': False, 'error': '기초자료 파일을 찾을 수 없습니다. 먼저 기초자료를 업로드해주세요.'}), 404
#     
#     try:
#         converter = DataConverter(str(raw_excel_path))
#         analysis_output = str(UPLOAD_FOLDER / f"분석표_{converter.year}년_{converter.quarter}분기_자동생성.xlsx")
#         
#         # 이미 유효한 분석표가 있는지 확인 (세션에서 생성된 파일)
#         download_path = session.get('download_analysis_path')
#         raw_file_mtime = session.get('raw_file_mtime')  # 원본 파일 수정 시간
#         need_regenerate = True
#         
#         if download_path and Path(download_path).exists():
#             # 원본 파일이 변경되었는지 확인
#             current_raw_mtime = Path(raw_excel_path).stat().st_mtime if Path(raw_excel_path).exists() else None
#             file_changed = (raw_file_mtime is None or current_raw_mtime is None or 
#                           abs(current_raw_mtime - raw_file_mtime) > 1.0)  # 1초 이상 차이
#             
#             if file_changed:
#                 print(f"[다운로드] 원본 파일이 변경되었습니다, 재생성 필요")
#                 need_regenerate = True
#             else:
#                 # 기존 파일 유효성 검사
#                 try:
#                     with zipfile.ZipFile(download_path, 'r') as zf:
#                         # zip 파일이 유효한지 테스트
#                         if zf.testzip() is None:
#                             need_regenerate = False
#                             analysis_output = download_path
#                             print(f"[다운로드] 기존 분석표 재사용: {download_path}")
#                 except (zipfile.BadZipFile, EOFError):
#                     print(f"[다운로드] 기존 파일 손상됨, 재생성 필요")
#                     need_regenerate = True
#         
#         if need_regenerate:
#             # 분석표 생성
#             analysis_path = converter.convert_all(analysis_output, weight_settings=None)
#             
#             # 파일 저장 완료 대기 (파일 시스템 동기화)
#             time.sleep(0.3)
#             
#             # 분석 시트 수식 계산 (집계 시트 값을 분석 시트로 복사)
#             _calculate_analysis_sheets(analysis_path)
#             
#             # 세션에 저장 (원본 파일 수정 시간 포함)
#             session['download_analysis_path'] = analysis_path
#             try:
#                 session['raw_file_mtime'] = Path(raw_excel_path).stat().st_mtime
#             except OSError:
#                 pass  # 파일 시간 확인 실패는 무시
#         else:
#             analysis_path = analysis_output
#         
#         filename = Path(analysis_path).name
#         
#         return send_file_with_korean_filename(
#             analysis_path,
#             filename,
#             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#     except Exception as e:
#         import traceback
#         traceback.print_exc()
#         return jsonify({'success': False, 'error': f'분석표 생성 실패: {str(e)}'}), 500


# 레거시 엔드포인트 - data_converter 모듈이 제거되어 비활성화됨
# @api_bp.route('/generate-analysis-with-weights', methods=['POST'])
# def generate_analysis_with_weights():
#     """분석표 생성 + 다운로드 (가중치 기본값 제거, 결측치는 N/A로 표시)"""
#     import time
#     
#     data = request.get_json()
#     # weight_settings는 더 이상 사용하지 않음 (기본값 없이 결측치는 N/A로 표시)
#     
#     raw_excel_path = session.get('raw_excel_path')
#     if not raw_excel_path or not Path(raw_excel_path).exists():
#         return jsonify({'success': False, 'error': '기초자료 파일을 찾을 수 없습니다.'}), 404
#     
#     try:
#         converter = DataConverter(str(raw_excel_path))
#         
#         # 분석표 생성 (가중치 기본값 없이, 결측치는 N/A로 표시)
#         analysis_output = str(UPLOAD_FOLDER / f"분석표_{converter.year}년_{converter.quarter}분기_자동생성.xlsx")
#         analysis_path = converter.convert_all(analysis_output, weight_settings=None)
#         
#         # 파일 저장 완료 대기 (파일 시스템 동기화)
#         time.sleep(0.3)
#         
#         # 분석 시트 수식 계산 (집계 시트 값을 분석 시트로 복사)
#         _calculate_analysis_sheets(analysis_path)
#         
#         # 파일 무결성 확인
#         import zipfile
#         try:
#             with zipfile.ZipFile(analysis_path, 'r') as zf:
#                 if zf.testzip() is not None:
#                     raise Exception("생성된 파일이 손상되었습니다.")
#         except zipfile.BadZipFile:
#             raise Exception("생성된 파일이 손상되었습니다. 다시 시도해주세요.")
#         
#         session['download_analysis_path'] = analysis_path
#         
#         return jsonify({
#             'success': True,
#             'filename': Path(analysis_path).name,
#             'message': '분석표가 성공적으로 생성되었습니다.'
#         })
#     except Exception as e:
#         import traceback
#         traceback.print_exc()
#         return jsonify({'success': False, 'error': f'분석표 생성 실패: {str(e)}'}), 500


@api_bp.route('/report-order', methods=['GET'])
def get_report_order():
    """현재 보도자료 순서 반환"""
    return jsonify({'reports': REPORT_ORDER, 'regional_reports': REGIONAL_REPORTS})


@api_bp.route('/report-order', methods=['POST'])
def update_report_order():
    """보도자료 순서 업데이트"""
    from config import reports as reports_module
    data = request.get_json(silent=True)
    if data is None:
        return jsonify({'success': False, 'error': 'JSON 형식의 요청 데이터가 필요합니다.'}), 400
    
    new_order = data.get('order', [])
    
    if new_order:
        order_map = {r['id']: idx for idx, r in enumerate(new_order)}
        reports_module.REPORT_ORDER = sorted(reports_module.REPORT_ORDER, key=lambda x: order_map.get(x['id'], 999))
    
    return jsonify({'success': True, 'reports': reports_module.REPORT_ORDER})


@api_bp.route('/session-info', methods=['GET'])
def get_session_info():
    """현재 세션 정보 반환 (파일 존재 여부 확인)"""
    excel_path = session.get('excel_path')
    
    # 파일이 실제로 존재하는지 확인
    has_file = False
    if excel_path:
        file_path = Path(excel_path)
        if file_path.exists() and file_path.is_file():
            has_file = True
        else:
            # 파일이 존재하지 않으면 세션에서도 제거
            session.pop('excel_path', None)
            session.pop('year', None)
            session.pop('quarter', None)
            session.pop('file_type', None)
            excel_path = None
    
    return jsonify({
        'excel_path': excel_path,
        'year': session.get('year') if has_file else None,
        'quarter': session.get('quarter') if has_file else None,
        'has_file': has_file
    })


def _generate_all_reports_core(year, quarter, cleanup_after=True):
    """모든 보도자료 생성 공통 로직 (옵션: 업로드 정리 여부)"""
    from services.excel_cache import get_excel_file, clear_excel_cache

    excel_path = session.get('excel_path')
    if not excel_path or not Path(excel_path).exists():
        return {'success': False, 'error': '엑셀 파일을 먼저 업로드하세요', 'generated': [], 'errors': [], 'cleanup': cleanup_after}

    generated_reports = []
    errors = []
    excel_file = None
    result_missing = False
    temp_cleaned = False

    try:
        excel_file = get_excel_file(excel_path, use_data_only=True)
        if excel_file is None:
            error_msg = f"엑셀 파일을 로드할 수 없습니다: {excel_path}"
            print(f"[ERROR] {error_msg}")
            return {
                'success': False,
                'error': error_msg,
                'generated': [],
                'errors': [{'report_id': 'all', 'report_name': '전체', 'error': error_msg}],
                'cleanup': cleanup_after
            }
        print(f"[보도자료 생성] 엑셀 파일 캐싱 완료: {excel_path}")
    except Exception as e:
        import traceback
        error_msg = f"엑셀 파일 로드 실패: {str(e)}"
        print(f"[ERROR] {error_msg}")
        traceback.print_exc()
        return {
            'success': False,
            'error': error_msg,
            'generated': [],
            'errors': [{'report_id': 'all', 'report_name': '전체', 'error': error_msg}],
            'cleanup': cleanup_after
        }

    try:
        for report_config in SECTOR_REPORTS:
            try:
                report_name = report_config.get('name', report_config.get('id', 'Unknown'))
                report_id = report_config.get('id', 'Unknown')

                print(f"[보도자료 생성] 시작: {report_name} ({report_id})")

                html_content, error, _ = generate_report_html(
                    excel_path, report_config, year, quarter, None, excel_file=excel_file
                )

                if error:
                    import traceback
                    error_msg = f"{report_name} 생성 실패: {error}"
                    print(f"[ERROR] {error_msg}")
                    traceback.print_exc()
                    errors.append({'report_id': report_id, 'report_name': report_name, 'error': str(error)})
                elif html_content is None:
                    error_msg = f"{report_name} 생성 실패: HTML 내용이 None입니다"
                    print(f"[ERROR] {error_msg}")
                    errors.append({'report_id': report_id, 'report_name': report_name, 'error': 'HTML 내용이 None입니다'})
                else:
                    try:
                        report_name_safe = report_config.get('name', 'unknown')
                        if not report_name_safe or not isinstance(report_name_safe, str):
                            report_name_safe = 'unknown'
                        report_name_safe = report_name_safe.replace('/', '_').replace('\\', '_').replace('..', '_')
                        TEMP_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
                        output_path = TEMP_OUTPUT_DIR / f"{report_name_safe}_output.html"
                        output_path.parent.mkdir(parents=True, exist_ok=True)
                        with open(output_path, 'w', encoding='utf-8') as f:
                            f.write(html_content if html_content else '<!-- Empty content -->')
                        print(f"[보도자료 생성] 성공: {report_name} → {output_path}")
                        generated_reports.append({'report_id': report_id, 'name': report_name, 'path': str(output_path)})
                    except Exception as write_error:
                        import traceback
                        error_msg = f"{report_name} 파일 저장 실패: {str(write_error)}"
                        print(f"[ERROR] {error_msg}")
                        traceback.print_exc()
                        errors.append({'report_id': report_id, 'report_name': report_name, 'error': f"파일 저장 실패: {str(write_error)}"})
            except Exception as e:
                import traceback
                error_message = str(e)
                print(f"[ERROR] {report_config.get('name', report_config.get('id', 'Unknown'))} 생성 중 예외 발생: {error_message}")
                traceback.print_exc()
                errors.append({'report_id': report_config.get('id', 'Unknown'), 'report_name': report_config.get('name', report_config.get('id', 'Unknown')), 'error': f"예외 발생: {error_message}"})
                continue

        print(f"[보도자료 생성] 시도별 보도자료 생성 시작...")
        TEMP_REGIONAL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_dir = TEMP_REGIONAL_OUTPUT_DIR

        for region_config in REGIONAL_REPORTS:
            try:
                region_name = region_config.get('name', region_config.get('id', 'Unknown'))
                region_id = region_config.get('id', 'Unknown')
                print(f"[시도별 보도자료 생성] 시작: {region_name} ({region_id})")
                html_content, error = generate_regional_report_html(excel_path, region_name, is_reference=False, year=year, quarter=quarter, excel_file=excel_file)

                if error:
                    import traceback
                    error_msg = f"{region_name} 생성 실패: {error}"
                    print(f"[ERROR] {error_msg}")
                    traceback.print_exc()
                    errors.append({'report_id': region_id, 'report_name': f'시도별-{region_name}', 'error': str(error)})
                elif html_content is None:
                    error_msg = f"{region_name} 생성 실패: HTML 내용이 None입니다"
                    print(f"[ERROR] {error_msg}")
                    errors.append({'report_id': region_id, 'report_name': f'시도별-{region_name}', 'error': 'HTML 내용이 None입니다'})
                else:
                    try:
                        region_name_safe = region_name.replace('/', '_').replace('\\', '_').replace('..', '_')
                        output_path = output_dir / f"{region_name_safe}_output.html"
                        output_path.parent.mkdir(parents=True, exist_ok=True)
                        with open(output_path, 'w', encoding='utf-8') as f:
                            f.write(html_content if html_content else '<!-- Empty content -->')
                        print(f"[시도별 보도자료 생성] 성공: {region_name} → {output_path}")
                        generated_reports.append({'report_id': region_id, 'name': f'시도별-{region_name}', 'path': str(output_path)})
                    except Exception as write_error:
                        import traceback
                        error_msg = f"{region_name} 파일 저장 실패: {str(write_error)}"
                        print(f"[ERROR] {error_msg}")
                        traceback.print_exc()
                        errors.append({'report_id': region_id, 'report_name': f'시도별-{region_name}', 'error': f"파일 저장 실패: {str(write_error)}"})
            except Exception as e:
                import traceback
                error_message = str(e)
                print(f"[ERROR] {region_config.get('name', region_config.get('id', 'Unknown'))} 생성 중 예외 발생: {error_message}")
                traceback.print_exc()
                errors.append({'report_id': region_config.get('id', 'Unknown'), 'report_name': f"시도별-{region_config.get('name', region_config.get('id', 'Unknown'))}", 'error': f"예외 발생: {error_message}"})
                continue

        print(f"[보도자료 생성] 요약 보도자료 생성 시작...")
        for report_config in SUMMARY_REPORTS:
            try:
                report_name = report_config.get('name', report_config.get('id', 'Unknown'))
                report_id = report_config.get('id', 'Unknown')

                print(f"[보도자료 생성] 시작: {report_name} ({report_id})")

                html_content, error, _ = generate_report_html(
                    excel_path, report_config, year, quarter, None, excel_file=excel_file
                )

                if error:
                    import traceback
                    error_msg = f"{report_name} 생성 실패: {error}"
                    print(f"[ERROR] {error_msg}")
                    traceback.print_exc()
                    errors.append({'report_id': report_id, 'report_name': report_name, 'error': str(error)})
                elif html_content is None:
                    error_msg = f"{report_name} 생성 실패: HTML 내용이 None입니다"
                    print(f"[ERROR] {error_msg}")
                    errors.append({'report_id': report_id, 'report_name': report_name, 'error': 'HTML 내용이 None입니다'})
                else:
                    try:
                        report_name_safe = report_config.get('name', 'unknown')
                        if not report_name_safe or not isinstance(report_name_safe, str):
                            report_name_safe = 'unknown'
                        report_name_safe = report_name_safe.replace('/', '_').replace('\\', '_').replace('..', '_')
                        TEMP_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
                        output_path = TEMP_OUTPUT_DIR / f"{report_name_safe}_output.html"
                        output_path.parent.mkdir(parents=True, exist_ok=True)
                        with open(output_path, 'w', encoding='utf-8') as f:
                            f.write(html_content if html_content else '<!-- Empty content -->')
                        print(f"[보도자료 생성] 성공: {report_name} → {output_path}")
                        generated_reports.append({'report_id': report_id, 'name': report_name, 'path': str(output_path)})
                    except Exception as write_error:
                        import traceback
                        error_msg = f"{report_name} 파일 저장 실패: {str(write_error)}"
                        print(f"[ERROR] {error_msg}")
                        traceback.print_exc()
                        errors.append({'report_id': report_id, 'report_name': report_name, 'error': f"파일 저장 실패: {str(write_error)}"})
            except Exception as e:
                import traceback
                error_message = str(e)
                print(f"[ERROR] {report_config.get('name', report_config.get('id', 'Unknown'))} 생성 중 예외 발생: {error_message}")
                traceback.print_exc()
                errors.append({'report_id': report_config.get('id', 'Unknown'), 'report_name': report_config.get('name', report_config.get('id', 'Unknown')), 'error': f"예외 발생: {error_message}"})
                continue
    finally:
        try:
            expected_count = len(SECTOR_REPORTS) + len(REGIONAL_REPORTS) + len(SUMMARY_REPORTS)
            result_missing = len(errors) > 0 or len(generated_reports) < expected_count
        except Exception:
            result_missing = True

        if result_missing:
            try:
                print("[정리] 결과 누락 감지 - 캐시 및 임시 파일 즉시 삭제")
                cleanup_temp_artifacts(excel_path)
                temp_cleaned = True
            except Exception as cleanup_error:
                print(f"[경고] 결과 누락 정리 중 오류 (무시): {cleanup_error}")
        else:
            clear_excel_cache(excel_path, preserve_calculated_path=not cleanup_after)

        if cleanup_after:
            try:
                if not temp_cleaned:
                    print(f"[정리] 작업 완료 - 임시 파일 정리 시작...")
                    cleanup_temp_artifacts(excel_path)
                    temp_cleaned = True
                print(f"[정리] 작업 완료 - 업로드 파일 정리 시작...")
                deleted_count = cleanup_upload_folder(keep_current_files=False, cleanup_excel_only=True)
                if deleted_count > 0:
                    print(f"[정리] 작업 완료 후 업로드 파일 {deleted_count}개 삭제 완료")
                session.pop('excel_path', None)
                session.pop('year', None)
                session.pop('quarter', None)
                session.pop('file_type', None)
            except Exception as cleanup_error:
                print(f"[경고] 업로드 파일 정리 중 오류 (무시): {cleanup_error}")

    return {'success': len(errors) == 0, 'generated': generated_reports, 'errors': errors, 'cleanup': cleanup_after}


@api_bp.route('/generate-all', methods=['POST'])
def generate_all_reports():
    """모든 보도자료 일괄 생성 (최적화 버전 - 엑셀 파일 캐싱)"""
    data = request.get_json(silent=True)
    if data is None:
        data = {}
    year = data.get('year', session.get('year'))
    quarter = data.get('quarter', session.get('quarter'))
    cleanup_after = data.get('cleanup_after', True)

    if year is None or quarter is None:
        excel_path = session.get('excel_path')
        year, quarter, resolve_err = _resolve_year_quarter(excel_path, year, quarter)
        if year is None or quarter is None:
            return jsonify({
                'success': False,
                'error': resolve_err or '연도/분기 정보가 없습니다',
                'generated': [],
                'errors': [{'report_id': 'all', 'report_name': '전체', 'error': resolve_err or '연도/분기 정보가 없습니다'}],
                'cleanup': cleanup_after
            })

    result = _generate_all_reports_core(year, quarter, cleanup_after=cleanup_after)
    return jsonify(result)


@api_bp.route('/generate-all-regional', methods=['POST'])
def generate_all_regional_reports():
    """시도별 보도자료 전체 생성"""
    excel_path = session.get('excel_path')
    if not excel_path or not Path(excel_path).exists():
        return jsonify({'success': False, 'error': '엑셀 파일을 먼저 업로드하세요'})
    
    generated_reports = []
    errors = []
    
    TEMP_REGIONAL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_dir = TEMP_REGIONAL_OUTPUT_DIR
    
    gen_year, gen_quarter, resolve_err = _resolve_year_quarter(excel_path, session.get('year'), session.get('quarter'))
    if gen_year is None or gen_quarter is None:
        return jsonify({'success': False, 'error': resolve_err or '연도/분기 정보가 없습니다'})

    for region_config in REGIONAL_REPORTS:
        html_content, error = generate_regional_report_html(excel_path, region_config['name'], is_reference=False, year=gen_year, quarter=gen_quarter)
        
        if error:
            errors.append({'region_id': region_config['id'], 'error': error})
        else:
            output_path = output_dir / f"{region_config['name']}_output.html"
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            generated_reports.append({
                'region_id': region_config['id'],
                'name': region_config['name'],
                'path': str(output_path)
            })
    
    # 작업 완료 후 업로드 파일 삭제
    try:
        print(f"[정리] 시도별 보도자료 생성 완료 - 임시 파일 정리 시작...")
        cleanup_temp_artifacts(excel_path)
        print(f"[정리] 시도별 보도자료 생성 완료 - 업로드 파일 정리 시작...")
        deleted_count = cleanup_upload_folder(keep_current_files=False, cleanup_excel_only=True)
        if deleted_count > 0:
            print(f"[정리] 작업 완료 후 업로드 파일 {deleted_count}개 삭제 완료")
        # 세션에서도 파일 경로 제거
        session.pop('excel_path', None)
        session.pop('year', None)
        session.pop('quarter', None)
        session.pop('file_type', None)
    except Exception as cleanup_error:
        print(f"[경고] 업로드 파일 정리 중 오류 (무시): {cleanup_error}")
    
    return jsonify({
        'success': len(errors) == 0,
        'generated': generated_reports,
        'errors': errors
    })


@api_bp.route('/export-final', methods=['POST'])
def export_final_document():
    """모든 보도자료를 HTML 문서로 합치기 (standalone 옵션 지원)"""
    try:
        data = request.get_json(silent=True)
        if data is None:
            return jsonify({'success': False, 'error': 'JSON 형식의 요청 데이터가 필요합니다.'}), 400
        
        pages = data.get('pages', [])
        year = data.get('year', session.get('year'))
        quarter = data.get('quarter', session.get('quarter'))
        if year is None or quarter is None:
            excel_path = session.get('excel_path')
            year, quarter, resolve_err = _resolve_year_quarter(excel_path, year, quarter)
            if year is None or quarter is None:
                return jsonify({'success': False, 'error': resolve_err or '연도/분기 정보가 없습니다.'}), 400
        standalone = data.get('standalone', False)  # 완전한 standalone HTML 여부
        
        if not pages:
            return jsonify({'success': False, 'error': '페이지 데이터가 없습니다.'})
        
        # 모든 페이지의 스타일 수집
        all_styles = set()
        
        # standalone 모드: 외부 의존성 없는 완전한 HTML
        # 일반 모드: 외부 폰트 사용
        font_style = ''
        if not standalone:
            font_style = "@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;600;700&display=swap');"
        
        final_html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{year}년 {quarter}/4분기 지역경제동향</title>
    <style>
        {font_style}
        
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        html, body {{
            width: 210mm;
            background: white;
        }}
        
        body {{
            font-family: 'Noto Sans KR', '맑은 고딕', 'Malgun Gothic', '나눔고딕', 'NanumGothic', sans-serif;
        }}
        
        /* PDF 출력용 페이지 스타일 */
        .pdf-page {{
            width: 210mm;
            min-height: 297mm;
            max-height: 297mm;
            padding: 12mm 15mm 15mm 15mm;
            margin: 0 auto 5mm auto;
            background: white;
            position: relative;
            overflow: hidden;
            page-break-after: always;
            page-break-inside: avoid;
        }}
        
        .pdf-page:last-child {{
            page-break-after: auto;
            margin-bottom: 0;
        }}
        
        .pdf-page-content {{
            width: 100%;
            height: calc(297mm - 32mm);
            overflow: hidden;
        }}
        
        .pdf-page-content > * {{
            max-width: 100%;
        }}
        
        /* 페이지 번호 */
        .pdf-page-number {{
            position: absolute;
            bottom: 8mm;
            left: 0;
            right: 0;
            text-align: center;
            font-size: 9pt;
            color: #333;
        }}
        
        /* 변환된 차트 이미지 스타일 */
        img[data-converted-from="canvas"] {{
            max-width: 100%;
            height: auto;
            display: block;
        }}
        
        /* 화면 미리보기용 */
        @media screen {{
            body {{
                background: #f0f0f0;
                padding: 20px;
            }}
            
            .pdf-page {{
                box-shadow: 0 2px 10px rgba(0,0,0,0.15);
                border: 1px solid #ddd;
            }}
        }}
        
        /* 인쇄/PDF 저장용 */
        @media print {{
            html, body {{
                width: 210mm;
                background: white !important;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }}
            
            body {{
                padding: 0;
                margin: 0;
            }}
            
            .pdf-page {{
                width: 210mm;
                height: 297mm;
                min-height: 297mm;
                max-height: 297mm;
                padding: 12mm 15mm 15mm 15mm;
                margin: 0;
                box-shadow: none;
                border: none;
                page-break-after: always;
                page-break-inside: avoid;
            }}
            
            .pdf-page:last-child {{
                page-break-after: auto;
            }}
            
            /* 이미지 색상 유지 */
            img {{
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }}
        }}
        
        @page {{
            size: A4 portrait;
            margin: 0;
        }}
        
        /* 표 스타일 공통 */
        table {{
            border-collapse: collapse;
            width: 100%;
        }}
        
        th, td {{
            border: 1px solid #333;
            padding: 4px 6px;
            font-size: 9pt;
            text-align: center;
        }}
        
        th {{
            background: #f5f5f5;
            font-weight: 600;
        }}
        
        /* 차트 이미지 크기 조정 */
        .chart-container, .chart-wrapper {{
            max-width: 100%;
            overflow: hidden;
        }}
        
        .chart-container img, .chart-wrapper img {{
            max-width: 100%;
            height: auto;
        }}
        
        /* A4 최적화 - 이미지 리사이징 */
        img {{
            max-width: 100%;
            height: auto;
            object-fit: contain;
        }}
        
        /* A4 최적화 - 표 리사이징 */
        table {{
            width: 100%;
            table-layout: fixed;
            font-size: 8pt;
            word-wrap: break-word;
        }}
        
        th, td {{
            overflow: hidden;
            text-overflow: ellipsis;
            word-break: keep-all;
        }}
        
        /* A4 최적화 - 콘텐츠 영역 */
        .pdf-page-content {{
            width: 180mm;
            max-width: 180mm;
        }}
        
        .pdf-page-content img {{
            max-width: 180mm;
            max-height: 240mm;
        }}
        
        .pdf-page-content table {{
            max-width: 180mm;
        }}
        
        /* A4 최적화 - SVG 리사이징 */
        svg {{
            max-width: 100%;
            height: auto;
        }}
        
        /* 큰 요소 overflow 방지 */
        .pdf-page-content > * {{
            max-width: 100%;
            overflow: hidden;
        }}
        
        /* 그래프/차트 컨테이너 */
        .chart-area, .graph-container {{
            max-width: 100%;
            overflow: hidden;
        }}
    </style>
'''
        
        # 각 페이지에서 스타일 추출하여 추가
        for idx, page in enumerate(pages):
            page_html = page.get('html', '')
            if '<style' in page_html:
                style_matches = re.findall(r'<style[^>]*>(.*?)</style>', page_html, re.DOTALL)
                for style in style_matches:
                    # 중복 방지를 위해 hash 사용
                    style_hash = hash(style.strip())
                    if style_hash not in all_styles:
                        all_styles.add(style_hash)
                        # standalone 모드에서는 외부 폰트 import 제거
                        if standalone:
                            style = re.sub(r'@import\s+url\([^)]+\);?', '', style)
                        final_html += f'    <style>/* Page {idx+1} styles */\n{style}\n    </style>\n'
        
        final_html += '''</head>
<body>
'''
        
        for idx, page in enumerate(pages, 1):
            page_html = page.get('html', '')
            page_title = page.get('title', f'페이지 {idx}')
            padding: 0;
            margin: 0;
            if '<body' in page_html.lower():
                body_match = re.search(r'<body[^>]*>(.*?)</body>', page_html, re.DOTALL | re.IGNORECASE)
                if body_match:
                    body_content = body_match.group(1)
            
            # 내용에서 style, script 태그 제거 (이미 head에 추가됨)
            body_content = re.sub(r'<style[^>]*>.*?</style>', '', body_content, flags=re.DOTALL)
            
            # standalone 모드에서는 script 태그도 제거 (Chart.js 등 불필요)
            if standalone:
                body_content = re.sub(r'<script[^>]*>.*?</script>', '', body_content, flags=re.DOTALL)
            
            # 페이지 래퍼 추가
            final_html += f'''
    <!-- Page {idx}: {page_title} -->
    <div class="pdf-page" data-page="{idx}" data-title="{page_title}">
        <div class="pdf-page-content">
{body_content}
        </div>
        <div class="pdf-page-number">- {idx} -</div>
    </div>
'''
        
        # standalone 모드에서는 최소한의 스크립트만 추가
        if standalone:
            final_html += '''
    <script>
        // 인쇄 전 준비
        window.onbeforeprint = function() {
            document.body.style.background = 'white';
        };
        console.log('이 HTML 파일은 오프라인에서도 사용 가능합니다.');
        console.log('PDF 저장: Ctrl+P (또는 Cmd+P) → "PDF로 저장" 선택');
    </script>
'''
        else:
            final_html += '''
    <script>
        // 인쇄 전 준비
        window.onbeforeprint = function() {
            document.body.style.background = 'white';
        };
        
        // Ctrl+P로 PDF 저장 안내
        console.log('PDF 저장: Ctrl+P (또는 Cmd+P) → "PDF로 저장" 선택');
    </script>
'''
        
        final_html += '''</body>
</html>
'''
        
        # 파일명 설정
        if standalone:
            output_filename = f'지역경제동향_{year}년_{quarter}분기.html'
        else:
            output_filename = f'지역경제동향_{year}년_{quarter}분기_PDF용.html'
        
        output_path = UPLOAD_FOLDER / output_filename
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(final_html)
        
        return jsonify({
            'success': True,
            'html': final_html,
            'filename': output_filename,
            'download_url': f'/uploads/{output_filename}',
            'total_pages': len(pages),
            'standalone': standalone
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


@api_bp.route('/export-xlsx', methods=['POST'])
def export_xlsx_document():
    """모든 보도자료를 XLSX 파일로 내보내기"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.drawing.image import Image as XLImage
        from bs4 import BeautifulSoup
        import io
        import base64
        from PIL import Image as PILImage
        
        data = request.get_json(silent=True)
        if data is None:
            return jsonify({'success': False, 'error': 'JSON 형식의 요청 데이터가 필요합니다.'}), 400
        
        pages = data.get('pages', [])
        year = data.get('year', session.get('year'))
        quarter = data.get('quarter', session.get('quarter'))
        if year is None or quarter is None:
            excel_path = session.get('excel_path')
            year, quarter, resolve_err = _resolve_year_quarter(excel_path, year, quarter)
            if year is None or quarter is None:
                return jsonify({'success': False, 'error': resolve_err or '연도/분기 정보가 없습니다.'}), 400
        
        if not pages:
            return jsonify({'success': False, 'error': '페이지 데이터가 없습니다.'})
        
        # 워크북 생성
        wb = Workbook()
        wb.remove(wb.active)  # 기본 시트 제거
        
        # 스타일 정의
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        image_counter = 0
        
        for idx, page in enumerate(pages, 1):
            page_html = page.get('html', '')
            page_title = page.get('title', f'페이지{idx}')
            category = page.get('category', 'unknown')
            
            # 시트 이름 (최대 31자, 특수문자 제거)
            sheet_name = re.sub(r'[\\/*?:\[\]]', '', page_title)[:31]
            if sheet_name in [ws.title for ws in wb.worksheets]:
                sheet_name = f"{sheet_name[:28]}_{idx}"
            
            ws = wb.create_sheet(title=sheet_name)
            
            # HTML 파싱
            soup = BeautifulSoup(page_html, 'html.parser')
            
            current_row = 1
            
            # 제목 추가
            ws.cell(row=current_row, column=1, value=page_title)
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            current_row += 2
            
            # 표 추출 및 변환
            tables = soup.find_all('table')
            for table in tables:
                # 표 제목 찾기 (이전 요소에서)
                table_title = None
                prev = table.find_previous(['h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'p', 'div'])
                if prev and prev.get_text(strip=True):
                    table_title = prev.get_text(strip=True)[:100]
                
                if table_title:
                    ws.cell(row=current_row, column=1, value=table_title)
                    ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
                    current_row += 1
                
                # 표 데이터 추출
                rows = table.find_all('tr')
                for row_idx, tr in enumerate(rows):
                    cells = tr.find_all(['th', 'td'])
                    col_pointer = 1  # 현재 입력할 엑셀 컬럼 위치
                    for cell in cells:
                        col_idx = col_pointer
                        cell_text = cell.get_text(strip=True)
                        
                        # 숫자 변환 시도
                        try:
                            if '.' in cell_text:
                                cell_value = float(cell_text.replace(',', '').replace('%', ''))
                            elif cell_text.replace(',', '').replace('-', '').isdigit():
                                cell_value = int(cell_text.replace(',', ''))
                            else:
                                cell_value = cell_text
                        except:
                            cell_value = cell_text
                        
                        excel_cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                        excel_cell.border = thin_border
                        excel_cell.alignment = center_align
                        
                        # 헤더 스타일
                        if cell.name == 'th' or row_idx == 0:
                            excel_cell.font = header_font
                            excel_cell.fill = header_fill
                        
                        # colspan 처리: 병합 영역에는 값 입력을 반복하지 않도록 포인터만 이동
                        colspan = int(cell.get('colspan', 1))
                        if colspan > 1:
                            ws.merge_cells(
                                start_row=current_row, start_column=col_idx,
                                end_row=current_row, end_column=col_idx + colspan - 1
                            )
                            col_pointer += colspan
                        else:
                            col_pointer += 1
                    
                    current_row += 1
                
                current_row += 2  # 표 간 간격
            
            # 인포그래픽/이미지 처리 (base64 이미지)
            images = soup.find_all('img')
            for img in images:
                src = img.get('src', '')
                if src.startswith('data:image'):
                    try:
                        # base64 이미지 추출
                        match = re.match(r'data:image/([^;]+);base64,(.+)', src)
                        if match:
                            img_format = match.group(1)
                            img_data = base64.b64decode(match.group(2))
                            
                            # 이미지를 임시 파일로 저장
                            img_buffer = io.BytesIO(img_data)
                            pil_img = PILImage.open(img_buffer)
                            
                            # PNG로 변환
                            png_buffer = io.BytesIO()
                            pil_img.save(png_buffer, format='PNG')
                            png_buffer.seek(0)
                            
                            # 엑셀에 이미지 추가
                            xl_img = XLImage(png_buffer)
                            
                            # 이미지 크기 조정 (최대 너비 500px)
                            max_width = 500
                            if xl_img.width > max_width:
                                ratio = max_width / xl_img.width
                                xl_img.width = max_width
                                xl_img.height = int(xl_img.height * ratio)
                            
                            ws.add_image(xl_img, f'A{current_row}')
                            image_counter += 1
                            
                            # 이미지 높이만큼 행 건너뛰기
                            rows_needed = max(1, xl_img.height // 20)
                            current_row += rows_needed + 2
                    except Exception as e:
                        print(f"이미지 처리 오류: {e}")
                        continue
            
            # 열 너비 자동 조정 (병합된 셀 위치에서도 안전하게 동작하도록 보강)
            for idx, col in enumerate(ws.iter_cols(1, ws.max_column)):
                max_length = 0
                column_letter = get_column_letter(idx + 1)
                for cell in col:
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        continue
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        output_filename = f'지역경제동향_{year}년_{quarter}분기.xlsx'
        output_path = UPLOAD_FOLDER / output_filename
        
        wb.save(output_path)
        
        # 파일을 바이트로 읽어서 base64로 인코딩
        with open(output_path, 'rb') as f:
            xlsx_data = base64.b64encode(f.read()).decode('utf-8')
        
        return jsonify({
            'success': True,
            'filename': output_filename,
            'download_url': f'/uploads/{output_filename}',
            'total_pages': len(pages),
            'xlsx_data': xlsx_data,
            'image_count': image_counter
        })
        
    except ImportError as e:
        return jsonify({
            'success': False, 
            'error': f'필요한 라이브러리가 설치되지 않았습니다: {str(e)}. pip install openpyxl pillow beautifulsoup4'
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


@api_bp.route('/cleanup-uploads', methods=['POST'])
def cleanup_uploads():
    """업로드 폴더 정리 API (작업 완료 후 호출)"""
    try:
        deleted_count = cleanup_upload_folder(keep_current_files=True, cleanup_excel_only=True)
        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'{deleted_count}개 파일이 정리되었습니다.'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })


@api_bp.route('/render-chart-image', methods=['POST'])
def render_chart_image():
    """차트/인포그래픽을 이미지로 렌더링"""
    try:
        data = request.get_json(silent=True)
        if data is None:
            return jsonify({'success': False, 'error': 'JSON 형식의 요청 데이터가 필요합니다.'}), 400
        
        image_data = data.get('image_data', '')
        filename = data.get('filename', 'chart.png')
        
        if not image_data:
            return jsonify({'success': False, 'error': '이미지 데이터가 없습니다.'})
        
        match = re.match(r'data:([^;]+);base64,(.+)', image_data)
        if match:
            mimetype = match.group(1)
            img_data = base64.b64decode(match.group(2))

            safe_name = Path(filename).name or 'chart.png'
            img_path = UPLOAD_FOLDER / safe_name
            with open(img_path, 'wb') as f:
                f.write(img_data)
            
            return jsonify({
                'success': True,
                'filename': safe_name,
                'path': str(img_path),
                'url': f'/uploads/{safe_name}'
            })
        else:
            return jsonify({'success': False, 'error': '잘못된 이미지 데이터 형식입니다.'})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


# 레거시 엔드포인트 - 기초자료 수집표는 사용하지 않으므로 비활성화됨
# @api_bp.route('/get-industry-weights', methods=['GET'])
# def get_industry_weights():
#     """기초자료에서 업종별 가중치 정보 추출"""
#     import pandas as pd
#     
#     sheet_type = request.args.get('sheet_type', '광공업생산')
#     raw_excel_path = session.get('raw_excel_path')
#     
#     if not raw_excel_path or not Path(raw_excel_path).exists():
#         return jsonify({
#             'success': False, 
#             'error': '기초자료 파일을 찾을 수 없습니다. 먼저 파일을 업로드하세요.'
#         })
#     
#     try:
#         xl = pd.ExcelFile(raw_excel_path)
#         
#         # 시트 매핑
#         sheet_mapping = {
#             '광공업생산': '광공업생산',
#             '서비스업생산': '서비스업생산'
#         }
#         
#         sheet_name = sheet_mapping.get(sheet_type)
#         if not sheet_name or sheet_name not in xl.sheet_names:
#             return jsonify({
#                 'success': False,
#                 'error': f'시트를 찾을 수 없습니다: {sheet_type}'
#             })
#         
#         df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
#         
#         # 업종별 정보 추출 (열 구조에 따라 다름)
#         industries = []
#         
#         if sheet_type == '광공업생산':
#             # 광공업생산 시트: 열 4=업종명, 열 8=가중치 (또는 해당 열 확인 필요)
#             name_col = 4  # 업종명 열
#             weight_col = 8  # 가중치 열
#             
#             for i, row in df.iterrows():
#                 if i < 3:  # 헤더 행 건너뛰기
#                     continue
#                     
#                 name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
#                 if not name or name in ['nan', 'NaN', '업종이름', '업종명']:
#                     continue
#                     
#                 weight = None
#                 if weight_col < len(row) and pd.notna(row[weight_col]):
#                     try:
#                         weight = float(row[weight_col])
#                     except (ValueError, TypeError):
#                         pass
#                 
#                 industries.append({
#                     'row': i + 1,
#                     'name': name,
#                     'weight': weight
#                 })
#                 
#         elif sheet_type == '서비스업생산':
#             # 서비스업생산 시트: 열 4=업종명, 열 8=가중치
#             name_col = 4  # 업종명 열
#             weight_col = 8  # 가중치 열
#             
#             for i, row in df.iterrows():
#                 if i < 3:  # 헤더 행 건너뛰기
#                     continue
#                     
#                 name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
#                 if not name or name in ['nan', 'NaN', '업종이름', '업종명']:
#                     continue
#                     
#                 weight = None
#                 if weight_col < len(row) and pd.notna(row[weight_col]):
#                     try:
#                         weight = float(row[weight_col])
#                     except (ValueError, TypeError):
#                         pass
#                 
#                 industries.append({
#                     'row': i + 1,
#                     'name': name,
#                     'weight': weight
#                 })
#         
#         return jsonify({
#             'success': True,
#             'sheet_type': sheet_type,
#             'industries': industries[:100]  # 최대 100개
#         })
#         
#     except Exception as e:
#         import traceback
#         traceback.print_exc()
#         return jsonify({'success': False, 'error': f'업종 정보 추출 실패: {str(e)}'})


@api_bp.route('/export-hwp-import', methods=['POST'])
def export_hwp_import():
    """한글 프로그램에서 열 수 있는 XML/HTML 문서 생성 - 차트는 이미지로 변환됨"""
    try:
        from datetime import datetime
        
        data = request.get_json(silent=True)
        if data is None:
            return jsonify({'success': False, 'error': 'JSON 형식의 요청 데이터가 필요합니다.'}), 400
        
        pages = data.get('pages', [])
        year = data.get('year', session.get('year'))
        quarter = data.get('quarter', session.get('quarter'))
        if year is None or quarter is None:
            excel_path = session.get('excel_path')
            year, quarter, resolve_err = _resolve_year_quarter(excel_path, year, quarter)
            if year is None or quarter is None:
                return jsonify({'success': False, 'error': resolve_err or '연도/분기 정보가 없습니다.'}), 400
        doc_format = data.get('format', 'hwp-xml')  # hwp-xml 형식
        
        if not pages:
            return jsonify({'success': False, 'error': '페이지 데이터가 없습니다.'})
        
        # 한글 호환 XML/HTML 생성 (모든 이미지가 이미 Base64로 포함됨)
        final_html = f'''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE html PUBLIC "-//W3C//DTD XHTML 1.0 Transitional//EN" "http://www.w3.org/TR/xhtml1/DTD/xhtml1-transitional.dtd">
<html xmlns="http://www.w3.org/1999/xhtml" lang="ko" xml:lang="ko">
<head>
    <meta http-equiv="Content-Type" content="text/html; charset=UTF-8" />
    <meta name="generator" content="지역경제동향 보도자료 시스템" />
    <title>{year}년 {quarter}/4분기 지역경제동향</title>
    <style type="text/css">
        /* 한글 호환 기본 스타일 */
        @page {{
            size: A4 portrait;
            margin: 20mm 15mm 20mm 15mm;
        }}
        
        body {{
            font-family: '맑은 고딕', 'Malgun Gothic', '바탕', 'Batang', serif;
            font-size: 10pt;
            line-height: 160%;
            color: #000000;
            background-color: #ffffff;
            margin: 0;
            padding: 0;
        }}
        
        /* 페이지 컨테이너 */
        .page-container {{
            width: 180mm;
            margin: 0 auto;
            padding: 10mm 0;
            page-break-after: always;
        }}
        
        .page-container:last-child {{
            page-break-after: auto;
        }}
        
        /* 제목 스타일 */
        h1 {{
            font-family: '맑은 고딕', 'Malgun Gothic', sans-serif;
            font-size: 18pt;
            font-weight: bold;
            color: #000000;
            margin: 0 0 15px 0;
            padding: 0;
            line-height: 140%;
        }}
        
        h2 {{
            font-family: '맑은 고딕', 'Malgun Gothic', sans-serif;
            font-size: 14pt;
            font-weight: bold;
            color: #000000;
            margin: 20px 0 10px 0;
            padding: 8px 10px;
            background-color: #f0f0f0;
            border-left: 4px solid #0066cc;
        }}
        
        h3 {{
            font-family: '맑은 고딕', 'Malgun Gothic', sans-serif;
            font-size: 12pt;
            font-weight: bold;
            color: #000000;
            margin: 15px 0 8px 0;
        }}
        
        h4 {{
            font-family: '맑은 고딕', 'Malgun Gothic', sans-serif;
            font-size: 11pt;
            font-weight: bold;
            color: #000000;
            margin: 10px 0 5px 0;
        }}
        
        /* 문단 스타일 */
        p {{
            font-family: '맑은 고딕', 'Malgun Gothic', sans-serif;
            font-size: 10pt;
            margin: 5px 0;
            line-height: 160%;
            text-align: justify;
        }}
        
        /* 표 스타일 - 한글 완벽 호환 */
        table {{
            border-collapse: collapse;
            width: 100%;
            margin: 10px 0;
            font-size: 9pt;
            border: 1px solid #000000;
            table-layout: fixed;
        }}
        
        th {{
            border: 1px solid #000000;
            padding: 6px 4px;
            text-align: center;
            vertical-align: middle;
            background-color: #d9d9d9;
            font-weight: bold;
            font-family: '맑은 고딕', 'Malgun Gothic', sans-serif;
        }}
        
        td {{
            border: 1px solid #000000;
            padding: 5px 4px;
            text-align: center;
            vertical-align: middle;
            font-family: '맑은 고딕', 'Malgun Gothic', sans-serif;
        }}
        
        /* 이미지 스타일 - 원본 비율 유지 */
        img {{
            max-width: 100%;
            height: auto;
            display: block;
            margin: 10px auto;
        }}
        
        .chart-image-converted, .svg-image-converted {{
            max-width: 100%;
            width: auto;
            height: auto;
            display: block;
            margin: 10px auto;
        }}
        
        /* 리스트 스타일 */
        ul, ol {{
            margin: 10px 0 10px 20px;
            padding: 0;
        }}
        
        li {{
            margin: 3px 0;
            line-height: 160%;
        }}
        
        /* 페이지 번호 */
        .page-number {{
            text-align: center;
            font-size: 9pt;
            color: #666666;
            margin-top: 20px;
            padding-top: 10px;
            border-top: 1px solid #cccccc;
        }}
        
        /* 인쇄 스타일 */
        @media print {{
            body {{
                background-color: #ffffff;
            }}
            .page-container {{
                page-break-after: always;
            }}
        }}
    </style>
</head>
<body>
'''
        
        # 각 페이지 처리
        for idx, page in enumerate(pages, 1):
            page_html = page.get('html', '')
            page_title = page.get('title', f'페이지 {idx}')
            category = page.get('category', '')
            
            # body 내용 추출
            body_content = page_html
            if '<body' in page_html.lower():
                body_match = re.search(r'<body[^>]*>(.*?)</body>', page_html, re.DOTALL | re.IGNORECASE)
                if body_match:
                    body_content = body_match.group(1)
            
            # 불필요한 태그 제거
            body_content = re.sub(r'<style[^>]*>.*?</style>', '', body_content, flags=re.DOTALL)
            body_content = re.sub(r'<script[^>]*>.*?</script>', '', body_content, flags=re.DOTALL)
            body_content = re.sub(r'<link[^>]*/?>', '', body_content)
            body_content = re.sub(r'<meta[^>]*/?>', '', body_content)
            body_content = re.sub(r'<!DOCTYPE[^>]*>', '', body_content)
            body_content = re.sub(r'<html[^>]*>', '', body_content)
            body_content = re.sub(r'</html>', '', body_content)
            body_content = re.sub(r'<head[^>]*>.*?</head>', '', body_content, flags=re.DOTALL)
            
            # 그래프/차트 요소 제거
            body_content = _strip_chart_elements(body_content)
            
            # 표에 인라인 스타일 강화 (한글 완벽 호환)
            body_content = _add_hwp_compatible_styles(body_content)
            
            # 카테고리 한글명
            category_names = {
                'summary': '요약',
                'sectoral': '부문별',
                'regional': '시도별',
                'statistics': '통계표'
            }
            category_name = category_names.get(category, '')
            
            # 페이지 래퍼 추가
            final_html += f'''
    <!-- 페이지 {idx}: {page_title} -->
    <div class="page-container">
        <h2>[{category_name}] {page_title}</h2>
        {body_content}
        <p class="page-number">- {idx} / {len(pages)} -</p>
    </div>
'''
        
        final_html += '''
</body>
</html>
'''
        
        # 파일 저장
        output_filename = f'지역경제동향_{year}년_{quarter}분기_한글용.html'
        output_path = UPLOAD_FOLDER / output_filename
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(final_html)
        
        return jsonify({
            'success': True,
            'html': final_html,
            'filename': output_filename,
            'download_url': f'/uploads/{output_filename}',
            'total_pages': len(pages),
            'message': '한글용 문서가 생성되었습니다. 한글에서 파일 → 불러오기로 열 수 있습니다.'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


def _add_hwp_compatible_styles(html_content):
    """한글 프로그램 완벽 호환을 위한 인라인 스타일 추가"""
    
    # table 태그에 인라인 스타일 추가
    html_content = re.sub(
        r'<table([^>]*)>',
        r'<table\1 style="border-collapse: collapse; width: 100%; margin: 10px 0; font-size: 9pt; border: 1px solid #000000; table-layout: fixed;">',
        html_content
    )
    
    # th 태그에 인라인 스타일 추가
    html_content = re.sub(
        r'<th([^>]*)>',
        r'<th\1 style="border: 1px solid #000000; padding: 6px 4px; text-align: center; vertical-align: middle; background-color: #d9d9d9; font-weight: bold; font-family: 맑은 고딕, Malgun Gothic, sans-serif;">',
        html_content
    )
    
    # td 태그에 인라인 스타일 추가
    html_content = re.sub(
        r'<td([^>]*)>',
        r'<td\1 style="border: 1px solid #000000; padding: 5px 4px; text-align: center; vertical-align: middle; font-family: 맑은 고딕, Malgun Gothic, sans-serif;">',
        html_content
    )
    
    # 제목 태그들에 인라인 스타일 추가
    html_content = re.sub(
        r'<h1([^>]*)>',
        r'<h1\1 style="font-family: 맑은 고딕, Malgun Gothic, sans-serif; font-size: 18pt; font-weight: bold; margin: 0 0 15px 0;">',
        html_content
    )
    html_content = re.sub(
        r'<h2([^>]*)>',
        r'<h2\1 style="font-family: 맑은 고딕, Malgun Gothic, sans-serif; font-size: 14pt; font-weight: bold; margin: 20px 0 10px 0;">',
        html_content
    )
    html_content = re.sub(
        r'<h3([^>]*)>',
        r'<h3\1 style="font-family: 맑은 고딕, Malgun Gothic, sans-serif; font-size: 12pt; font-weight: bold; margin: 15px 0 8px 0;">',
        html_content
    )
    html_content = re.sub(
        r'<h4([^>]*)>',
        r'<h4\1 style="font-family: 맑은 고딕, Malgun Gothic, sans-serif; font-size: 11pt; font-weight: bold; margin: 10px 0 5px 0;">',
        html_content
    )
    
    # p 태그에 스타일 추가
    html_content = re.sub(
        r'<p([^>]*)>',
        r'<p\1 style="font-family: 맑은 고딕, Malgun Gothic, sans-serif; font-size: 10pt; margin: 5px 0; line-height: 160%;">',
        html_content
    )
    
    # img 태그에 스타일 추가 (원본 비율 유지)
    html_content = re.sub(
        r'<img([^>]*)>',
        r'<img\1 style="max-width: 100%; height: auto; display: block; margin: 10px auto;">',
        html_content
    )
    
    # ul, ol 태그에 스타일 추가
    html_content = re.sub(
        r'<ul([^>]*)>',
        r'<ul\1 style="margin: 10px 0 10px 20px; font-family: 맑은 고딕, Malgun Gothic, sans-serif;">',
        html_content
    )
    html_content = re.sub(
        r'<ol([^>]*)>',
        r'<ol\1 style="margin: 10px 0 10px 20px; font-family: 맑은 고딕, Malgun Gothic, sans-serif;">',
        html_content
    )
    html_content = re.sub(
        r'<li([^>]*)>',
        r'<li\1 style="margin: 3px 0; line-height: 160%;">',
        html_content
    )
    
    return html_content


def _strip_chart_elements(html_content: str) -> str:
    """그래프/차트 관련 요소 제거"""
    if not html_content:
        return html_content

    chart_class_pattern = (
        r'chart-container|chart-wrapper|chart-area|graph-container|'
        r'chart-canvas-wrapper|chart-title|chart-image-converted|svg-image-converted'
    )

    html_content = re.sub(
        rf'<div[^>]*class=["\"][^"\"]*(?:{chart_class_pattern})[^"\"]*["\"][^>]*>.*?</div>',
        '',
        html_content,
        flags=re.DOTALL | re.IGNORECASE
    )

    html_content = re.sub(r'<canvas[^>]*>.*?</canvas>', '', html_content, flags=re.DOTALL | re.IGNORECASE)
    html_content = re.sub(r'<canvas[^>]*/?>', '', html_content, flags=re.IGNORECASE)
    html_content = re.sub(r'<svg[^>]*>.*?</svg>', '', html_content, flags=re.DOTALL | re.IGNORECASE)
    html_content = re.sub(
        rf'<img[^>]*class=["\"][^"\"]*(?:{chart_class_pattern})[^"\"]*["\"][^>]*>',
        '',
        html_content,
        flags=re.DOTALL | re.IGNORECASE
    )

    return html_content


def _strip_placeholders(html_content: str) -> str:
    """편집용 placeholder 제거"""
    if not html_content:
        return html_content

    html_content = re.sub(
        r'<span[^>]*class=["\"][^"\"]*\beditable-placeholder\b[^"\"]*["\"][^>]*>.*?</span>',
        '',
        html_content,
        flags=re.DOTALL | re.IGNORECASE
    )
    html_content = re.sub(r'\[[^\]]*입력 필요\]', '', html_content)
    html_content = re.sub(r'\[\s*\]', '', html_content)

    return html_content


def _strip_page_wrapper(html_content: str) -> str:
    """페이지 래퍼 div 제거 (중첩 방지)"""
    if not html_content:
        return html_content

    page_open_pattern = r'<div[^>]*class=["\"][^"\"]*\bpage\b[^"\"]*["\"][^>]*>'
    if not re.search(page_open_pattern, html_content, flags=re.IGNORECASE):
        return html_content

    html_content = re.sub(page_open_pattern, '', html_content, count=1, flags=re.IGNORECASE)
    html_content = re.sub(r'</div>\s*$', '', html_content.strip(), count=1)
    return html_content


def _create_placeholder_image(image_path):
    """플레이스홀더 이미지 생성"""
    try:
        from PIL import Image, ImageDraw, ImageFont
        import os
        
        # 800x400 크기의 플레이스홀더 이미지 생성
        img = Image.new('RGB', (800, 400), color='#f5f5f5')
        draw = ImageDraw.Draw(img)
        
        # 테두리
        draw.rectangle([(0, 0), (799, 399)], outline='#666', width=2)
        
        # 텍스트
        try:
            # 한글 폰트 시도 (시스템 폰트)
            font = ImageFont.truetype('/System/Library/Fonts/AppleGothic.ttf', 24)
        except:
            try:
                font = ImageFont.truetype('/usr/share/fonts/truetype/nanum/NanumGothic.ttf', 24)
            except:
                font = ImageFont.load_default()
        
        text = '📊 차트 영역\n\n브라우저에서 HTML을 열어\nCanvas를 이미지로 변환하세요'
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        position = ((800 - text_width) // 2, (400 - text_height) // 2)
        
        draw.text(position, text, fill='#666', font=font, align='center')
        
        image_path.parent.mkdir(parents=True, exist_ok=True)
        img.save(str(image_path), 'PNG')
        return True
    except ImportError:
        # PIL이 없으면 기본 이미지 생성 스킵
        return False
    except Exception as e:
        print(f"[경고] 플레이스홀더 이미지 생성 실패: {e}")
        return False
def _export_hwp_ready_core(pages, year, quarter, output_folder=EXPORT_FOLDER):
    """한글(HWP) 복붙용 HTML을 생성하고 지정 폴더에 저장"""
    try:
        if not pages:
            def _safe_output_name(name: str) -> str:
                if not name or not isinstance(name, str):
                    name = 'unknown'
                return name.replace('/', '_').replace('\\', '_').replace('..', '_')

            ordered_files = []

            # 1) 요약 → 2) 부문별 → 3) 시도별 (요청 순서)
            for report in SUMMARY_REPORTS:
                report_name = report.get('name', report.get('id', 'Unknown'))
                safe_name = _safe_output_name(report_name)
                path = TEMP_OUTPUT_DIR / f"{safe_name}_output.html"
                if path.exists():
                    ordered_files.append((report_name, path))

            for report in SECTOR_REPORTS:
                report_name = report.get('name', report.get('id', 'Unknown'))
                safe_name = _safe_output_name(report_name)
                path = TEMP_OUTPUT_DIR / f"{safe_name}_output.html"
                if path.exists():
                    ordered_files.append((report_name, path))

            for region in REGIONAL_REPORTS:
                region_name = region.get('name', region.get('id', 'Unknown'))
                safe_name = _safe_output_name(region_name)
                path = TEMP_REGIONAL_OUTPUT_DIR / f"{safe_name}_output.html"
                if path.exists():
                    ordered_files.append((f"시도별-{region_name}", path))

            if not ordered_files:
                return {'success': False, 'error': '생성된 보도자료가 없습니다. 먼저 "전체 생성"을 실행하세요.'}

            for title, output_file in ordered_files:
                with open(output_file, 'r', encoding='utf-8') as f:
                    html_content = f.read()
                    pages.append({'title': title, 'html': html_content})

        if not pages:
            return {'success': False, 'error': '페이지 데이터가 없습니다.'}

        final_html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{year}년 {quarter}/4분기 지역경제동향</title>
    <style>
        /* 브라우저 미리보기용 스타일 (한글 복붙 시에는 인라인 스타일 적용됨) */
        body {{
            font-family: 'Malgun Gothic', '맑은 고딕', 'Dotum', '돋움', sans-serif;
            font-size: 10pt;
            line-height: 1.5;
            color: #000;
            background: #fff;
            padding: 20px;
            max-width: 210mm;
            margin: 0 auto;
        }}
        .copy-btn {{
            position: fixed;
            top: 10px;
            right: 10px;
            background: #0066cc;
            color: white;
            padding: 12px 20px;
            border: none;
            border-radius: 5px;
            font-size: 12pt;
            cursor: pointer;
            z-index: 9999;
            box-shadow: 0 2px 10px rgba(0,0,0,0.3);
        }}
        .copy-btn:hover {{ background: #0055aa; }}
        @media print {{ .copy-btn {{ display: none; }} }}
    </style>
</head>
<body>
    <button class="copy-btn" onclick="copyAll()">📋 전체 복사 (클릭)</button>
    
    <div id="hwp-content">
'''

        excluded_report_ids = {'cover', 'toc', 'stat_toc', 'guide', 'infographic', 'stat_appendix', 'stat_grdp'}
        is_first_page = True

        for idx, page in enumerate(pages, 1):
            page_html = page.get('html', '')
            page_title = page.get('title', f'페이지 {idx}')
            report_id = page.get('report_id', '')

            if report_id in excluded_report_ids:
                print(f"[HTML 내보내기] 제외: {report_id} ({page_title})")
                continue

            body_content = page_html
            if '<body' in page_html.lower():
                body_match = re.search(r'<body[^>]*>(.*?)</body>', page_html, re.DOTALL | re.IGNORECASE)
                if body_match:
                    body_content = body_match.group(1)

            body_content = re.sub(r'<style[^>]*>.*?</style>', '', body_content, flags=re.DOTALL)
            body_content = re.sub(r'<script[^>]*>.*?</script>', '', body_content, flags=re.DOTALL)
            body_content = re.sub(r'<link[^>]*>', '', body_content)
            body_content = re.sub(r'<meta[^>]*>', '', body_content)

            body_content = _strip_chart_elements(body_content)
            body_content = _strip_placeholders(body_content)
            body_content = _strip_page_wrapper(body_content)

            body_content = _add_table_inline_styles(body_content)

            if not is_first_page:
                final_html += '\n<div style="height: 1em;"></div>\n'
            is_first_page = False

            final_html += f'''
            <!-- 페이지 {idx}: {page_title} -->
{body_content}
'''

        final_html += '''
    </div>
    
    <script>
        function copyAll() {
            const content = document.getElementById('hwp-content');
            const range = document.createRange();
            range.selectNodeContents(content);
            const selection = window.getSelection();
            selection.removeAllRanges();
            selection.addRange(range);
            
            try {
                document.execCommand('copy');
                alert('복사 완료!\\n\\n한글(HWP)에서 Ctrl+V로 붙여넣기 하세요.\\n※ 표와 서식이 유지됩니다.');
            } catch (e) {
                alert('자동 복사 실패.\\nCtrl+A로 전체 선택 후 Ctrl+C로 복사하세요.');
            }
            
            selection.removeAllRanges();
        }
        
        document.addEventListener('keydown', function(e) {
            if (e.ctrlKey && e.key === 'a') {
                e.preventDefault();
                copyAll();
            }
        });
    </script>
</body>
</html>
'''

        output_filename = f'지역경제동향_{year}년_{quarter}분기.html'
        output_folder.mkdir(parents=True, exist_ok=True)
        output_path = output_folder / output_filename

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(final_html)

        return {
            'success': True,
            'html': final_html,
            'filename': output_filename,
            'view_url': f'/exports/{output_filename}',
            'download_url': f'/exports/{output_filename}',
            'total_pages': len(pages),
            'output_path': str(output_path)
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}


@api_bp.route('/export-hwp-ready', methods=['POST'])
def export_hwp_ready():
    """한글(HWP) 복붙용 HTML 문서 생성 - 인라인 스타일 최적화"""
    data = request.get_json(silent=True)
    if data is None:
        data = {}
    pages = data.get('pages', [])
    year = data.get('year', session.get('year'))
    quarter = data.get('quarter', session.get('quarter'))

    if year is None or quarter is None:
        excel_path = session.get('excel_path')
        year, quarter, resolve_err = _resolve_year_quarter(excel_path, year, quarter)
        if year is None or quarter is None:
            return jsonify({'success': False, 'error': resolve_err or '연도/분기 정보가 없습니다.'}), 400

    result = _export_hwp_ready_core(pages, year, quarter, output_folder=EXPORT_FOLDER)
    status = 200 if result.get('success') else 500
    return jsonify(result), status


@api_bp.route('/save-html-to-project', methods=['POST'])
def save_html_to_project():
    """HTML 보도자료를 프로젝트 메인 디렉토리에 저장"""
    try:
        from datetime import datetime
        
        data = request.get_json(silent=True)
        if data is None:
            return jsonify({'success': False, 'error': 'JSON 형식의 요청 데이터가 필요합니다.'}), 400
        
        pages = data.get('pages', [])
        year = data.get('year', session.get('year'))
        quarter = data.get('quarter', session.get('quarter'))
        if year is None or quarter is None:
            excel_path = session.get('excel_path')
            year, quarter, resolve_err = _resolve_year_quarter(excel_path, year, quarter)
            if year is None or quarter is None:
                return jsonify({'success': False, 'error': resolve_err or '연도/분기 정보가 없습니다.'}), 400
        
        if not pages:
            return jsonify({'success': False, 'error': '페이지 데이터가 없습니다.'})
        
        # 모든 페이지의 스타일 수집
        all_styles = set()
        
        final_html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{year}년 {quarter}/4분기 지역경제동향</title>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;600;700&display=swap');
        
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        html, body {{
            width: 210mm;
            background: white;
        }}
        
        body {{
            font-family: 'Malgun Gothic', '맑은 고딕', 'Dotum', '돋움', sans-serif;
            font-size: 10pt;
            line-height: 1.5;
            color: #000;
        }}
        
        /* PDF 출력용 페이지 스타일 - STYLE_GUIDE 적용 */
        /* STYLE_GUIDE: Container width: 210mm; min-height: 297mm; padding: 20mm 15mm; */
        .pdf-page {{
            width: 210mm;
            min-height: 297mm;
            max-height: 297mm;
            padding: 20mm 15mm;
            margin: 0 auto 5mm auto;
            background: white;
            position: relative;
            overflow: hidden;
            page-break-after: always;
            page-break-inside: avoid;
        }}
        
        .pdf-page:last-child {{
            page-break-after: auto;
            margin-bottom: 0;
        }}
        
        .pdf-page-content {{
            width: 100%;
            height: calc(297mm - 32mm);
            overflow: hidden;
        }}
        
        .pdf-page-content > * {{
            max-width: 100%;
        }}
        
        /* 페이지 번호 */
        .pdf-page-number {{
            position: absolute;
            bottom: 8mm;
            left: 0;
            right: 0;
            text-align: center;
            font-size: 9pt;
            color: #333;
        }}
        
        /* 화면 미리보기용 */
        @media screen {{
            body {{
                background: #f0f0f0;
                padding: 20px;
            }}
            
            .pdf-page {{
                box-shadow: 0 2px 10px rgba(0,0,0,0.15);
                border: 1px solid #ddd;
            }}
        }}
        
        /* 인쇄/PDF 저장용 */
        @media print {{
            html, body {{
                width: 210mm;
                background: white !important;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }}
            
            body {{
                padding: 0;
                margin: 0;
            }}
            
            .pdf-page {{
                width: 210mm;
                height: 297mm;
                min-height: 297mm;
                max-height: 297mm;
                padding: 12mm 15mm 15mm 15mm;
                margin: 0;
                box-shadow: none;
                border: none;
                page-break-after: always;
                page-break-inside: avoid;
            }}
            
            .pdf-page:last-child {{
                page-break-after: auto;
            }}
            
            /* 차트 색상 유지 */
            canvas {{
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }}
            
            .pdf-page {{
                padding: 20mm 15mm !important; /* STYLE_GUIDE: padding: 20mm 15mm */
            }}
        }}
        
        @page {{
            size: A4 portrait;
            margin: 0;
        }}
        
        /* 표 스타일 공통 - STYLE_GUIDE 적용 */
        /* STYLE_GUIDE: Table Borders - Top/Bottom 2px solid #000, Header Separator 1px solid #888, Inner Grid 1px solid #DDDDDD */
        table {{
            border-collapse: collapse;
            width: 100%;
            border-top: 2px solid #000000;
            border-bottom: 2px solid #000000;
        }}
        
        th, td {{
            border-left: 1px solid #DDDDDD;
            border-right: 1px solid #DDDDDD;
            padding: 4px 6px;
            font-size: 11pt; /* STYLE_GUIDE: Table Body 11pt */
            vertical-align: middle;
        }}
        
        th {{
            border-top: 1px solid #888;
            border-bottom: 1px solid #888;
            background: #F5F7FA; /* STYLE_GUIDE: Background (Th): #F5F7FA */
            text-align: center; /* STYLE_GUIDE: th Center / Middle */
            font-weight: normal;
        }}
        
        td {{
            text-align: center; /* STYLE_GUIDE: td (Text): Center (default) */
        }}
        
        td.number, td[data-type="number"] {{
            text-align: right; /* STYLE_GUIDE: td (Number): Right */
            padding-right: 4px;
        }}
        
        /* 차트 크기 조정 */
        .chart-container, .chart-wrapper {{
            max-width: 100%;
        }}
        
        canvas {{
            max-width: 100% !important;
            height: auto !important;
        }}
    </style>
'''
        
        # 각 페이지에서 스타일 추출하여 추가
        for idx, page in enumerate(pages):
            page_html = page.get('html', '')
            if '<style' in page_html:
                style_matches = re.findall(r'<style[^>]*>(.*?)</style>', page_html, re.DOTALL)
                for style in style_matches:
                    # 중복 방지를 위해 hash 사용
                    style_hash = hash(style.strip())
                    if style_hash not in all_styles:
                        all_styles.add(style_hash)
                        final_html += f'    <style>/* Page {idx+1} styles */\n{style}\n    </style>\n'
        
        final_html += '''</head>
<body>
'''
        
        for idx, page in enumerate(pages, 1):
            page_html = page.get('html', '')
            page_title = page.get('title', f'페이지 {idx}')
            
            # body 내용 추출
            body_content = page_html
            if '<body' in page_html.lower():
                body_match = re.search(r'<body[^>]*>(.*?)</body>', page_html, re.DOTALL | re.IGNORECASE)
                if body_match:
                    body_content = body_match.group(1)
            
            # 내용에서 style 태그 제거 (이미 head에 추가됨)
            body_content = re.sub(r'<style[^>]*>.*?</style>', '', body_content, flags=re.DOTALL)
            
            # 페이지 래퍼 추가
            final_html += f'''
    <!-- Page {idx}: {page_title} -->
    <div class="pdf-page" data-page="{idx}" data-title="{page_title}">
        <div class="pdf-page-content">
{body_content}
        </div>
        <div class="pdf-page-number">- {idx} -</div>
    </div>
'''
        
        final_html += '''
    <script>
        // 인쇄 전 준비
        window.onbeforeprint = function() {
            document.body.style.background = 'white';
        };
        
        // Ctrl+P로 PDF 저장 안내
        console.log('PDF 저장: Ctrl+P (또는 Cmd+P) → "PDF로 저장" 선택');
    </script>
</body>
</html>
'''
        
        # 프로젝트 메인 디렉토리에 저장
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f'지역경제동향_{year}년_{quarter}분기_{timestamp}.html'
        output_path = BASE_DIR / output_filename
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(final_html)
        
        print(f"[HTML 저장] 프로젝트 메인 디렉토리에 저장됨: {output_path}")
        
        return jsonify({
            'success': True,
            'filename': output_filename,
            'path': str(output_path),
            'total_pages': len(pages),
            'message': f'HTML 파일이 프로젝트 메인 디렉토리에 저장되었습니다: {output_filename}'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


def _add_table_inline_styles(html_content):
    """표에 인라인 스타일 추가 (한글 복붙 최적화) - STYLE_GUIDE_FROM_PNG.md 적용"""
    # table 태그에 인라인 스타일 추가
    # STYLE_GUIDE: Font Family: 'Malgun Gothic', 'Dotum', sans-serif
    # STYLE_GUIDE: Base Size: 10pt (Body), 11pt (Table Body), 9pt (Dense Table)
    html_content = re.sub(
        r'<table([^>]*)>',
        r'<table\1 style="border-collapse: collapse; width: 100%; margin: 10px 0; border-top: 2px solid #000000; border-bottom: 2px solid #000000; font-family: \'Malgun Gothic\', \'맑은 고딕\', \'Dotum\', \'돋움\', sans-serif; font-size: 11pt;">',
        html_content
    )
    
    # th 태그에 인라인 스타일 추가
    # STYLE_GUIDE: Background (Th): #F5F7FA, Header Separator: 1px solid #888
    # STYLE_GUIDE: Alignment: th Center / Middle
    html_content = re.sub(
        r'<th([^>]*)>',
        r'<th\1 style="border-top: 1px solid #888; border-bottom: 1px solid #888; border-left: 1px solid #DDDDDD; border-right: 1px solid #DDDDDD; padding: 4px 6px; text-align: center; vertical-align: middle; background-color: #F5F7FA; font-weight: normal; font-family: \'Malgun Gothic\', \'맑은 고딕\', \'Dotum\', \'돋움\', sans-serif; font-size: 11pt;">',
        html_content
    )
    
    # td 태그에 인라인 스타일 추가
    # STYLE_GUIDE: Inner Grid: 1px solid #DDDDDD
    # STYLE_GUIDE: td (Text): Center, td (Number): Right (padding-right: 4px)
    html_content = re.sub(
        r'<td([^>]*)>',
        r'<td\1 style="border: 1px solid #DDDDDD; padding: 4px 6px; text-align: center; vertical-align: middle; font-family: \'Malgun Gothic\', \'맑은 고딕\', \'Dotum\', \'돋움\', sans-serif; font-size: 11pt;">',
        html_content
    )
    
    # 제목 태그들에 인라인 스타일 추가 - STYLE_GUIDE 적용
    # STYLE_GUIDE: Font Family: 'Malgun Gothic', 'Dotum', sans-serif
    # STYLE_GUIDE: Primary (Header): #2B4E88 or #000000
    html_content = re.sub(
        r'<h1([^>]*)>',
        r'<h1\1 style="font-family: \'Malgun Gothic\', \'맑은 고딕\', \'Dotum\', \'돋움\', sans-serif; font-size: 14pt; font-weight: bold; color: #000000; margin: 15px 0 10px 0;">',
        html_content
    )
    html_content = re.sub(
        r'<h2([^>]*)>',
        r'<h2\1 style="font-family: \'Malgun Gothic\', \'맑은 고딕\', \'Dotum\', \'돋움\', sans-serif; font-size: 13pt; font-weight: bold; color: #000000; margin: 15px 0 10px 0;">',
        html_content
    )
    html_content = re.sub(
        r'<h3([^>]*)>',
        r'<h3\1 style="font-family: \'Malgun Gothic\', \'맑은 고딕\', \'Dotum\', \'돋움\', sans-serif; font-size: 11.5pt; font-weight: bold; color: #000000; border-bottom: 1px solid #000000; padding-bottom: 3px; margin: 10px 0 8px 0;">',
        html_content
    )
    html_content = re.sub(
        r'<h4([^>]*)>',
        r'<h4\1 style="font-family: \'Malgun Gothic\', \'맑은 고딕\', \'Dotum\', \'돋움\', sans-serif; font-size: 11pt; font-weight: bold; color: #000000; margin: 10px 0 5px 0;">',
        html_content
    )
    
    # p 태그에 스타일 추가 - STYLE_GUIDE: Base Size 10pt, Line Height 1.5 ~ 1.6
    html_content = re.sub(
        r'<p([^>]*)>',
        r'<p\1 style="font-family: \'Malgun Gothic\', \'맑은 고딕\', \'Dotum\', \'돋움\', sans-serif; font-size: 10pt; margin: 5px 0; line-height: 1.5;">',
        html_content
    )
    
    # ul, ol 태그에 스타일 추가
    html_content = re.sub(
        r'<ul([^>]*)>',
        r'<ul\1 style="margin: 10px 0 10px 25px; font-family: \'Malgun Gothic\', \'맑은 고딕\', \'Dotum\', \'돋움\', sans-serif; font-size: 10pt;">',
        html_content
    )
    html_content = re.sub(
        r'<ol([^>]*)>',
        r'<ol\1 style="margin: 10px 0 10px 25px; font-family: \'Malgun Gothic\', \'맑은 고딕\', \'Dotum\', \'돋움\', sans-serif; font-size: 10pt;">',
        html_content
    )
    
    # 숫자 셀에 대해 right align 적용 (STYLE_GUIDE: td Number Right, padding-right: 4px)
    # 숫자 패턴이 포함된 td에 text-align: right 추가
    html_content = re.sub(
        r'<td([^>]*style="[^"]*)"([^>]*)>(-?\d+[\.%]?)',
        r'<td\1 text-align: right; padding-right: 4px;"\2>\3',
        html_content
    )
    html_content = re.sub(
        r'<li([^>]*)>',
        r'<li\1 style="margin: 3px 0;">',
        html_content
    )
    
    return html_content

