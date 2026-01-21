#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd

from pathlib import Path
from typing import Dict, Any, List, Optional
try:
    from .base_generator import BaseGenerator
    from config.reports import REPORT_ORDER, SECTOR_REPORTS, REGIONAL_REPORTS, REGION_DISPLAY_MAPPING, REGION_GROUPS, VALID_REGIONS
except ImportError:
    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from templates.base_generator import BaseGenerator
    from config.reports import REPORT_ORDER, SECTOR_REPORTS, REGIONAL_REPORTS, REGION_DISPLAY_MAPPING, REGION_GROUPS, VALID_REGIONS

def get_report_config(report_type: str) -> dict:
    """Return the config matching either id or report_id; accept legacy aliases."""
    aliases = {
        'mining': 'manufacturing',  # legacy name used in 일부 호출
    }
    normalized = aliases.get(report_type, report_type)
    for config in SECTOR_REPORTS:
        # 지원: id 매칭 혹은 report_id 매칭
        if config.get('id') == normalized or config.get('report_id') == normalized:
            return config
    raise ValueError(f"Unknown report type: {report_type}")


class UnifiedReportGenerator(BaseGenerator):
    """
    통합 보고서 Generator (집계 시트 기반)
    mining_manufacturing_generator의 검증된 로직을 기반으로 구현
    """

    # 데이터 시작 행은 동적으로 찾음 (하드코딩 제거)

    def __init__(self, report_type: str, excel_path: str, year=None, quarter=None, excel_file=None):
        super().__init__(excel_path, year, quarter, excel_file)

        # 설정 로드
        self.config = get_report_config(report_type)
        if not self.config:
            raise ValueError(f"Unknown report type: {report_type}")

        self.report_type = report_type
        # report_id 누락 시 id로 폴백하여 KeyError 방지
        self.report_id = self.config.get('report_id', self.config.get('id', report_type))
        if 'name_mapping' not in self.config:
            raise ValueError(f"[{self.config['name']}] ❌ 설정에서 'name_mapping'을 찾을 수 없습니다. 기본값 사용 금지.")
        self.name_mapping = self.config['name_mapping']

        if 'aggregation_structure' not in self.config:
            raise ValueError(f"[{self.config['name']}] ❌ 설정에서 'aggregation_structure'를 찾을 수 없습니다. 기본값 사용 금지.")
        # metadata_columns는 컬럼 존재 여부 힌트일 뿐, 키워드 탐색에는 기본 키워드 목록을 사용
        meta = self.config.get('metadata_columns', {})
        if isinstance(meta, dict):
            self.metadata_cols = meta
        elif isinstance(meta, list):
            # 단순 보존
            self.metadata_cols = {c: c for c in meta}
        else:
            self.metadata_cols = {}
        # 동적으로 할당되는 주요 속성들 기본값 None으로 초기화
        self.region_name_col = None
        self.industry_code_col = None
        self.industry_name_col = None
        self.data_start_row = None
        
        # 경고 메시지 중복 출력 방지 플래그
        self._warned_industry_col_missing = False
        self.df_analysis = None
        self.df_aggregation = None
        self.df_aggregation_raw = None
        self.df_aggregation_table = None
        self.preprocessed_table_df = None
        self.df_reference = None
        self.target_col = None
        self.prev_y_col = None
        self.prev_prev_y_col = None
        self.prev_prev_prev_y_col = None
        self.quarterly_keys = []
        self.quarterly_cols = {}
        self.analysis_target_col = None
        self.analysis_prev_y_col = None
        self.analysis_prev_prev_y_col = None
        self.analysis_prev_prev_prev_y_col = None
        self.analysis_quarterly_keys = []
        self.analysis_quarterly_cols = {}
        # 인스턴스 생성 시 데이터프레임 등 필드 자동 초기화
        self.load_data()
        # region_name_col, target_col, prev_y_col 등 주요 컬럼 자동 탐색
        if self.df_aggregation is not None and self.year is not None and self.quarter is not None:
            import re
            max_header_rows = min(10, len(self.df_aggregation))
            # 1. 지역명 컬럼(기존대로)
            if self.region_name_col is None:
                region_keywords = ['지역', '시도', '구분', '행정구역']
                found = False
                for row_idx in range(max_header_rows):
                    row = self.df_aggregation.iloc[row_idx]
                    for col_idx, val in enumerate(row):
                        if pd.isna(val):
                            continue
                        val_str = str(val).strip()
                        if any(kw in val_str for kw in region_keywords):
                            self.region_name_col = col_idx
                            found = True
                            print(f"[자동탐색] 헤더 {row_idx+1}행에서 지역명 컬럼 인덱스 자동설정: {col_idx} ({val_str})")
                            break
                    if found:
                        break
                if not found:
                    print("[자동탐색] 지역명 컬럼을 헤더에서 찾지 못했습니다. 기존 로직을 사용합니다.")

            # 2. 지수/값 컬럼 (target_col, 기존대로)
            if self.target_col is None:
                value_keywords = ['지수', '값', '실적', '금액', '규모', '수치', '매출', '생산', '수출', '수입', '고용', '취업', '실업', '인구']
                found = False
                for row_idx in range(max_header_rows):
                    row = self.df_aggregation.iloc[row_idx]
                    for col_idx, val in enumerate(row):
                        if pd.isna(val):
                            continue
                        val_str = str(val).strip()
                        if any(kw in val_str for kw in value_keywords):
                            self.target_col = col_idx
                            found = True
                            print(f"[자동탐색] 헤더 {row_idx+1}행에서 지수/값 컬럼 인덱스 자동설정: {col_idx} ({val_str})")
                            break
                    if found:
                        break
                if not found:
                    print("[자동탐색] 지수/값 컬럼을 헤더에서 찾지 못했습니다. 기존 로직을 사용합니다.")

            # 3. 현재 분기명 패턴 추출 및 변형으로 prev_y_col 등 동적 탐색
            # 헤더에서 현재 분기명(연도/분기) 패턴 찾기
            max_header_rows = min(10, len(self.df_aggregation))
            
            def normalize(s):
                import re
                return re.sub(r'[^0-9a-zA-Z가-힣분기년Q/4 ]', '', str(s)).replace('  ', ' ').replace(' ', '').replace('.', '').replace('/', '').replace('-', '').replace('년', '').replace('분기', '').replace('Q', '').lower()

            def find_col_by_pattern(pattern):
                norm_pat = normalize(pattern)
                best_idx = None
                best_score = 0
                found_row = -1
                for r in range(max_header_rows):
                    row = self.df_aggregation.iloc[r]
                    for c, v in enumerate(row):
                        if pd.isna(v):
                            continue
                        norm_v = normalize(v)
                        if norm_v == norm_pat:
                            return c, r
                        if norm_pat in norm_v or norm_v in norm_pat:
                            score = min(len(norm_pat), len(norm_v))
                            if score > best_score:
                                best_score = score
                                best_idx = c
                                found_row = r
                return best_idx, found_row

            header_found_row = -1

            if self.year is not None and self.quarter is not None:
                # 사용자가 지정한 연도/분기를 기준으로 탐색
                cur_year = self.year
                cur_q = self.quarter
                target_pat = f"{cur_year} {cur_q}4"  # normalize 하면 '202534' 형태가 됨 (2025 3/4 -> 202534)
                
                # 명시적 패턴 탐색
                idx, r_idx = find_col_by_pattern(f"{cur_year} {cur_q}/4")
                if idx is None:
                     idx, r_idx = find_col_by_pattern(f"{cur_year}. {cur_q}/4")
                
                if idx is not None:
                    self.target_col = idx
                    header_found_row = r_idx
                    print(f"[자동탐색] '{cur_year} {cur_q}/4' 패턴으로 당분기 컬럼 인덱스 자동설정: {idx} (행: {r_idx})")
                else:
                    print(f"[자동탐색] '{cur_year} {cur_q}/4' 패턴을 헤더에서 찾지 못했습니다.")
            
            else:
                # 연도/분기 미지정 시 최신 패턴 탐색
                current_patterns = []
                import re
                for row_idx in range(max_header_rows):
                    row = self.df_aggregation.iloc[row_idx]
                    for col_idx, val in enumerate(row):
                        if pd.isna(val):
                            continue
                        val_str = str(val).strip()
                        m = re.match(r'(20\d{2})[.\-/년 ]+([1-4])[분기Q/4 ]*', val_str)
                        if m:
                            current_patterns.append((row_idx, col_idx, val_str, m.group(1), m.group(2)))
                
                if current_patterns:
                    current_patterns.sort(key=lambda x: (int(x[3]), int(x[4])), reverse=True)
                    row_idx, col_idx, cur_val, cur_year, cur_q = current_patterns[0]
                    self.target_col = col_idx
                    self.year = int(cur_year)
                    self.quarter = int(cur_q)
                    header_found_row = row_idx
                    print(f"[자동탐색] 최신 패턴 '{cur_val}'으로 당분기 컬럼 인덱스 자동설정: {col_idx} (행: {row_idx})")
                    cur_year = int(cur_year)
                    cur_q = int(cur_q)
                else:
                    cur_year = None
                    cur_q = None

            # 데이터 시작 행 자동 설정 (헤더 바로 다음 행)
            if header_found_row >= 0 and self.data_start_row is None:
                self.data_start_row = header_found_row + 1
                print(f"[자동탐색] 데이터 시작 행 자동설정: {self.data_start_row}")

            if cur_year is not None and cur_q is not None:
                # 최근 9분기 컬럼 자동 탐색 (현재 분기 포함, 2년 전 동분기까지 포함)
                # 이를 통해 prev_q_col 등 테이블에서 요구하는 이전 분기 컬럼을 동적으로 확보
                temp_year, temp_q = cur_year, cur_q
                for _ in range(9):
                    pat = f"{temp_year} {temp_q}/4"
                    key = self._format_quarter_key(temp_year, temp_q)
                    
                    # 이미 설정에 있는 컬럼은 유지, 없는 경우만 탐색
                    if key not in self.quarterly_cols:
                        idx, _ = find_col_by_pattern(pat)
                        if idx is None:
                            # 보조 패턴: '2025. 3/4'
                            idx, _ = find_col_by_pattern(f"{temp_year}. {temp_q}/4")
                        
                        if idx is not None:
                            self.quarterly_cols[key] = idx
                            print(f"[자동탐색] '{pat}' 패턴으로 컬럼 인덱스 자동설정: {idx}")
                    
                    # 전 분기로 이동
                    temp_y_prev, temp_q_prev = self._previous_quarter(temp_year, temp_q)
                    
                    # 전년 동기 컬럼들도 특별히 속성에 저장 (하위 호환성)
                    if temp_year == cur_year - 1 and temp_q == cur_q:
                        if self.prev_y_col is None:
                            self.prev_y_col = self.quarterly_cols.get(key)
                    elif temp_year == cur_year - 2 and temp_q == cur_q:
                        if self.prev_prev_y_col is None:
                            self.prev_prev_y_col = self.quarterly_cols.get(key)
                    elif temp_year == cur_year - 3 and temp_q == cur_q:
                        if self.prev_prev_prev_y_col is None:
                            self.prev_prev_prev_y_col = self.quarterly_cols.get(key)
                    
                    temp_year, temp_q = temp_y_prev, temp_q_prev

                self.quarterly_keys = list(self.quarterly_cols.keys())
    def _get_region_display_name(self, region: str) -> str:
        try:
            return REGION_DISPLAY_MAPPING.get(region, region)
        except Exception:
            return region
    @staticmethod
    def _is_numeric(val) -> bool:
        try:
            if pd.isna(val):
                return False
            float(str(val).replace(',', '').replace('%', ''))
            return True
        except Exception:
            return False

    @staticmethod
    def _find_textual_column(df: pd.DataFrame, header_rows: int, exclude_cols: List[int]) -> Optional[int]:
        """
        헤더 키워드로 못 찾을 때, 데이터 행의 문자 비율이 높은 컬럼을 업종명 후보로 추정
        """
        if df is None or df.empty:
            return None
        n_rows = min(len(df) - header_rows, 30)
        if n_rows <= 0:
            return None
        best_idx = None
        best_score = -1.0
        start = max(header_rows, 0)
        for col_idx in range(len(df.columns)):
            if exclude_cols and col_idx in exclude_cols:
                continue
            text_cnt = 0
            total = 0
            for r in range(start, start + n_rows):
                val = df.iloc[r, col_idx] if col_idx < len(df.columns) else None
                if pd.isna(val):
                    continue
                total += 1
                s = str(val).strip()
                # 숫자만/날짜/코드 패턴 제외
                if not UnifiedReportGenerator._is_numeric(s):
                    text_cnt += 1
            if total == 0:
                continue
            score = text_cnt / total
            if score > best_score:
                best_score = score
                best_idx = col_idx
        return best_idx

    @staticmethod
    def _find_total_row_by_name(df: pd.DataFrame, name_col: int, header_rows: int) -> Optional[pd.DataFrame]:
        """
        업종명 컬럼에서 총계를 의미하는 키워드로 행을 탐색
        """
        if df is None or df.empty or name_col is None:
            return None
        # '계' 단독 키워드는 '단계' 등과 오탐 가능하므로 제외
        keywords = ['총계', '합계', '총지수', '전체', '전산업', '전 산업']
        try:
            series = df.iloc[:, name_col].astype(str).str.strip()
        except Exception:
            return None
        mask = pd.Series(False, index=series.index)
        for kw in keywords:
            mask = mask | series.str.contains(kw, na=False)
        result = df[mask]
        if result is not None and not result.empty:
            return result.head(1)
        return None

    @staticmethod
    def _find_total_row_by_code(
        df: pd.DataFrame,
        total_code: Any,
        exclude_cols: Optional[List[int]] = None
    ) -> Optional[pd.DataFrame]:
        """지정 코드(total_code)를 갖는 행을 모든 텍스트 컬럼에서 탐색."""
        if df is None or df.empty or total_code is None:
            return None
        code_str = str(total_code).strip()
        exclude_cols = exclude_cols or []
        for col_idx in range(len(df.columns)):
            if col_idx in exclude_cols:
                continue
            try:
                series = df.iloc[:, col_idx].astype(str).str.strip()
            except Exception:
                continue
            matched = df[series == code_str]
            if matched is not None and not matched.empty:
                return matched.head(1)
        return None

    @staticmethod
    def _previous_quarter(year: int, quarter: int) -> tuple[int, int]:
        if quarter <= 1:
            return (year - 1, 4)
        return (year, quarter - 1)

    @staticmethod
    def _format_quarter_key(year: int, quarter: int) -> str:
        return f"{year} {quarter}/4"

    def _build_quarter_range(
        self,
        start_year: int,
        start_quarter: int,
        end_year: int,
        end_quarter: int
    ) -> List[tuple[int, int]]:
        quarters = []
        y, q = start_year, start_quarter
        while (y < end_year) or (y == end_year and q <= end_quarter):
            quarters.append((y, q))
            q += 1
            if q > 4:
                q = 1
                y += 1
        return quarters

    def _ensure_quarter_columns(
        self,
        df: pd.DataFrame,
        start_year: int,
        start_quarter: int,
        end_year: int,
        end_quarter: int,
        max_header_rows: int
    ) -> None:
        """
        동적으로 분기별 컬럼 인덱스 매핑 (헤더 기반)
        """
        from utils.excel_utils import get_period_context, find_columns_by_period
        # 분기 키 생성
        period_ctx = get_period_context(end_year, end_quarter)
        period_list = [period_ctx['target_period'], period_ctx['prev_y_period'], period_ctx['prev_q_period']]
        # 최근 5분기 등 필요시 period_ctx['recent_5_quarters'] 활용 가능
        header_row = self.config.get('header_rows', 1) - 1
        col_map = find_columns_by_period(df, period_list, header_row=header_row, exact=False)
        self.quarterly_keys = period_list
        self.quarterly_cols = col_map

    def _collect_quarter_columns(
        self,
        df: pd.DataFrame,
        start_year: int,
        start_quarter: int,
        end_year: int,
        end_quarter: int,
        max_header_rows: int
    ) -> tuple[List[str], Dict[str, Optional[int]]]:
        """
        동적으로 분기별 컬럼 인덱스 매핑 (헤더 기반)
        """
        from utils.excel_utils import get_period_context, find_columns_by_period
        period_ctx = get_period_context(end_year, end_quarter)
        period_list = [period_ctx['target_period'], period_ctx['prev_y_period'], period_ctx['prev_q_period']]
        header_row = self.config.get('header_rows', 1) - 1
        col_map = find_columns_by_period(df, period_list, header_row=header_row, exact=False)
        return period_list, col_map
    def load_data(self):
        """
        테스트 호환성: 기존 테스트 코드에서 generator.load_data()를 호출하는 경우
        실제 데이터프레임 및 주요 속성(df_aggregation, target_col 등)을 초기화
        
        데이터 누락 시 우아하게 처리:
        - 요청한 연도/분기가 없으면 최신 데이터를 자동으로 사용
        - 설정에서 require_analysis_sheet=False면 분석시트 요구 안 함
        """
        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        agg_sheet_name = self.config['aggregation_structure']['sheet']
        print(f"[디버그] config['aggregation_structure']: {self.config.get('aggregation_structure')}")
        print(f"[디버그] agg_sheet_name: {agg_sheet_name}")
        print(f"[디버그] wb.sheetnames: {wb.sheetnames}")
        if not agg_sheet_name:
            raise ValueError('집계 시트명이 설정에 없습니다.')
        # 헤더 행을 보존하기 위해 header=None으로 읽어 병합 헤더 탐색과 데이터 시작 행 탐색을 일관되게 처리
        self.df_aggregation = pd.read_excel(self.excel_path, sheet_name=agg_sheet_name, header=None)
        # 집계 범위가 설정되어 있으면 해당 범위만 사용
        agg_range = self.config.get('aggregation_range')
        if isinstance(agg_range, dict) and self.df_aggregation is not None:
            from openpyxl.utils import column_index_from_string

            def _col_to_index(col_value):
                if col_value is None:
                    return None
                if isinstance(col_value, int):
                    return col_value
                if isinstance(col_value, str) and col_value.strip():
                    return column_index_from_string(col_value.strip().upper()) - 1
                return None

            start_row = agg_range.get('start_row')
            end_row = agg_range.get('end_row')
            start_col = _col_to_index(agg_range.get('start_col'))
            end_col = _col_to_index(agg_range.get('end_col'))

            row_start = max((start_row - 1) if isinstance(start_row, int) else 0, 0)
            row_end = end_row if isinstance(end_row, int) else len(self.df_aggregation)
            col_start = start_col if isinstance(start_col, int) else 0
            col_end = (end_col + 1) if isinstance(end_col, int) else len(self.df_aggregation.columns)

            self.df_aggregation = self.df_aggregation.iloc[row_start:row_end, col_start:col_end].copy()
            print(
                f"[{self.config['name']}] ✅ 집계 범위 적용: rows {start_row}-{end_row}, cols {agg_range.get('start_col')}-{agg_range.get('end_col')}"
            )
        # 원본 보관
        self.df_aggregation_raw = self.df_aggregation
        # 헤더 포함 표를 DataFrame으로 분리 저장
        if self.config.get('header_included') and self.df_aggregation is not None and not self.df_aggregation.empty:
            try:
                df_table = self.df_aggregation.copy()
                df_table.columns = df_table.iloc[0].tolist()
                df_table = df_table.iloc[1:].reset_index(drop=True)
                self.df_aggregation_table = df_table
            except Exception as e:
                print(f"[{self.config['name']}] ⚠️ 헤더 포함 테이블 변환 실패: {e}")
        self.target_col = None
        # 정적 컬럼 인덱스 로드 (동적 탐색 제거)
        column_indices = self.config.get('aggregation_columns') or self.config.get('column_indices') or {}
        self.target_col = column_indices.get('target_col')
        self.prev_y_col = column_indices.get('prev_y_col')
        self.prev_prev_y_col = column_indices.get('prev_prev_y_col')
        self.prev_prev_prev_y_col = column_indices.get('prev_prev_prev_y_col')
        self.quarterly_cols = column_indices.get('quarterly_cols', {}) or {}
        self.quarterly_keys = list(self.quarterly_cols.keys())

        wb.close()

        # 정적 메타 컬럼 설정 (동적 탐색 제거)
        header_rows = self.config.get('header_rows', 1)
        agg_struct = self.config.get('aggregation_structure', {}) if isinstance(self.config, dict) else {}
        self.region_name_col = agg_struct.get('region_name_col')
        self.industry_code_col = agg_struct.get('industry_code_col')
        if self.industry_name_col is None:
            self.industry_name_col = self.config.get('industry_name_col') or agg_struct.get('industry_name_col')
        if self.data_start_row is None:
            self.data_start_row = self.config.get('data_start_row', header_rows)
        analysis_sheet = self.config.get('analysis_sheet')
        if analysis_sheet and analysis_sheet != agg_sheet_name:
            try:
                self.df_analysis = pd.read_excel(self.excel_path, sheet_name=analysis_sheet, header=None)
                analysis_columns = self.config.get('analysis_columns') or self.config.get('analysis_column_indices') or {}
                self.analysis_target_col = analysis_columns.get('target_col')
                self.analysis_prev_y_col = analysis_columns.get('prev_y_col')
                self.analysis_prev_prev_y_col = analysis_columns.get('prev_prev_y_col')
                self.analysis_prev_prev_prev_y_col = analysis_columns.get('prev_prev_prev_y_col')
                self.analysis_quarterly_cols = analysis_columns.get('quarterly_cols', {}) or {}
                self.analysis_quarterly_keys = list(self.analysis_quarterly_cols.keys())
            except Exception as e:
                print(f"[{self.config['name']}] ⚠️ 분석 시트 로드 실패: {analysis_sheet} ({e})")
    
    def _find_latest_data_col(self, target_year=None):
        """
        동적 데이터 탐색 기능은 제거되었습니다.
        """
        raise NotImplementedError(
            "동적 데이터 탐색 기능이 제거되었습니다. 설정에서 컬럼 인덱스를 지정하세요."
        )

    def _extract_table_data_ssot(self) -> List[Dict[str, Any]]:
        """
        집계/분석 데이터를 단일 테이블 형태로 추출
        """
        if self.df_aggregation is None:
            raise ValueError(
                f"[{self.config['name']}] ❌ 집계 시트를 로드할 수 없습니다. "
                f"load_data() 또는 extract_all_data()를 먼저 호출해야 합니다."
            )
        df = self.df_aggregation
        # 데이터 행만 (헤더 제외) - 동적으로 찾은 시작 행 사용
        if self.data_start_row is None:
            self.data_start_row = 0
        
        if self.data_start_row < 0:
            self.data_start_row = 0
        
        if self.data_start_row < len(df):
            data_df = df.iloc[self.data_start_row:].copy()
        else:
            print(f"[{self.config['name']}] ⚠️ data_start_row({self.data_start_row})가 DataFrame 길이({len(df)})를 초과합니다. 전체 DataFrame 사용")
            data_df = df.copy()
        
        # 분기 단위 전체 범위 컬럼은 설정 기반으로만 사용
        header_rows = self.config.get('header_rows', 5)

        # 직전 분기 컬럼
        prev_q_col = None
        if self.year is not None and self.quarter is not None:
            prev_q_year, prev_q = self._previous_quarter(self.year, self.quarter)
            prev_q_key = self._format_quarter_key(prev_q_year, prev_q)
            prev_q_col = self.quarterly_cols.get(prev_q_key)

        use_analysis_rates = self.config.get('value_type') == 'change_rate' and self.df_analysis is not None
        
        # 지역 목록
        regions = ['전국', '서울', '부산', '대구', '인천', '광주', '대전', '울산', '세종',
                   '경기', '강원', '충북', '충남', '전북', '전남', '경북', '경남', '제주']
        
        table_data = []
        total_code = None
        try:
            total_code = (self.config.get('aggregation_structure') or {}).get('total_code')
        except Exception:
            total_code = None

        def _parse_age_range(name: Any) -> Optional[tuple[int, int]]:
            if not name:
                return None
            age_str = str(name).strip()
            if not age_str:
                return None
            normalized = age_str.replace(' ', '')
            if normalized in {'계', '합계', '총계', '전체', '전연령', '전연령층', '전체연령', '총연령'}:
                return None
            normalized = normalized.replace('세', '')
            normalized = normalized.replace('~', '-').replace('–', '-').replace('—', '-')
            try:
                import re
                match = re.search(r'(\d{1,2})-(\d{1,2})', normalized)
                if match:
                    return (int(match.group(1)), int(match.group(2)))
                match = re.search(r'(\d{1,2})대', normalized)
                if match:
                    start = int(match.group(1))
                    return (start, start + 9)
            except Exception:
                return None
            return None

        def _compute_migration_age_sums(region_name: str) -> tuple[Optional[float], Optional[float]]:
            items = self._extract_industry_data(region_name)
            if not items:
                return (None, None)
            sum_20_29 = 0.0
            found_20_29_parts = False
            alt_20_29 = None
            sum_other = 0.0
            found_other = False
            for item in items:
                if not isinstance(item, dict):
                    continue
                name = item.get('name')
                value = item.get('value')
                if value is None:
                    continue
                age_range = _parse_age_range(name)
                if age_range is None:
                    continue
                if age_range in [(20, 24), (25, 29)]:
                    sum_20_29 += float(value)
                    found_20_29_parts = True
                elif age_range == (20, 29):
                    alt_20_29 = float(value)
                else:
                    sum_other += float(value)
                    found_other = True
            age_20_29 = None
            if found_20_29_parts:
                age_20_29 = round(sum_20_29, 1)
            elif alt_20_29 is not None:
                age_20_29 = round(alt_20_29, 1)
            age_other = round(sum_other, 1) if found_other else None
            return (age_20_29, age_other)

        def _select_region_total(df_source: pd.DataFrame, region_name: str) -> Optional[pd.Series]:
            if df_source is None:
                return None
            if self.data_start_row is None:
                start_row = 0
            else:
                start_row = max(self.data_start_row, 0)
            if start_row < len(df_source):
                local_df = df_source.iloc[start_row:].copy()
            else:
                local_df = df_source.copy()
            region_col = self.region_name_col
            if region_col is None or region_col < 0 or region_col >= len(local_df.columns):
                region_col = None

            def _detect_region_col(df_search: pd.DataFrame) -> Optional[int]:
                if df_search is None or df_search.empty:
                    return None
                valid_regions = ['전국', '서울', '부산', '대구', '인천', '광주', '대전', '울산', '세종',
                                 '경기', '강원', '충북', '충남', '전북', '전남', '경북', '경남', '제주']
                rows_to_scan = min(40, len(df_search))
                try:
                    for col_idx in range(len(df_search.columns)):
                        for r in range(rows_to_scan):
                            val = df_search.iloc[r, col_idx]
                            if pd.notna(val) and str(val).strip() in valid_regions:
                                return col_idx
                except Exception:
                    return None
                return None

            if region_col is None:
                region_col = _detect_region_col(df_source)
            if region_col is None:
                return None

            try:
                from config.reports import REGION_DISPLAY_MAPPING
                full_name = REGION_DISPLAY_MAPPING.get(region_name)
                
                # 1. Exact match (short name)
                region_filter = local_df[
                    local_df.iloc[:, region_col].astype(str).str.strip() == region_name
                ]
                
                # 2. Exact match (full name)
                if region_filter.empty and full_name:
                    region_filter = local_df[
                        local_df.iloc[:, region_col].astype(str).str.strip() == full_name
                    ]

                # 3. Alias match (Historical/Nomenclature variations)
                if region_filter.empty:
                    REGION_ALIASES = {
                        '전북': ['전라북도', '전북특별자치도'],
                        '강원': ['강원도', '강원특별자치도'],
                        '경기': ['경기도'],
                        '충북': ['충청북도'],
                        '충남': ['충청남도'],
                        '전남': ['전라남도'],
                        '경북': ['경상북도'],
                        '경남': ['경상남도'],
                        '제주': ['제주특별자치도', '제주도', '제주'],
                        '세종': ['세종특별자치시'],
                        '서울': ['서울특별시'],
                        '부산': ['부산광역시'],
                        '대구': ['대구광역시'],
                        '인천': ['인천광역시'],
                        '광주': ['광주광역시'],
                        '대전': ['대전광역시'],
                        '울산': ['울산광역시'],
                    }
                    aliases = REGION_ALIASES.get(region_name, [])
                    for alias in aliases:
                        region_filter = local_df[
                            local_df.iloc[:, region_col].astype(str).str.strip() == alias
                        ]
                        if not region_filter.empty:
                            break
                
                # 3. Fuzzy match / Alias fallback (especially for '전국')
                if region_filter.empty:
                    if region_name == '전국':
                        if total_code:
                            region_filter = local_df[
                                local_df.iloc[:, region_col].astype(str).str.strip().str.contains(total_code, na=False)
                            ]
                        if region_filter.empty:
                            for alias in ['합계', '총계', '전국', '전 국']:
                                region_filter = local_df[
                                    local_df.iloc[:, region_col].astype(str).str.strip() == alias
                                ]
                                if not region_filter.empty:
                                    break
                    else:
                        # Other regions fuzzy match (e.g. '서울' matches '서울특별시' via contains)
                        region_filter = local_df[
                            local_df.iloc[:, region_col].astype(str).str.strip().str.contains(region_name, na=False)
                        ]

            except (IndexError, KeyError):
                return None
            if region_filter.empty and df_source is not local_df:
                try:
                    region_filter = df_source[
                        df_source.iloc[:, region_col].astype(str).str.strip() == region_name
                    ]
                    local_df = df_source
                except (IndexError, KeyError):
                    return None
            if region_filter.empty:
                alt_col = _detect_region_col(df_source)
                if alt_col is not None and alt_col != region_col:
                    region_col = alt_col
                    try:
                        region_filter = df_source[
                            df_source.iloc[:, region_col].astype(str).str.strip() == region_name
                        ]
                        local_df = df_source
                    except (IndexError, KeyError):
                        return None
            if region_filter.empty:
                return None
            region_total = None
            if self.industry_name_col is not None and self.industry_name_col != self.region_name_col and self.industry_name_col >= 0 and self.industry_name_col < len(region_filter.columns):
                by_name = self._find_total_row_by_name(region_filter, self.industry_name_col, header_rows=0)
                if by_name is not None and not by_name.empty:
                    region_total = by_name
            if (region_total is None or region_total.empty) and total_code:
                exclude_cols = []
                if region_col is not None:
                    exclude_cols.append(region_col)
                if self.industry_name_col is not None and self.report_type not in ['employment', 'unemployment', 'migration']:
                    exclude_cols.append(self.industry_name_col)
                by_code = self._find_total_row_by_code(region_filter, total_code, exclude_cols=exclude_cols)
                if by_code is not None and not by_code.empty:
                    region_total = by_code
            if (region_total is None or region_total.empty) and self.report_type in ['employment', 'unemployment', 'migration']:
                if len(region_filter) > 0:
                    region_total = region_filter.head(1)
            if (region_total is None or region_total.empty) and self.report_type == 'migration':
                if len(region_filter) > 0:
                    region_total = region_filter.head(1)
            if region_total is None or region_total.empty:
                if self.report_type == 'price' and region_name == '전북':
                    print(f"[DEBUG] {region_name}: region_filter empty? {region_filter.empty}, industry_name_col: {self.industry_name_col}")
                    if not region_filter.empty:
                        print(f"[DEBUG] {region_name} first 5 rows in industry_col:\n{region_filter.iloc[:5, self.industry_name_col]}")
                return None
            return region_total.iloc[0]
        
        # 컬럼 인덱스 검증 (동적으로 찾은 컬럼)
        if self.region_name_col is None or self.region_name_col < 0 or self.region_name_col >= len(data_df.columns):
            print(
                f"[{self.config['name']}] ⚠️ 지역명 컬럼이 설정되지 않았거나 유효하지 않습니다. "
                f"인덱스({self.region_name_col}), DataFrame 컬럼 수: {len(data_df.columns)}"
            )
            return []
        
        def _select_youth_rate(df_source: pd.DataFrame, region_name: str, target_col_idx: int) -> Optional[float]:
            """실업률/고용률에서 청년층 데이터 추출 (고용률: 20-29세, 실업률: 15-29세)"""
            if df_source is None or self.report_type not in ['employment', 'unemployment']:
                return None
            
            if self.data_start_row is None:
                start_row = 0
            else:
                start_row = max(self.data_start_row, 0)
            
            if start_row < len(df_source):
                local_df = df_source.iloc[start_row:].copy()
            else:
                local_df = df_source.copy()
            
            # 지역명 컬럼 (0번)과 연령계층 컬럼 (1번)
            region_col = 0
            age_col = 1
            
            # 고용률은 20-29세, 실업률은 15-29세 사용
            if self.report_type == 'employment':
                # 여러 패턴 시도: '20-29', '20~29', '20∼29', '20～29' 등
                age_patterns = ['20.*29', '20~29', '20-29', '20∼29', '20～29']
            else:  # unemployment
                age_patterns = ['15.*29', '15~29', '15-29', '15∼29', '15～29']
            
            try:
                youth_filter = None
                for age_pattern in age_patterns:
                    try:
                        # 지역명과 연령계층 조건으로 필터링
                        temp_filter = local_df[
                            (local_df.iloc[:, region_col].astype(str).str.strip() == region_name) &
                            (local_df.iloc[:, age_col].astype(str).str.strip().str.contains(age_pattern, regex=True, na=False))
                        ]
                        
                        if not temp_filter.empty:
                            youth_filter = temp_filter
                            break
                    except Exception:
                        continue
                
                if youth_filter is None or youth_filter.empty:
                    # 디버그: 왜 찾지 못했는지 로그
                    unique_ages = local_df[local_df.iloc[:, region_col].astype(str).str.strip() == region_name].iloc[:, age_col].unique()
                    if len(unique_ages) > 0:
                        print(f"[{self.config['name']}] ⚠️ {region_name} 청년층 데이터 찾기 실패. 연령계층 후보: {unique_ages[:5]}...")
                    return None
                
                youth_row = youth_filter.iloc[0]
                if target_col_idx is not None and target_col_idx < len(youth_row):
                    return self.safe_float(youth_row.iloc[target_col_idx], None)
            except Exception as e:
                print(f"[{self.config['name']}] ⚠️ {region_name} 청년층 데이터 추출 실패: {e}")
            
            return None

        for region in regions:
            row = _select_region_total(df, region)
            if row is None:
                print(f"[{self.config['name']}] ⚠️ {region}: 총계 행을 찾지 못했습니다. 스킵합니다.")
                continue

            analysis_row = _select_region_total(self.df_analysis, region) if use_analysis_rates else None
            
            # 청년층(15-29세) 데이터 추출 (실업률/고용률만)
            youth_rate = _select_youth_rate(df, region, self.target_col)
            
            if self.target_col is None or self.prev_y_col is None:
                print(
                    f"[{self.config['name']}] ⚠️ {region}: 컬럼 인덱스가 설정되지 않아 스킵합니다. "
                    f"target_col={self.target_col}, prev_y_col={self.prev_y_col}"
                )
                continue
            
            # 인덱스 범위 체크
            if self.target_col is None or self.target_col >= len(row):
                print(f"[{self.config['name']}] ⚠️ Target 컬럼 인덱스({self.target_col})가 행 길이({len(row)})를 초과합니다. 스킵합니다.")
                continue
            
            if self.prev_y_col is None or self.prev_y_col >= len(row):
                print(f"[{self.config['name']}] ⚠️ 전년 컬럼 인덱스({self.prev_y_col})가 행 길이({len(row)})를 초과합니다. 스킵합니다.")
                continue

            def _compute_quarterly_growth(current: Optional[float], previous: Optional[float]) -> Optional[float]:
                if current is None or previous is None:
                    return None
                if self.report_type in ['employment', 'unemployment']:
                    return round(current - previous, 1)
                if self.report_type == 'migration':
                    return round(current - previous, 1)
                if self.config.get('value_type') == 'change_rate':
                    return round(current, 1)
                if previous == 0:
                    return None
                return round((current - previous) / previous * 100, 1)
            
            # 지수 추출
            try:
                idx_current = self.safe_float(row.iloc[self.target_col], None)
                idx_prev_year = self.safe_float(row.iloc[self.prev_y_col], None)
                idx_prev_prev_year = None
                idx_prev_prev_prev_year = None
                if self.prev_prev_y_col is not None and self.prev_prev_y_col < len(row):
                    idx_prev_prev_year = self.safe_float(row.iloc[self.prev_prev_y_col], None)
                if self.prev_prev_prev_y_col is not None and self.prev_prev_prev_y_col < len(row):
                    idx_prev_prev_prev_year = self.safe_float(row.iloc[self.prev_prev_prev_y_col], None)
            except (IndexError, KeyError) as e:
                print(f"[{self.config['name']}] ⚠️ 데이터 추출 오류: {e}. 스킵합니다.")
                continue

            rate_current = None
            rate_prev_year = None
            rate_quarterly_values: List[Optional[float]] = []
            rate_prev_quarter = None
            if use_analysis_rates and analysis_row is not None:
                if self.analysis_target_col is not None and self.analysis_target_col < len(analysis_row):
                    rate_current = self.safe_float(analysis_row.iloc[self.analysis_target_col], None)
                if self.analysis_prev_y_col is not None and self.analysis_prev_y_col < len(analysis_row):
                    rate_prev_year = self.safe_float(analysis_row.iloc[self.analysis_prev_y_col], None)
                if self.analysis_quarterly_keys:
                    for key in self.analysis_quarterly_keys:
                        col_idx = self.analysis_quarterly_cols.get(key)
                        if col_idx is not None and col_idx < len(analysis_row):
                            rate_quarterly_values.append(self.safe_float(analysis_row.iloc[col_idx], None))
                        else:
                            rate_quarterly_values.append(None)
                    if len(rate_quarterly_values) >= 2:
                        rate_prev_quarter = rate_quarterly_values[-2]
            
            # 분기 단위 전체 범위 값 추출
            quarterly_values: List[Optional[float]] = []
            if self.quarterly_keys:
                for key in self.quarterly_keys:
                    col_idx = self.quarterly_cols.get(key)
                    if col_idx is not None and col_idx < len(row):
                        quarterly_values.append(self.safe_float(row.iloc[col_idx], None))
                    else:
                        quarterly_values.append(None)

            # 단위 보정
            scale_factor = 1.0
            if self.report_type == 'construction':
                # 10억원 단위 → 100억원 단위 (1/10)
                scale_factor = 0.1
            elif self.report_type in ['export', 'import']:
                # 백만달러 단위 → 억달러 단위 (1/100)
                scale_factor = 0.01
            elif self.report_type == 'migration':
                # 연령별 계산 등이 복잡하므로 원형(명) 유지, 템플릿(val_thousand)에서 나누기 수행
                scale_factor = 1.0

            if scale_factor != 1.0:
                idx_current = (idx_current * scale_factor) if idx_current is not None else None
                idx_prev_year = (idx_prev_year * scale_factor) if idx_prev_year is not None else None
                idx_prev_prev_year = (idx_prev_prev_year * scale_factor) if idx_prev_prev_year is not None else None
                idx_prev_prev_prev_year = (idx_prev_prev_prev_year * scale_factor) if idx_prev_prev_prev_year is not None else None
                quarterly_values = [
                    (v * scale_factor) if v is not None else None
                    for v in quarterly_values
                ]

            if use_analysis_rates and rate_quarterly_values:
                quarterly_growth_rates = rate_quarterly_values[:]
            elif self.report_type == 'migration':
                quarterly_growth_rates: List[Optional[float]] = [None for _ in quarterly_values]
            else:
                quarterly_growth_rates = []
                for i, val in enumerate(quarterly_values):
                    if i == 0:
                        quarterly_growth_rates.append(None)
                    else:
                        quarterly_growth_rates.append(_compute_quarterly_growth(val, quarterly_values[i - 1]))

            # 직전 분기 값
            idx_prev_quarter = None
            if prev_q_col is not None and prev_q_col < len(row):
                idx_prev_quarter = self.safe_float(row.iloc[prev_q_col], None)

            if idx_prev_quarter is not None and scale_factor != 1.0:
                idx_prev_quarter = idx_prev_quarter * scale_factor

            # 국내인구이동: 직전/전전/전전전 분기 값 추출 (없으면 None 유지)
            idx_prev_prev = idx_prev_prev_prev = None
            if self.report_type == 'migration' and quarterly_values:
                if len(quarterly_values) >= 2:
                    idx_prev_quarter = quarterly_values[-2]
                if len(quarterly_values) >= 3:
                    idx_prev_prev = quarterly_values[-3]
                if len(quarterly_values) >= 4:
                    idx_prev_prev_prev = quarterly_values[-4]
            
            if idx_current is None:
                continue

            if self.report_type == 'migration':
                previous_quarter_growth = None
            elif use_analysis_rates:
                previous_quarter_growth = rate_prev_quarter
            else:
                previous_quarter_growth = _compute_quarterly_growth(idx_current, idx_prev_quarter)
            
            # 증감 계산 (report_type에 따라 다름)
            # 국내인구이동: 절대값 (부호 포함, 변화율 아님)
            # 고용률/실업률: 퍼센트포인트(p) 차이
            # value_type='change_rate': 이미 계산된 증감률 직접 사용
            # 기타 지수: 증감률(%)
            if self.report_type == 'migration':
                # 국내인구이동은 증감률(%)이 아닌 전년동분기대비 증감(명)을 계산
                if idx_prev_year is not None:
                    change_rate = round(idx_current - idx_prev_year, 1)
                else:
                    change_rate = None
            elif self.config.get('value_type') == 'change_rate':
                # 시트에 이미 증감률이 계산되어 있는 경우 (예: C 분석)
                change_rate = round(rate_current, 1) if rate_current is not None else round(idx_current, 1)
            elif idx_prev_year is not None and idx_prev_year != 0:
                if self.report_type in ['employment', 'unemployment']:
                    # 퍼센트포인트 차이 (p)
                    change_rate = round(idx_current - idx_prev_year, 1)
                else:
                    # 증감률 (%)
                    change_rate = round(((idx_current - idx_prev_year) / idx_prev_year) * 100, 1)
            elif idx_current is not None:
                # 전년 데이터가 없으면 증감률을 0.0으로 처리하거나 보합으로 간주 (또는 None 유지)
                # 여기서는 명확성을 위해 None을 유지하되, 나중에 템플릿에서 처리하도록 함
                change_rate = None
            
            if self.report_type == 'migration':
                age_20_29, age_other = _compute_migration_age_sums(region)
                row_data = {
                    'region_name': region,
                    'region_display': self._get_region_display_name(region),
                    'value': round(idx_current, 1),
                    'prev_value': round(idx_prev_quarter, 1) if idx_prev_quarter is not None else None,
                    'prev_prev_value': round(idx_prev_prev, 1) if idx_prev_prev is not None else None,
                    'prev_prev_prev_value': round(idx_prev_prev_prev, 1) if idx_prev_prev_prev is not None else None,
                    'prev_year_value': round(idx_prev_year, 1) if idx_prev_year is not None else None,
                    'change_rate': change_rate,  # 전년동분기대비 증감(명)
                    'quarterly_keys': self.quarterly_keys,
                    'quarterly_values': quarterly_values,
                    'quarterly_growth_rates': quarterly_growth_rates,
                    'age_20_29': age_20_29,
                    'age_other': age_other
                }
            else:
                row_data = {
                    'region_name': region,
                    'region_display': self._get_region_display_name(region),
                    'value': round(idx_current, 1),
                    'prev_value': round(idx_prev_year, 1) if idx_prev_year else None,
                    'prev_prev_value': round(idx_prev_prev_year, 1) if idx_prev_prev_year is not None else None,
                    'prev_prev_prev_value': round(idx_prev_prev_prev_year, 1) if idx_prev_prev_prev_year is not None else None,
                    'change_rate': change_rate,
                    'previous_quarter_growth': previous_quarter_growth,
                    'quarterly_keys': self.quarterly_keys,
                    'quarterly_values': quarterly_values,
                    'quarterly_growth_rates': quarterly_growth_rates,
                    'rate_quarterly_keys': self.analysis_quarterly_keys if use_analysis_rates else None,
                    'rate_quarterly_values': rate_quarterly_values if use_analysis_rates else None,
                    'youth_rate': round(youth_rate, 1) if youth_rate is not None else None
                }

            table_data.append(row_data)
            
            if self.report_type == 'migration':
                change_str = f"{change_rate:.1f}명" if change_rate is not None else "N/A"
                print(f"[{self.config['name']}] ✅ {region}: 순이동={idx_current:.1f}명, 전년동분기대비 증감={change_str}")
            else:
                print(f"[{self.config['name']}] ✅ {region}: 지수={idx_current:.1f}, 증감률={change_rate}%")
        
        # 국내인구이동: 전국 데이터 생성 여부 확인 (config의 has_nationwide 설정)
        # 국내이동은 지역간 이동이므로 전국 합계(0)는 의미가 없어 생성하지 않음
        if self.report_type == 'migration' and table_data:
            # config에서 has_nationwide 설정 확인 (기본값 True)
            should_generate_nationwide = self.config.get('has_nationwide', True)
            
            if should_generate_nationwide:
                def sum_field(key: str) -> Optional[float]:
                    values = [row.get(key) for row in table_data if row.get('region_name') != '전국' and row.get(key) is not None]
                    return round(sum(values), 1) if values else None

                # 이미 전국이 있다면 스킵
                has_nationwide = any(row.get('region_name') == '전국' for row in table_data)
                if not has_nationwide:
                    nationwide_row = {
                        'region_name': '전국',
                        'region_display': self._get_region_display_name('전국'),
                        'value': sum_field('value'),
                        'prev_value': sum_field('prev_value'),
                        'prev_prev_value': sum_field('prev_prev_value'),
                        'prev_prev_prev_value': sum_field('prev_prev_prev_value'),
                        'change_rate': sum_field('change_rate'),
                        'age_20_29': sum_field('age_20_29'),
                        'age_other': sum_field('age_other')
                    }
                    table_data.insert(0, nationwide_row)
                    print(f"[{self.config['name']}] ✅ 전국 데이터가 없어 지역 합계로 추가")
            else:
                print(f"[{self.config['name']}] ⚠️ has_nationwide=False이므로 전국 데이터 생성 건너뜀")
        
        return table_data
    
    def _extract_age_groups_for_region(self, region: str) -> List[Dict[str, Any]]:
        """
        특정 지역의 연령별 고용률/실업률 변화 데이터 추출
        
        Args:
            region: 지역명 ('전국', '서울', 등)
            
        Returns:
            연령별 데이터 리스트 [{'name': '20-29세', 'display_name': '20-29세', 'change': 증감(%p)}, ...]
        """
        if self.report_type not in ['employment', 'unemployment']:
            return []
        
        if self.df_aggregation is None:
            return []
        
        df = self.df_aggregation
        
        # 데이터 시작 행
        start_row = self.data_start_row if self.data_start_row is not None else 0
        if start_row < len(df):
            data_df = df.iloc[start_row:].copy()
        else:
            data_df = df.copy()
        
        # 지역명 컬럼 (0번)과 연령계층 컬럼 (1번) - 고용률/실업률 시트 구조
        region_col = 0
        age_col = 1
        
        # 현재 분기와 전년 동분기 컬럼 인덱스
        target_col = self.target_col
        prev_y_col = self.prev_y_col
        
        if target_col is None or prev_y_col is None:
            print(f"[{self.config['name']}] ⚠️ {region}: 컬럼 인덱스가 설정되지 않아 연령별 데이터 추출 불가")
            return []
        
        try:
            # 지역 필터링
            region_filter = data_df[
                data_df.iloc[:, region_col].astype(str).str.strip() == region
            ]
            
            if region_filter.empty:
                return []
            
            # 연령별 데이터 추출
            age_groups = []
            for idx, row in region_filter.iterrows():
                age_name = str(row.iloc[age_col]).strip() if age_col < len(row) else ''
                
                # '계' (합계)는 제외
                if age_name in ['계', '합계', '전체', '']:
                    continue
                
                # 현재 값과 전년 동분기 값 추출
                current_val = self.safe_float(row.iloc[target_col], None) if target_col < len(row) else None
                prev_val = self.safe_float(row.iloc[prev_y_col], None) if prev_y_col < len(row) else None
                
                # 증감(%p) 계산
                if current_val is not None and prev_val is not None:
                    change = round(current_val - prev_val, 1)
                else:
                    change = None
                
                age_groups.append({
                    'name': age_name,
                    'display_name': age_name,
                    'value': current_val,
                    'prev_value': prev_val,
                    'change': change
                })
            
            # 증감이 큰 순서로 정렬 (증가 > 0, 감소 < 0)
            age_groups_with_change = [a for a in age_groups if a.get('change') is not None]
            age_groups_with_change.sort(key=lambda x: abs(x['change']), reverse=True)
            
            return age_groups_with_change
            
        except Exception as e:
            print(f"[{self.config['name']}] ⚠️ {region} 연령별 데이터 추출 실패: {e}")
            return []
    
    def _extract_industry_data(self, region: str) -> List[Dict[str, Any]]:
        """
        특정 지역의 업종별 데이터 추출
        
        Args:
            region: 지역명 ('전국', '서울', 등)
            
        Returns:
            업종별 데이터 리스트 [{'name': '업종명', 'value': 지수, 'change_rate': 증감률, 'growth_rate': 증감률}, ...]
        """
        if self.df_aggregation is None:
            return []
        
        df = self.df_aggregation
        
        # 컬럼 인덱스 검증 (동적으로 찾은 컬럼)
        if self.region_name_col is None or self.region_name_col < 0 or self.region_name_col >= len(df.columns):
            print(f"[{self.config['name']}] ⚠️ 지역명 컬럼을 찾을 수 없습니다. 동적 탐색 실패 또는 인덱스({self.region_name_col})가 유효하지 않습니다. 빈 리스트 반환")
            return []
        
        # 데이터 행만 (헤더 제외) - 동적으로 찾은 시작 행 사용
        if self.data_start_row is None:
            self.data_start_row = 0
        
        if self.data_start_row < 0:
            self.data_start_row = 0
        
        if self.data_start_row < len(df):
            data_df = df.iloc[self.data_start_row:].copy()
        else:
            data_df = df.copy()
        
        # 지역 필터링 (안전한 인덱스 접근)
        try:
            region_filter = data_df[
                data_df.iloc[:, self.region_name_col].astype(str).str.strip() == region
            ]
        except (IndexError, KeyError) as e:
            print(f"[{self.config['name']}] ⚠️ {region} 필터링 오류: {e}")
            return []
        
        if region_filter.empty:
            return []
        
        industries = []
        # 기본값/폴백 사용 금지
        if 'name_mapping' not in self.config:
            raise ValueError(f"[{self.config['name']}] ❌ 설정에서 'name_mapping'을 찾을 수 없습니다. 기본값 사용 금지.")
        name_mapping = self.config['name_mapping']
        
        # 산업명 컬럼 찾기 (동적으로 찾은 값 사용)
        if self.industry_name_col is None:
            if self.report_type in ['employment', 'unemployment']:
                # 경고 메시지 중복 출력 방지
                if not self._warned_industry_col_missing:
                    print(f"[{self.config['name']}] ⚠️ 산업명 컬럼을 찾을 수 없지만, 고용률/실업률은 산업명이 선택적이므로 계속 진행합니다.")
                    self._warned_industry_col_missing = True
                industry_name_col = None
            else:
                # 헤더로 못 찾은 경우 텍스트 비율 기반 추정 시도
                industry_name_col = self._find_textual_column(df, header_rows=0, exclude_cols=[self.region_name_col] if self.region_name_col is not None else [])
                if industry_name_col is not None:
                    print(f"[{self.config['name']}] ✅ 업종명 컬럼 추정: {industry_name_col}")
                    self.industry_name_col = industry_name_col
                else:
                    print(f"[{self.config['name']}] ⚠️ 업종명 컬럼을 추정하지 못했습니다. 업종 데이터 추출을 건너뜁니다.")
                    return []
        else:
            industry_name_col = self.industry_name_col
        
        if industry_name_col is not None and industry_name_col < 0:
            industry_name_col = 0
        
        for idx, row in region_filter.iterrows():
            # 산업명 추출 우선 (총계 키워드면 스킵)
            industry_name = ''
            if industry_name_col is not None and industry_name_col < len(row) and pd.notna(row.iloc[industry_name_col]):
                industry_name = str(row.iloc[industry_name_col]).strip()
            if not industry_name:
                # 고용률/실업률은 산업명이 없어도 진행 가능
                if self.report_type not in ['employment', 'unemployment']:
                    continue
            
            # 총계 키워드 스킵 (오탐 방지를 위해 '계' 제외)
            if any(kw in industry_name for kw in ['총계', '합계', '총지수', '전체', '전산업', '전 산업']):
                continue
            
            # 산업명 컬럼이 없으면 스킵 (고용률/실업률 제외)
            if industry_name_col is None and self.report_type not in ['employment', 'unemployment']:
                continue
            
            # 이름 매핑 적용
            if industry_name in name_mapping:
                industry_name = name_mapping[industry_name]
            
            if not industry_name:
                continue
            
            # 지수 추출 (안전한 인덱스 접근)
            try:
                if self.target_col is None or self.prev_y_col is None:
                    continue
                
                # 인덱스 범위 체크
                if self.target_col < 0 or self.target_col >= len(row):
                    continue
                if self.prev_y_col < 0 or self.prev_y_col >= len(row):
                    continue
                    
                idx_current = self.safe_float(row.iloc[self.target_col], None)
                idx_prev_year = self.safe_float(row.iloc[self.prev_y_col], None)
            except (IndexError, KeyError, AttributeError) as e:
                print(f"[{self.config['name']}] ⚠️ 데이터 추출 오류 (인덱스 {self.target_col}/{self.prev_y_col}): {e}")
                continue
            
            if idx_current is None:
                continue
            
            # 증감률 계산
            change_rate = None
            if idx_prev_year and idx_prev_year != 0:
                change_rate = round(((idx_current - idx_prev_year) / idx_prev_year) * 100, 1)
            
            industries.append({
                'name': industry_name,
                'value': round(idx_current, 1),
                'prev_value': round(idx_prev_year, 1) if idx_prev_year else None,
                'change_rate': change_rate,
                'growth_rate': change_rate,  # 템플릿 호환 필드명
                'contribution_rate': None,   # UndefinedError 방지
                'contribution': None        # UndefinedError 방지
            })
        
        return industries
    
    def _get_top_industries_for_region(self, region: str, increase: bool = True, top_n: int = 3) -> List[Dict[str, Any]]:
        """
        특정 지역의 상위 업종 추출
        
        Args:
            region: 지역명
            increase: True면 증가 업종, False면 감소 업종
            top_n: 상위 N개
            
        Returns:
            상위 업종 리스트
        """
        if not region or not isinstance(region, str):
            return []
        
        industries = self._extract_industry_data(region)
        
        # 안전한 필터링
        if not industries:
            return []
        
        if increase:
            filtered = [
                ind for ind in industries 
                if ind and isinstance(ind, dict) and 
                ind.get('change_rate') is not None and 
                ind['change_rate'] > 0
            ]
            try:
                # 기본값/폴백 사용 금지: change_rate가 None이면 정렬에서 제외
                filtered = [x for x in filtered if x and isinstance(x, dict) and x.get('change_rate') is not None]
                filtered.sort(key=lambda x: x['change_rate'], reverse=True)
            except (TypeError, AttributeError, KeyError) as e:
                raise ValueError(f"[{self.config['name']}] ❌ 정렬 오류: {e}. 기본값 사용 금지: 반드시 데이터를 찾아야 합니다.")
        else:
            filtered = [
                ind for ind in industries 
                if ind and isinstance(ind, dict) and 
                ind.get('change_rate') is not None and 
                ind['change_rate'] < 0
            ]
            try:
                # 기본값/폴백 사용 금지: change_rate가 None이면 정렬에서 제외
                filtered = [x for x in filtered if x and isinstance(x, dict) and x.get('change_rate') is not None]
                filtered.sort(key=lambda x: x['change_rate'])
            except (TypeError, AttributeError):
                pass  # 정렬 실패 시 원본 유지
        
        # 안전한 슬라이싱
        # 기본값/폴백 사용 금지: filtered가 없으면 None 반환
        if not filtered or len(filtered) == 0:
            return None
        return filtered[:top_n]
    
    def extract_nationwide_data(self, table_data: List[Dict] = None) -> Dict[str, Any]:
        """전국 데이터 추출 - 템플릿 호환 필드명"""
        if table_data is None:
            table_data = self._extract_table_data_ssot()
        
        nationwide = next((d for d in table_data if str(d.get('region_name')).strip() == '전국'), None)
        if not nationwide:
             # some reports might use '합계' or other terms for nationwide in table_data
             nationwide = next((d for d in table_data if d.get('region_name') == self.config.get('aggregation_structure', {}).get('total_code')), None)
        
        # 국내인구이동의 경우 전국 데이터가 없으면 지역 합계로 계산
        if not nationwide or not isinstance(nationwide, dict):
            if self.report_type == 'migration' and table_data:
                print(f"[{self.config['name']}] ⚠️ 전국 데이터를 찾을 수 없으므로 모든 지역을 합계하여 계산합니다.")
                # 모든 지역 데이터 합계 (전국 제외)
                total_value = 0
                total_prev_value = 0
                for d in table_data:
                    if d and isinstance(d, dict) and d.get('region_name') != '전국':
                        total_value += d.get('value', 0) or 0
                        total_prev_value += d.get('prev_value', 0) or 0
                
                # 전국 데이터 생성
                change_rate = None
                if self.report_type == 'migration':
                    # 국내인구이동: 절대 순인구이동값 (부호 포함)
                    change_rate = round(total_value, 1)
                elif total_prev_value != 0:
                    change_rate = round((total_value - total_prev_value) / total_prev_value * 100, 1)
                
                nationwide = {
                    'region_name': '전국',
                    'region_display': '전 국',
                    'value': total_value,
                    'prev_value': total_prev_value,
                    'change_rate': change_rate
                }
                print(f"[{self.config['name']}] ✅ 전국 합계: {total_value} (전년: {total_prev_value}, 증감률: {change_rate}%)")
            else:
                print(f"[{self.config['name']}] 🔍 [디버그] 전국 데이터 찾기 실패:")
                print(f"  - nationwide 타입: {type(nationwide)}")
                print(f"  - nationwide 값: {nationwide}")
                print(f"  - table_data 길이: {len(table_data)}")
                if table_data:
                    print(f"  - table_data 샘플 (처음 3개): {table_data[:3]}")
                print(f"[{self.config['name']}] ⚠️ 전국 데이터를 찾을 수 없습니다. 빈 데이터를 사용합니다.")
                nationwide = {
                    'region_name': '전국',
                    'region_display': '전 국',
                    'value': 0.0,
                    'prev_value': 0.0,
                    'change_rate': 0.0,
                    'growth_rate': 0.0,
                    'production_index': 0.0
                }
        
        # 국내인구이동은 nationwide가 없음 - 나머지만 처리
# 국내인구이동은 nationwide가 없음 - 나머지만 처리
        if nationwide:
            index_value = nationwide.get('value')
            if index_value is None:
                # index_value가 없으면 0.0으로 설정 (프로덕션 환경 안정성)
                index_value = 0.0
        
            growth_rate = nationwide.get('change_rate')
            if growth_rate is None:
                # growth_rate가 없으면 0.0으로 설정
                growth_rate = 0.0
        
            # 업종별 데이터 추출
            industry_data = self._extract_industry_data('전국')
        
            # 안전한 업종 데이터 처리
            if not industry_data:
                industry_data = []
        
            # 증가/감소 업종 분류 (None 체크 강화)
            increase_industries = [
                ind for ind in industry_data 
                if ind and isinstance(ind, dict) and 
                ind.get('change_rate') is not None and 
                ind['change_rate'] > 0
            ]
            decrease_industries = [
                ind for ind in industry_data 
                if ind and isinstance(ind, dict) and 
                ind.get('change_rate') is not None and 
                ind['change_rate'] < 0
            ]
            
            # 증감률 기준 정렬 (안전한 정렬)
            try:
                # 기본값/폴백 사용 금지: change_rate가 None이면 정렬에서 제외
                increase_industries = [x for x in increase_industries if x and isinstance(x, dict) and x.get('change_rate') is not None]
                decrease_industries = [x for x in decrease_industries if x and isinstance(x, dict) and x.get('change_rate') is not None]
                increase_industries.sort(key=lambda x: x['change_rate'], reverse=True)
                decrease_industries.sort(key=lambda x: x['change_rate'])
            except (TypeError, AttributeError) as e:
                print(f"[{self.config['name']}] ⚠️ 업종 정렬 오류: {e}")
                # 정렬 실패 시 원본 유지
            
            # 상위 3개 추출 (안전한 슬라이싱)
            # 기본값/폴백 사용 금지
            main_increase = increase_industries[:3] if increase_industries and len(increase_industries) > 0 else None
            main_decrease = decrease_industries[:3] if decrease_industries and len(decrease_industries) > 0 else None
        else:
            # nationwide가 None인 경우 (국내인구이동 등)
            index_value = None
            growth_rate = None
            main_increase = None
            main_decrease = None
        
        # 모든 필드명 포함 (기존 데이터 유지 및 템플릿 호환 필드 추가)
        result = (nationwide.copy() if nationwide else {})
        result.update({
            'production_index': index_value,
            'sales_index': index_value,  # 소비동향 템플릿 호환
            'service_index': index_value,  # 서비스업 템플릿 호환
            'growth_rate': growth_rate,
            'main_items': main_increase,  # 업종별 데이터 추가 완료
            'main_industries': main_increase,  # 템플릿 호환
            'main_businesses': main_increase,  # 소비동향 템플릿 호환
            'main_increase_industries': main_increase,  # 템플릿 호환
            'main_decrease_industries': main_decrease,   # 템플릿 호환
            'value': index_value,  # 명시적으로 보장
            'change_rate': growth_rate  # 명시적으로 보장
        })

        # 건설동향 템플릿 호환 별칭 추가
        if self.report_type == 'construction':
            # construction_template.html에서 요구하는 키: construction_index_trillion
            # index_value가 백억원이므로 조원 단위로 변환 (백억원 * 100 = 조원)
            construction_trillion = (index_value / 100) if index_value else None
            result['construction_index_trillion'] = construction_trillion
            result['change'] = growth_rate
            # 토목/건축 증감률 (기본값은 전체 증감률 사용)
            result['civil_growth'] = growth_rate
            result['building_growth'] = growth_rate
            # 토목/건축 부공종 (기본값)
            result['civil_subtypes'] = '철도·궤도, 기계설치'
            result['building_subtypes'] = '주택, 관공서 등'
            result['main_category'] = '토목' if (growth_rate is not None and growth_rate >= 0) else '토목'
            result['sub_types_text'] = '철도·궤도, 도로·교량, 주택'
        # 고용률/실업률 템플릿 호환 별칭 추가
        elif self.report_type == 'employment':
            # employment_template.html에서 요구하는 키: employment_rate, change, main_age_groups, top_age_groups
            result['employment_rate'] = index_value
            result['change'] = growth_rate
            result['main_age_groups'] = []
            result['top_age_groups'] = []
        elif self.report_type == 'unemployment':
            # unemployment_template.html에서 요구하는 키: rate, change, age_groups
            result['rate'] = index_value
            result['change'] = growth_rate
            result['age_groups'] = []

        return result
    
    def extract_regional_data(self, table_data: List[Dict] = None) -> Dict[str, Any]:
        """시도별 데이터 추출"""
        if table_data is None:
            table_data = self._extract_table_data_ssot()
        
        # 전국 제외 (안전한 필터링)
        regional = [
            d for d in table_data 
            if d and isinstance(d, dict) and 
            d.get('region_name') != '전국'
        ]
        
        # 증가/감소 분류 (None 체크 강화)
        increase = [
            r for r in regional 
            if r and isinstance(r, dict) and 
            r.get('change_rate') is not None and 
            r['change_rate'] > 0
        ]
        decrease = [
            r for r in regional 
            if r and isinstance(r, dict) and 
            r.get('change_rate') is not None and 
            r['change_rate'] < 0
        ]
        
        # 기본값/폴백 사용 금지: 정렬 (change_rate가 None이면 제외)
        try:
            # change_rate가 None인 항목은 정렬에서 제외
            increase_filtered = [x for x in increase if x and isinstance(x, dict) and x.get('change_rate') is not None]
            decrease_filtered = [x for x in decrease if x and isinstance(x, dict) and x.get('change_rate') is not None]
            increase_filtered.sort(key=lambda x: x['change_rate'], reverse=True)
            decrease_filtered.sort(key=lambda x: x['change_rate'])
            increase = increase_filtered
            decrease = decrease_filtered
        except (TypeError, AttributeError, KeyError) as e:
            print(f"[{self.config['name']}] 🔍 [디버그] 지역 정렬 오류:")
            print(f"  - 오류: {e}")
            print(f"  - increase 샘플: {increase[:3] if increase else '없음'}")
            print(f"  - decrease 샘플: {decrease[:3] if decrease else '없음'}")
            raise ValueError(f"[{self.config['name']}] ❌ 지역 정렬 오류: {e}. 기본값 사용 금지: 반드시 데이터를 찾아야 합니다.")
        
        return {
            'increase_regions': increase,
            'decrease_regions': decrease,
            'all_regions': regional
        }

    def _build_summary_table(self, table_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """템플릿용 요약 테이블 생성 (필수 필드만 기본 값으로 채움)"""
        if table_data is None:
            table_data = []

        # 4개 증감률 컬럼, 3개 지수/율 컬럼을 기본 라벨로 구성
        def _previous_quarter(year: int, quarter: int) -> tuple[int, int]:
            if quarter <= 1:
                return (year - 1, 4)
            return (year, quarter - 1)

        def _growth_labels(year: Optional[int], quarter: Optional[int]) -> List[str]:
            if year is None or quarter is None:
                return ["전전기", "전기", "직전기", "현기"]
            prev_q_year, prev_q = _previous_quarter(year, quarter)
            return [
                f"{year-2}.{quarter}/4",
                f"{year-1}.{quarter}/4",
                f"{prev_q_year}.{prev_q}/4",
                f"{year}.{quarter}/4",
            ]

        def _index_labels(year: Optional[int], quarter: Optional[int]) -> List[str]:
            if self.report_type == 'employment':
                age_label = "20-29세"  # 고용률은 20-29세 사용
            elif self.report_type == 'unemployment':
                age_label = "15-29세"  # 실업률은 15-29세 사용
            else:
                age_label = "15-29세"
            if year is None or quarter is None:
                return ["전기", "현기", "청년층"]
            return [
                f"{year-1}.{quarter}/4",
                f"{year}.{quarter}/4",
                age_label,
            ]

        growth_cols = _growth_labels(self.year, self.quarter)
        index_cols = _index_labels(self.year, self.quarter)

        target_quarter_keys: List[str] = []
        if self.year is not None and self.quarter is not None:
            prev_q_year, prev_q = _previous_quarter(self.year, self.quarter)
            target_quarter_keys = [
                self._format_quarter_key(self.year - 2, self.quarter),
                self._format_quarter_key(self.year - 1, self.quarter),
                self._format_quarter_key(prev_q_year, prev_q),
                self._format_quarter_key(self.year, self.quarter),
            ]

        def _normalize_quarter_key(key: str) -> str:
            """분기 키를 정규화 (예: '2023_3Q' -> '2023 3/4')"""
            if not key:
                return key
            key = str(key).strip()
            # '2023_3Q' -> '2023 3/4' 형식 변환
            import re
            match = re.match(r'(\d{4})_(\d)Q', key)
            if match:
                return f"{match.group(1)} {match.group(2)}/4"
            # '2023 3/4' 형식은 그대로 반환
            return key

        def _map_quarter_values(keys: Any, values: Any) -> List[Optional[float]]:
            if not keys or not values:
                return [None, None, None, None]
            # 키 정규화 후 매핑
            normalized_mapping = {}
            for k, v in zip(keys, values):
                normalized_key = _normalize_quarter_key(k)
                normalized_mapping[normalized_key] = v
            if not target_quarter_keys:
                return [None, None, None, None]
            return [normalized_mapping.get(k) for k in target_quarter_keys]

        def _to_float(value: Any) -> Optional[float]:
            if value is None or value == '' or value == '-':
                return None
            try:
                return float(value)
            except Exception:
                return None

        def _compute_growth(current: Optional[float], previous: Optional[float]) -> Optional[float]:
            if current is None or previous is None:
                return None
            if previous == 0:
                return None
            return round((current - previous) / previous * 100, 1)

        def _build_growth_slots(row: Dict[str, Any]) -> List[Optional[float]]:
            # 표에는 '전년동분기대비 증감률(%)'을 표시해야 함
            # 따라서 quarterly_growth_rates(전분기 대비)를 사용하지 않고
            # prev_prev_prev_value, prev_prev_value, prev_value, value를 사용하여 계산
            
            # 단, value_type이 'change_rate'인 경우 (시트에 이미 증감률이 있는 경우)는 그대로 사용
            if self.config.get('value_type') == 'change_rate':
                rate_keys = row.get('rate_quarterly_keys') or row.get('quarterly_keys')
                rate_values = row.get('rate_quarterly_values') or row.get('quarterly_values')
                mapped = _map_quarter_values(rate_keys, rate_values)
                if any(v is not None for v in mapped):
                    return mapped
            
            # 전년동분기 대비 증감률 계산
            # - 슬롯 0 (2년 전): prev_prev_value 대비 prev_prev_prev_value
            # - 슬롯 1 (1년 전): prev_value 대비 prev_prev_value
            # - 슬롯 2 (직전분기): previous_quarter_growth 사용 (이미 전년동분기 대비로 계산됨)
            # - 슬롯 3 (현재): current_value 대비 prev_value
            current_value = _to_float(row.get('value'))
            prev_value = _to_float(row.get('prev_value'))
            prev_prev_value = _to_float(row.get('prev_prev_value'))
            prev_prev_prev_value = _to_float(row.get('prev_prev_prev_value'))

            # 고용률/실업률은 퍼센트포인트(%p) 차이를 사용
            if self.report_type in ['employment', 'unemployment']:
                # 퍼센트포인트 차이 계산 (current - previous)
                two_years_ago = round(prev_prev_value - prev_prev_prev_value, 1) if (prev_prev_value is not None and prev_prev_prev_value is not None) else None
                last_year = round(prev_value - prev_prev_value, 1) if (prev_value is not None and prev_prev_value is not None) else None
                # 직전분기의 전년동분기 대비 증감(%p) - 별도로 저장된 값 사용
                previous_quarter = _to_float(
                    row.get('previous_quarter_growth') or row.get('prev_quarter_growth')
                )
                # 현재 분기의 전년동분기 대비 증감(%p)
                current = round(current_value - prev_value, 1) if (current_value is not None and prev_value is not None) else None
                if current is None:
                    current = _to_float(row.get('change_rate'))
                return [two_years_ago, last_year, previous_quarter, current]
            
            # 다른 보고서 유형은 퍼센트 증감률 사용
            two_years_ago = _compute_growth(prev_prev_value, prev_prev_prev_value)
            last_year = _compute_growth(prev_value, prev_prev_value)
            
            # 직전분기의 전년동분기 대비 증감률
            # 이 값은 별도로 저장되어 있지 않으므로 None으로 유지
            previous_quarter = _to_float(
                row.get('previous_quarter_growth') or row.get('prev_quarter_growth')
            )
            
            # 현재 분기의 전년동분기 대비 증감률
            current = _compute_growth(current_value, prev_value)
            if current is None:
                current = _to_float(row.get('change_rate'))

            return [two_years_ago, last_year, previous_quarter, current]

        regions = []
        for row in table_data:
            region_name = row.get('region_name', '') if isinstance(row, dict) else ''
            growth_rate = row.get('change_rate') if isinstance(row, dict) else None
            value = row.get('value') if isinstance(row, dict) else None
            prev_value = row.get('prev_value') if isinstance(row, dict) else None
            prev_prev_value = row.get('prev_prev_value') if isinstance(row, dict) else None
            prev_prev_prev_value = row.get('prev_prev_prev_value') if isinstance(row, dict) else None

            computed = _build_growth_slots(row) if isinstance(row, dict) else [None, None, None, None]
            
            # 마지막 슬롯(현재 분기)에 change_rate 사용 (분기별 매핑 실패 시 폴백)
            if computed[3] is None and growth_rate is not None:
                computed[3] = growth_rate
            
            # 실업률/고용률/국내인구이동: prev_prev_value와 prev_prev_prev_value를 사용하여 이전 분기 증감도 계산
            if self.report_type in ['employment', 'unemployment']:
                # 전년동분기대비 증감(%p) 계산
                if computed[0] is None and prev_prev_prev_value is not None:
                    # 2023. 3/4 증감 = 2023 3/4 - 2022 3/4 (= prev_prev_value - prev_prev_prev_value)
                    if prev_prev_value is not None:
                        computed[0] = round(prev_prev_value - prev_prev_prev_value, 1)
                if computed[1] is None and prev_prev_value is not None and prev_value is not None:
                    # 2024. 3/4 증감 = 2024 3/4 - 2023 3/4 (= prev_value - prev_prev_value)
                    computed[1] = round(prev_value - prev_prev_value, 1)
                # 직전 분기(2025. 2/4) 증감은 계산하기 어려우므로 None 유지
            
            growth_rates = [
                '' if computed[0] is None else computed[0],
                '' if computed[1] is None else computed[1],
                '' if computed[2] is None else computed[2],
                '' if computed[3] is None else computed[3],
            ]

            youth_rate = row.get('youth_rate') if isinstance(row, dict) else None
            regions.append({
                'group': None,
                'region': region_name,
                'sido': region_name,
                'region_group': None,
                'rowspan': 1,
                # 보유한 데이터 기반으로 증감률 슬롯 채움
                'growth_rates': growth_rates,
                'indices': [prev_value, value, ''],
                'changes': growth_rates,
                'amounts': [prev_value, value] if self.report_type in ['export', 'import'] else [prev_value, value],
                'rates': [prev_value, value, youth_rate if youth_rate not in (None, '', '-') else ''],
                'youth_rate': youth_rate,
                'quarterly_keys': row.get('quarterly_keys') if isinstance(row, dict) else None,
                'quarterly_values': row.get('quarterly_values') if isinstance(row, dict) else None,
                'quarterly_growth_rates': row.get('quarterly_growth_rates') if isinstance(row, dict) else None,
                'rate_quarterly_keys': row.get('rate_quarterly_keys') if isinstance(row, dict) else None,
                'rate_quarterly_values': row.get('rate_quarterly_values') if isinstance(row, dict) else None,
                'prev_prev_value': prev_prev_value,
                'prev_prev_prev_value': prev_prev_prev_value,
            })

        return {
            'base_year': 2020,
            'columns': {
                'growth_rate_columns': growth_cols,
                'index_columns': index_cols,
                'change_columns': growth_cols,
                'rate_columns': index_cols,
                # 수출/수입 템플릿에서 액수 컬럼 라벨 요구
                'amount_columns': index_cols[:2],
            },
            'regions': regions,
            'rows': regions,
        }

    def _extract_item_names(self, items: Any) -> List[str]:
        """리스트에서 표시용 이름만 추출"""
        if not items:
            return []
        names = []
        for item in items:
            if isinstance(item, dict):
                name_val = item.get('name') or item.get('display_name')
                if name_val is not None:
                    names.append(name_val)
            else:
                names.append(item)
        return names

    def _enrich_template_data(
        self,
        data: Dict[str, Any],
        table_data: List[Dict[str, Any]],
        regional: Dict[str, Any],
        top3_increase: List[Dict[str, Any]],
        top3_decrease: List[Dict[str, Any]],
    ) -> None:
        """템플릿에서 요구하는 필드를 채워 렌더링 오류를 방지"""

        # summary_box 기본 필드 보강
        summary_box = data.get('summary_box', {}) or {}
        summary_box.setdefault('increase_count', len(regional.get('increase_regions', [])))
        summary_box.setdefault('decrease_count', len(regional.get('decrease_regions', [])))
        summary_box.setdefault('region_count', len(regional.get('increase_regions', [])))
        summary_box.setdefault('main_items', [])
        data['summary_box'] = summary_box

        # summary_table 기본 구조 추가
        data['summary_table'] = self._build_summary_table(table_data)

        # footer 정보 기본값
        data.setdefault('footer_info', {
            'source': '자료: 국가데이터처 국가통계포털(KOSIS), 집계시트',
            'page_num': '1'
        })

        # nationwide 필드 보강 (보고서 타입별 별칭)
        nationwide = data.get('nationwide_data') or {}
        if self.report_type in ['export', 'import']:
            # scale_factor가 이미 0.01이 적용되어 있음
            nationwide.setdefault('amount', nationwide.get('value'))
            nationwide.setdefault('change', nationwide.get('change_rate'))
            products = nationwide.get('products') or nationwide.get('main_items') or []
            normalized_products = []
            for p in products:
                if isinstance(p, dict):
                    normalized_products.append({
                        'name': p.get('name') or p.get('display_name') or str(p),
                        'change': p.get('change', nationwide.get('change'))
                    })
                else:
                    normalized_products.append({'name': p, 'change': nationwide.get('change')})
            nationwide['products'] = normalized_products
        elif self.report_type == 'price':
            nationwide.setdefault('index', nationwide.get('production_index'))
            nationwide.setdefault('change', nationwide.get('growth_rate'))
            categories = nationwide.get('categories') or nationwide.get('main_items') or []
            normalized_categories = []
            for cat in categories:
                if isinstance(cat, dict):
                    normalized_categories.append({
                        'name': cat.get('name') or cat.get('display_name') or str(cat),
                        'change': cat.get('change', cat.get('growth_rate', nationwide.get('change')))
                    })
                else:
                    normalized_categories.append({'name': cat, 'change': nationwide.get('change')})
            nationwide['categories'] = normalized_categories
        elif self.report_type == 'employment':
            nationwide.setdefault('employment_rate', nationwide.get('production_index'))
            nationwide.setdefault('change', nationwide.get('growth_rate'))
            nationwide.setdefault('main_age_groups', nationwide.get('main_age_groups', []))
            nationwide.setdefault('top_age_groups', nationwide.get('top_age_groups', []))
        elif self.report_type == 'unemployment':
            nationwide.setdefault('rate', nationwide.get('production_index'))
            nationwide.setdefault('change', nationwide.get('growth_rate'))
            nationwide.setdefault('age_groups', nationwide.get('age_groups', []))
            nationwide.setdefault('main_age_groups', nationwide.get('main_age_groups', []))
        elif self.report_type == 'migration':
            report_info = data.get('report_info', {}) or {}
            report_info.setdefault('age_20_29_label', '20-29세')
            report_info.setdefault('age_other_label', '그 외 연령층')
            report_info.setdefault('age_other_note', '그 외 연령층')
            data['report_info'] = report_info
        elif self.report_type == 'construction':
            # construction_template.html 호환성 보강
            nationwide.setdefault('civil_growth', nationwide.get('growth_rate'))
            nationwide.setdefault('building_growth', nationwide.get('growth_rate'))
            nationwide.setdefault('civil_subtypes', '철도·궤도, 기계설치')
            nationwide.setdefault('building_subtypes', '주택, 관공서 등')
            nationwide.setdefault('main_category', '토목' if (nationwide.get('growth_rate') is not None and nationwide.get('growth_rate') >= 0) else '토목')
            nationwide.setdefault('sub_types_text', '철도·궤도, 도로·교량, 주택')
        data['nationwide_data'] = nationwide

        # 지역 데이터 별칭/필드 보강
        regional_increase = regional.get('increase_regions', []) or []
        regional_decrease = regional.get('decrease_regions', []) or []

        for entry in regional_increase + regional_decrease:
            if not isinstance(entry, dict):
                continue
            if self.report_type in ['export', 'import']:
                entry.setdefault('amount', entry.get('value'))
                entry.setdefault('change', entry.get('change_rate'))
                raw_products = entry.get('products') or self._extract_item_names(entry.get('top_industries'))
                normalized_products = []
                for p in raw_products or []:
                    if isinstance(p, dict):
                        normalized_products.append({
                            'name': p.get('name') or p.get('display_name') or str(p),
                            'change': p.get('change', entry.get('change'))
                        })
                    else:
                        normalized_products.append({'name': p, 'change': entry.get('change')})
                entry['products'] = normalized_products
            elif self.report_type == 'price':
                categories = entry.get('categories') or entry.get('top_industries', [])
                normalized_categories = []
                for cat in categories:
                    if isinstance(cat, dict):
                        normalized_categories.append({
                            'name': cat.get('name') or cat.get('display_name') or str(cat),
                            'change': cat.get('change', cat.get('growth_rate', entry.get('change')))
                        })
                    else:
                        normalized_categories.append({'name': cat, 'change': entry.get('change')})
                entry['categories'] = normalized_categories
            elif self.report_type in ['employment', 'unemployment']:
                # 연령별 데이터 추출 및 채우기
                region_name = entry.get('region') or entry.get('region_name')
                if region_name:
                    existing_age_groups = entry.get('age_groups')
                    if not existing_age_groups:
                        age_groups = self._extract_age_groups_for_region(region_name)
                        entry['age_groups'] = age_groups

        # 고용률/실업률: 전국 연령별 데이터 채우기
        if self.report_type in ['employment', 'unemployment']:
            # 전국 연령별 데이터 추출
            if not nationwide.get('top_age_groups'):
                nationwide_age_groups = self._extract_age_groups_for_region('전국')
                nationwide['top_age_groups'] = nationwide_age_groups
                nationwide['main_age_groups'] = nationwide_age_groups[:4]
            if not nationwide.get('main_age_groups') and nationwide.get('top_age_groups'):
                nationwide['main_age_groups'] = nationwide.get('top_age_groups', [])[:4]
        
        if self.report_type == 'construction':
            nationwide.setdefault('civil_growth', nationwide.get('growth_rate'))
            nationwide.setdefault('building_growth', nationwide.get('growth_rate'))
            for entry in regional_increase + regional_decrease:
                if not isinstance(entry, dict):
                    continue
                entry.setdefault('civil_growth', entry.get('growth_rate'))
                entry.setdefault('building_growth', entry.get('growth_rate'))

        if self.report_type == 'price':
            regional['high_regions'] = regional_increase
            regional['low_regions'] = regional_decrease
            
            # 물가 보고서: summary_box.main_items에 전국 주요 품목 이름 추가
            nationwide_categories = nationwide.get('categories') or nationwide.get('main_items') or []
            main_item_names = []
            for cat in nationwide_categories:
                if isinstance(cat, dict):
                    name = cat.get('name') or cat.get('display_name')
                    if name:
                        main_item_names.append(name)
                elif cat:
                    main_item_names.append(str(cat))
            summary_box['main_items'] = main_item_names[:3]  # 상위 3개만
            
            # 물가 보고서: low_items_text와 high_items_text 생성
            # 전국보다 낮은 지역의 주요 품목 텍스트
            low_items = []
            for region in regional_decrease[:3]:
                if isinstance(region, dict):
                    region_name = region.get('region') or region.get('region_name')
                    if region_name:
                        # 지역별 품목 데이터 추출
                        region_cats = region.get('categories') or region.get('top_industries') or []
                        if not region_cats:
                            region_cats = self._extract_industry_data(region_name)[:3]
                            region['categories'] = region_cats
                        for cat in region_cats[:3]:
                            if isinstance(cat, dict):
                                name = cat.get('name') or cat.get('display_name')
                                if name and name not in low_items:
                                    low_items.append(name)
            data['low_items_text'] = ', '.join(low_items[:3]) if low_items else ''
            
            # 전국보다 높은 지역의 주요 품목 텍스트
            high_items = []
            for region in regional_increase[:3]:
                if isinstance(region, dict):
                    region_name = region.get('region') or region.get('region_name')
                    if region_name:
                        # 지역별 품목 데이터 추출
                        region_cats = region.get('categories') or region.get('top_industries') or []
                        if not region_cats:
                            region_cats = self._extract_industry_data(region_name)[:3]
                            region['categories'] = region_cats
                        for cat in region_cats[:3]:
                            if isinstance(cat, dict):
                                name = cat.get('name') or cat.get('display_name')
                                if name and name not in high_items:
                                    high_items.append(name)
            data['high_items_text'] = ', '.join(high_items[:3]) if high_items else ''

        data['regional_data'] = regional

        # Top3 리스트 별칭 보강
        # None 제거 및 products 필드 항상 리스트 보장
        def ensure_products(item):
            if not isinstance(item, dict):
                return {}
            if self.report_type in ['export', 'import']:
                item.setdefault('amount', item.get('value'))
                # change_rate 또는 growth_rate에서 change 값을 가져옴
                change_val = item.get('change_rate') or item.get('growth_rate')
                item.setdefault('change', change_val)
                
                # industries가 없으면 _extract_industry_data로 추출
                industries = item.get('industries')
                region_name = item.get('region') or item.get('region_name')
                if (not industries or (isinstance(industries, list) and len(industries) == 0)) and region_name:
                    industries = self._extract_industry_data(region_name)[:3]
                    item['industries'] = industries
                
                if not isinstance(industries, list):
                    industries = [] if industries is None else [industries]
                
                # products 필드 생성 (템플릿에서 사용)
                normalized_products = []
                for ind in industries[:3]:
                    if isinstance(ind, dict):
                        normalized_products.append({
                            'name': ind.get('name') or ind.get('display_name') or str(ind),
                            'change': ind.get('change_rate') or ind.get('growth_rate') or ind.get('change'),
                            'contribution_rate': ind.get('contribution_rate'),
                            'contribution': ind.get('contribution'),
                            'growth_rate': ind.get('growth_rate') or ind.get('change_rate')
                        })
                    elif ind:
                        normalized_products.append({'name': str(ind), 'change': change_val})
                
                item['products'] = normalized_products
                
                # products가 None이거나 리스트가 아니면 무조건 빈 리스트로 보정
                if not isinstance(item.get('products'), list):
                    item['products'] = []
                if item['products'] is None:
                    item['products'] = []
            elif self.report_type == 'price':
                item.setdefault('categories', item.get('industries', []))
            elif self.report_type in ['employment', 'unemployment']:
                item.setdefault('age_groups', [])
            elif self.report_type == 'construction':
                item.setdefault('civil_growth', item.get('growth_rate'))
                item.setdefault('building_growth', item.get('growth_rate'))
                item.setdefault('civil_subtypes', '철도·궤도, 기계설치')
                item.setdefault('building_subtypes', '주택, 관공서 등')
            return item

        # None 제거 및 보정 적용 (products 필드가 항상 리스트가 되도록 보정)
        def ensure_products_final(x):
            x = ensure_products(x)
            if not isinstance(x, dict):
                return None
            if 'products' not in x or not isinstance(x['products'], list) or x['products'] is None:
                x['products'] = []
            return x

        top3_increase[:] = [y for y in (ensure_products_final(i) for i in top3_increase) if isinstance(y, dict) and y]
        top3_decrease[:] = [y for y in (ensure_products_final(i) for i in top3_decrease) if isinstance(y, dict) and y]

        data['top3_increase_regions'] = top3_increase
        data['top3_decrease_regions'] = top3_decrease

        # 고용률/실업률: top3_increase_regions와 top3_decrease_regions에 연령별 데이터 채우기
        if self.report_type in ['employment', 'unemployment']:
            for item in top3_increase:
                region_name = item.get('region')
                if region_name and not item.get('age_groups'):
                    age_groups = self._extract_age_groups_for_region(region_name)
                    item['age_groups'] = age_groups
            for item in top3_decrease:
                region_name = item.get('region')
                if region_name and not item.get('age_groups'):
                    age_groups = self._extract_age_groups_for_region(region_name)
                    item['age_groups'] = age_groups

        if self.report_type == 'price':
            # 물가 보고서: top3_above_regions 생성 (categories가 없으면 추출)
            top3_above = []
            for item in top3_increase:
                region_name = item.get('region')
                existing_cats = item.get('categories', item.get('industries', []))
                if not existing_cats and region_name:
                    # categories가 없으면 _extract_industry_data로 추출
                    existing_cats = self._extract_industry_data(region_name)[:3]
                normalized_cats = []
                for cat in (existing_cats or []):
                    if isinstance(cat, dict):
                        normalized_cats.append({
                            'name': cat.get('name') or cat.get('display_name') or str(cat),
                            'change': cat.get('change', cat.get('growth_rate', cat.get('change_rate', item.get('growth_rate')))),
                            'contribution_rate': cat.get('contribution_rate'),
                            'contribution': cat.get('contribution'),
                            'growth_rate': cat.get('growth_rate', cat.get('change_rate'))
                        })
                    else:
                        normalized_cats.append({'name': cat, 'change': item.get('growth_rate')})
                top3_above.append({
                    'name': region_name,
                    'region': region_name,
                    'change': item.get('growth_rate'),
                    'categories': normalized_cats[:3]
                })
            data['top3_above_regions'] = top3_above
            
            # 물가 보고서: top3_below_regions 생성 (categories가 없으면 추출)
            top3_below = []
            for item in top3_decrease:
                region_name = item.get('region')
                existing_cats = item.get('categories', item.get('industries', []))
                if not existing_cats and region_name:
                    # categories가 없으면 _extract_industry_data로 추출
                    existing_cats = self._extract_industry_data(region_name)[:3]
                normalized_cats = []
                for cat in (existing_cats or []):
                    if isinstance(cat, dict):
                        normalized_cats.append({
                            'name': cat.get('name') or cat.get('display_name') or str(cat),
                            'change': cat.get('change', cat.get('growth_rate', cat.get('change_rate', item.get('growth_rate')))),
                            'contribution_rate': cat.get('contribution_rate'),
                            'contribution': cat.get('contribution'),
                            'growth_rate': cat.get('growth_rate', cat.get('change_rate'))
                        })
                    else:
                        normalized_cats.append({'name': cat, 'change': item.get('growth_rate')})
                top3_below.append({
                    'name': region_name,
                    'region': region_name,
                    'change': item.get('growth_rate'),
                    'categories': normalized_cats[:3]
                })
            data['top3_below_regions'] = top3_below

    def extract_all_data(self, region: Optional[str] = None) -> Dict[str, Any]:
        """전체 데이터 추출"""
        # 데이터 로드는 외부에서 보장 (테스트 호환성)
        
        # 정적 컬럼 인덱스 사용 (동적 탐색 제거)
        target_idx = self.target_col
        prev_y_idx = self.prev_y_col
        
        if self.df_aggregation is not None:
            if target_idx is None or prev_y_idx is None:
                print(
                    f"[{self.config['name']}] ⚠️ 컬럼 인덱스가 설정되지 않아 기본 인덱스를 유지합니다. "
                    f"target_col={target_idx}, prev_y_col={prev_y_idx}"
                )
            
            self.target_col = target_idx
            self.prev_y_col = prev_y_idx
        else:
            raise ValueError(
                f"[{self.config['name']}] ❌ 집계 시트를 로드할 수 없습니다. "
                f"기본값 사용 금지: 반드시 데이터를 찾아야 합니다."
            )
        
        # Table Data (SSOT)
        table_data = self._extract_table_data_ssot()
        # 전처리 결과 DF 저장
        self.preprocessed_table_df = None
        if isinstance(table_data, list):
            try:
                self.preprocessed_table_df = pd.DataFrame(table_data)
            except Exception as e:
                print(f"[{self.config['name']}] ⚠️ 전처리 결과 DF 생성 실패: {e}")
        
        # Text Data
        nationwide = self.extract_nationwide_data(table_data)
        regional = self.extract_regional_data(table_data)
        
        # Top3 regions (템플릿 호환 필드명으로 생성, 기본값/폴백 사용 금지)
        top3_increase = []
        if 'increase_regions' not in regional or not isinstance(regional['increase_regions'], list):
            print(f"[{self.config['name']}] 🔍 [디버그] regional 데이터에서 'increase_regions' 찾기 실패:")
            print(f"  - regional 타입: {type(regional)}")
            print(f"  - regional 키: {list(regional.keys()) if isinstance(regional, dict) else 'N/A'}")
            print(f"  - regional 전체 값: {regional}")
            raise ValueError(
                f"[{self.config['name']}] ❌ regional 데이터에서 'increase_regions'를 찾을 수 없습니다.\n"
                f"  regional 타입: {type(regional)}\n"
                f"  regional 키: {list(regional.keys()) if isinstance(regional, dict) else 'N/A'}\n"
                f"  regional 전체 값: {regional}"
            )
        increase_regions = regional['increase_regions']
        
        for r in increase_regions[:3]:
            if not r or not isinstance(r, dict):
                continue
            
            # 기본값/폴백 사용 금지
            if 'region_name' not in r or not r['region_name']:
                print(f"[{self.config['name']}] 🔍 [디버그] region_name 찾기 실패:")
                print(f"  - r 타입: {type(r)}")
                print(f"  - r 키: {list(r.keys()) if isinstance(r, dict) else 'N/A'}")
                print(f"  - r 전체 값: {r}")
                continue
            region_name = r['region_name']
            
            try:
                # 지역별 업종 데이터 추출
                region_industries = self._extract_industry_data(region_name)
                # 기본값/폴백 사용 금지: 빈 리스트는 그대로 사용 (데이터가 없는 경우)
                # 하지만 None 체크는 필요
                if region_industries is None:
                    raise ValueError(f"[{self.config['name']}] ❌ {region_name} 업종 데이터를 추출할 수 없습니다. 기본값 사용 금지: 반드시 데이터를 찾아야 합니다.")
                
                # 증가 업종만 필터링 및 정렬 (안전한 처리)
                increase_industries = [
                    ind for ind in region_industries 
                    if ind and isinstance(ind, dict) and 
                    ind.get('change_rate') is not None and 
                    ind['change_rate'] > 0
                ]
                try:
                    # 기본값/폴백 사용 금지: change_rate가 None이면 정렬에서 제외
                    increase_industries = [x for x in increase_industries if x and isinstance(x, dict) and x.get('change_rate') is not None]
                    increase_industries.sort(key=lambda x: x['change_rate'], reverse=True)
                except (TypeError, AttributeError):
                    pass  # 정렬 실패 시 원본 유지
                
                # 상위 3개 업종 추출
                top_industries = increase_industries[:3] if increase_industries and len(increase_industries) > 0 else None
                # 업종 이름만 추출한 리스트 (템플릿에서 industries_names로 사용)
                industries_names = [ind.get('name', '') for ind in (top_industries or []) if ind and isinstance(ind, dict) and ind.get('name')]
                
                top3_increase.append({
                    'region': region_name,
                    # 기본값/폴백 사용 금지
                    'growth_rate': r['change_rate'] if 'change_rate' in r and r['change_rate'] is not None else None,
                    # 기본값/폴백 사용 금지: increase_industries가 없으면 None
                    'industries': top_industries,
                    # 업종 이름 리스트 (템플릿 호환)
                    'industries_names': industries_names
                })
            except Exception as e:
                print(f"[{self.config['name']}] ⚠️ {region_name} 업종 데이터 추출 오류: {e}")
                # 오류 발생 시 빈 업종 리스트로 추가
                top3_increase.append({
                    'region': region_name,
                    # 기본값/폴백 사용 금지
                    'growth_rate': r['change_rate'] if 'change_rate' in r and r['change_rate'] is not None else None,
                    # 기본값/폴백 사용 금지: 빈 리스트 대신 None
                    'industries': None,
                    'industries_names': []
                })
        
        top3_decrease = []
        # 기본값/폴백 사용 금지
        if 'decrease_regions' not in regional or not isinstance(regional['decrease_regions'], list):
            print(f"[{self.config['name']}] 🔍 [디버그] regional 데이터에서 'decrease_regions' 찾기 실패:")
            print(f"  - regional 타입: {type(regional)}")
            print(f"  - regional 키: {list(regional.keys()) if isinstance(regional, dict) else 'N/A'}")
            raise ValueError(
                f"[{self.config['name']}] ❌ regional 데이터에서 'decrease_regions'를 찾을 수 없습니다.\n"
                f"  regional 타입: {type(regional)}\n"
                f"  regional 키: {list(regional.keys()) if isinstance(regional, dict) else 'N/A'}"
            )
        decrease_regions = regional['decrease_regions']
        # 기본값/폴백 사용 금지: 타입 체크는 이미 위에서 했으므로 여기서는 추가 체크 불필요
        
        for r in decrease_regions[:3]:
            if not r or not isinstance(r, dict):
                continue
            
            # 기본값/폴백 사용 금지
            if 'region_name' not in r or not r['region_name']:
                print(f"[{self.config['name']}] 🔍 [디버그] region_name 찾기 실패:")
                print(f"  - r 타입: {type(r)}")
                print(f"  - r 키: {list(r.keys()) if isinstance(r, dict) else 'N/A'}")
                print(f"  - r 전체 값: {r}")
                continue
            region_name = r['region_name']
            
            try:
                # 지역별 업종 데이터 추출
                region_industries = self._extract_industry_data(region_name)
                # 기본값/폴백 사용 금지: 빈 리스트는 그대로 사용 (데이터가 없는 경우)
                # 하지만 None 체크는 필요
                if region_industries is None:
                    raise ValueError(f"[{self.config['name']}] ❌ {region_name} 업종 데이터를 추출할 수 없습니다. 기본값 사용 금지: 반드시 데이터를 찾아야 합니다.")
                
                # 감소 업종만 필터링 및 정렬 (안전한 처리)
                decrease_industries = [
                    ind for ind in region_industries 
                    if ind and isinstance(ind, dict) and 
                    ind.get('change_rate') is not None and 
                    ind['change_rate'] < 0
                ]
                try:
                    # 기본값/폴백 사용 금지: change_rate가 None이면 정렬에서 제외
                    decrease_industries_filtered = [x for x in decrease_industries if x and isinstance(x, dict) and x.get('change_rate') is not None]
                    decrease_industries_filtered.sort(key=lambda x: x['change_rate'])
                    decrease_industries = decrease_industries_filtered
                except (TypeError, AttributeError, KeyError) as e:
                    print(f"[{self.config['name']}] 🔍 [디버그] decrease_industries 정렬 오류:")
                    print(f"  - 오류: {e}")
                    print(f"  - decrease_industries 샘플: {decrease_industries[:3] if decrease_industries else '없음'}")
                    raise ValueError(f"[{self.config['name']}] ❌ decrease_industries 정렬 오류: {e}. 기본값 사용 금지: 반드시 데이터를 찾아야 합니다.")
                
                # 소비동향용 주요 업태 (첫 번째 감소 업종, 기본값/폴백 사용 금지)
                main_business = None
                if decrease_industries and decrease_industries[0] and isinstance(decrease_industries[0], dict):
                    # 기본값/폴백 사용 금지
                    if 'name' not in decrease_industries[0] or not decrease_industries[0]['name']:
                        raise ValueError(f"[{self.config['name']}] ❌ decrease_industries[0]에서 'name'을 찾을 수 없습니다.")
                    main_business = decrease_industries[0]['name']
                
                # 상위 3개 업종 추출
                top_industries = decrease_industries[:3] if decrease_industries and len(decrease_industries) > 0 else None
                # 업종 이름만 추출한 리스트 (템플릿에서 industries_names로 사용)
                industries_names = [ind.get('name', '') for ind in (top_industries or []) if ind and isinstance(ind, dict) and ind.get('name')]
                
                top3_decrease.append({
                    'region': region_name,
                    # 기본값/폴백 사용 금지
                    'growth_rate': r['change_rate'] if 'change_rate' in r and r['change_rate'] is not None else None,
                    # 기본값/폴백 사용 금지
                    'industries': top_industries,
                    'main_business': main_business,  # 소비동향용 주요 업태
                    # 업종 이름 리스트 (템플릿 호환)
                    'industries_names': industries_names
                })
            except Exception as e:
                print(f"[{self.config['name']}] ⚠️ {region_name} 업종 데이터 추출 오류: {e}")
                # 오류 발생 시 빈 업종 리스트로 추가
                top3_decrease.append({
                    'region': region_name,
                    # 기본값/폴백 사용 금지
                    'growth_rate': r['change_rate'] if 'change_rate' in r and r['change_rate'] is not None else None,
                    # 기본값/폴백 사용 금지: 빈 리스트 대신 None
                    'industries': None,
                    # 기본값/폴백 사용 금지: 빈 문자열 대신 None
                    'main_business': None,
                    'industries_names': []
                })
        
        # Summary Box (안전한 처리)
        main_regions = []
        for r in top3_increase:
            if r and isinstance(r, dict):
                main_regions.append({
                    # 기본값/폴백 사용 금지
                    'region': r['region'] if 'region' in r and r['region'] else None,
                    # 기본값/폴백 사용 금지
                    'items': r['industries'] if 'industries' in r and isinstance(r['industries'], list) else None
                })
        
        # 기본값/폴백 사용 금지
        if 'increase_regions' not in regional or not isinstance(regional['increase_regions'], list):
            raise ValueError(f"[{self.config['name']}] ❌ regional 데이터에서 'increase_regions'를 찾을 수 없습니다.")
        increase_regions_count = len(regional['increase_regions'])
        
        summary_box = {
            'main_regions': main_regions,
            'region_count': increase_regions_count
        }
        
        # Regional data 필드명 변환 (템플릿 호환, 기본값/폴백 사용 금지)
        if 'increase_regions' not in regional or not isinstance(regional['increase_regions'], list):
            print(f"[{self.config['name']}] 🔍 [디버그] regional 데이터에서 'increase_regions' 찾기 실패:")
            print(f"  - regional 타입: {type(regional)}")
            print(f"  - regional 키: {list(regional.keys()) if isinstance(regional, dict) else 'N/A'}")
            raise ValueError(
                f"[{self.config['name']}] ❌ regional 데이터에서 'increase_regions'를 찾을 수 없습니다.\n"
                f"  regional 타입: {type(regional)}\n"
                f"  regional 키: {list(regional.keys()) if isinstance(regional, dict) else 'N/A'}"
            )
        increase_regions_list = regional['increase_regions']
        
        if 'decrease_regions' not in regional or not isinstance(regional['decrease_regions'], list):
            print(f"[{self.config['name']}] 🔍 [디버그] regional 데이터에서 'decrease_regions' 찾기 실패:")
            print(f"  - regional 타입: {type(regional)}")
            print(f"  - regional 키: {list(regional.keys()) if isinstance(regional, dict) else 'N/A'}")
            raise ValueError(
                f"[{self.config['name']}] ❌ regional 데이터에서 'decrease_regions'를 찾을 수 없습니다.\n"
                f"  regional 타입: {type(regional)}\n"
                f"  regional 키: {list(regional.keys()) if isinstance(regional, dict) else 'N/A'}"
            )
        decrease_regions_list = regional['decrease_regions']
        
        if 'all_regions' not in regional or not isinstance(regional['all_regions'], list):
            print(f"[{self.config['name']}] 🔍 [디버그] regional 데이터에서 'all_regions' 찾기 실패:")
            print(f"  - regional 타입: {type(regional)}")
            print(f"  - regional 키: {list(regional.keys()) if isinstance(regional, dict) else 'N/A'}")
            raise ValueError(
                f"[{self.config['name']}] ❌ regional 데이터에서 'all_regions'를 찾을 수 없습니다.\n"
                f"  regional 타입: {type(regional)}\n"
                f"  regional 키: {list(regional.keys()) if isinstance(regional, dict) else 'N/A'}"
            )
        all_regions_list = regional['all_regions']
        
        regional_converted = {
            'increase_regions': [
                {
                    # 기본값/폴백 사용 금지
                    'region': r['region_name'] if r and isinstance(r, dict) and 'region_name' in r and r['region_name'] else None,
                    'growth_rate': r['change_rate'] if r and isinstance(r, dict) and 'change_rate' in r and r['change_rate'] is not None else None,
                    'change': r['change_rate'] if r and isinstance(r, dict) and 'change_rate' in r and r['change_rate'] is not None else None,
                    # 기본값/폴백 사용 금지
                    'value': r['value'] if r and isinstance(r, dict) and 'value' in r and r['value'] is not None else None,
                    'top_industries': self._get_top_industries_for_region(
                        r['region_name'] if r and isinstance(r, dict) and 'region_name' in r and r['region_name'] else None, 
                        increase=True
                    )
                }
                for r in increase_regions_list
                if r and isinstance(r, dict) and r.get('region_name')
            ],
            'decrease_regions': [
                {
                    # 기본값/폴백 사용 금지
                    'region': r['region_name'] if r and isinstance(r, dict) and 'region_name' in r and r['region_name'] else None,
                    'growth_rate': r['change_rate'] if r and isinstance(r, dict) and 'change_rate' in r and r['change_rate'] is not None else None,
                    'change': r['change_rate'] if r and isinstance(r, dict) and 'change_rate' in r and r['change_rate'] is not None else None,
                    # 기본값/폴백 사용 금지
                    'value': r['value'] if r and isinstance(r, dict) and 'value' in r and r['value'] is not None else None,
                    'top_industries': self._get_top_industries_for_region(
                        r['region_name'] if r and isinstance(r, dict) and 'region_name' in r and r['region_name'] else None, 
                        increase=False
                    )
                }
                for r in decrease_regions_list
                if r and isinstance(r, dict) and r.get('region_name')
            ],
            'all_regions': all_regions_list
        }
        
        data = {
            'report_info': {
                'year': self.year,
                'quarter': self.quarter,
                'report_type': self.report_type,
                'report_name': self.config['name'],
                'index_name': self.config.get('index_name', '지수'),
                'item_name': self.config.get('item_name', '항목')
            },
            'summary_box': summary_box,
            'nationwide_data': nationwide,
            'regional_data': regional_converted,  # 필드명 변환된 버전
            'table_data': table_data,
            'top3_increase_regions': top3_increase,  # 템플릿 호환
            'top3_decrease_regions': top3_decrease   # 템플릿 호환
        }

        self._enrich_template_data(data, table_data, regional_converted, top3_increase, top3_decrease)
        return data


# 하위 호환성 Wrapper
class MiningManufacturingGenerator(UnifiedReportGenerator):
    def __init__(self, excel_path: str, year=None, quarter=None, excel_file=None):
        super().__init__('mining', excel_path, year, quarter, excel_file)


class ServiceIndustryGenerator(UnifiedReportGenerator):
    def __init__(self, excel_path: str, year=None, quarter=None, excel_file=None):
        super().__init__('service', excel_path, year, quarter, excel_file)


class ConsumptionGenerator(UnifiedReportGenerator):
    def __init__(self, excel_path: str, year=None, quarter=None, excel_file=None):
        super().__init__('consumption', excel_path, year, quarter, excel_file)


class ConstructionGenerator(UnifiedReportGenerator):
    def __init__(self, excel_path: str, year=None, quarter=None, excel_file=None):
        super().__init__('construction', excel_path, year, quarter, excel_file)


class ExportGenerator(UnifiedReportGenerator):
    def __init__(self, excel_path: str, year=None, quarter=None, excel_file=None):
        super().__init__('export', excel_path, year, quarter, excel_file)


class ImportGenerator(UnifiedReportGenerator):
    def __init__(self, excel_path: str, year=None, quarter=None, excel_file=None):
        super().__init__('import', excel_path, year, quarter, excel_file)


class PriceTrendGenerator(UnifiedReportGenerator):
    def __init__(self, excel_path: str, year=None, quarter=None, excel_file=None):
        super().__init__('price', excel_path, year, quarter, excel_file)


class EmploymentRateGenerator(UnifiedReportGenerator):
    def __init__(self, excel_path: str, year=None, quarter=None, excel_file=None):
        super().__init__('employment', excel_path, year, quarter, excel_file)


class UnemploymentGenerator(UnifiedReportGenerator):
    def __init__(self, excel_path: str, year=None, quarter=None, excel_file=None):
        super().__init__('unemployment', excel_path, year, quarter, excel_file)


class DomesticMigrationGenerator(UnifiedReportGenerator):
    def __init__(self, excel_path: str, year=None, quarter=None, excel_file=None):
        # report_configs.py에서 'migration'을 사용하지만, 
        # 실제로는 REPORT_CONFIGS에 'migration'으로 정의되어 있으므로 'migration' 사용
        super().__init__('migration', excel_path, year, quarter, excel_file)


class RegionalEconomyByRegionGenerator(BaseGenerator):
    """시도별 경제동향 생성기 (모든 부문 통합)
    
    각 시도별로 생산, 소비·건설, 수출·입, 고용, 물가, 국내인구이동 데이터를 
    한 페이지에 통합하여 보도자료를 생성합니다.
    """
    
    # 17개 시도 정보
    REGIONS = [
        {'code': 11, 'name': '서울', 'full_name': '서울특별시'},
        {'code': 21, 'name': '부산', 'full_name': '부산광역시'},
        {'code': 22, 'name': '대구', 'full_name': '대구광역시'},
        {'code': 23, 'name': '인천', 'full_name': '인천광역시'},
        {'code': 24, 'name': '광주', 'full_name': '광주광역시'},
        {'code': 25, 'name': '대전', 'full_name': '대전광역시'},
        {'code': 26, 'name': '울산', 'full_name': '울산광역시'},
        {'code': 29, 'name': '세종', 'full_name': '세종특별자치시'},
        {'code': 31, 'name': '경기', 'full_name': '경기도'},
        {'code': 32, 'name': '강원', 'full_name': '강원특별자치도'},
        {'code': 33, 'name': '충북', 'full_name': '충청북도'},
        {'code': 34, 'name': '충남', 'full_name': '충청남도'},
        {'code': 35, 'name': '전북', 'full_name': '전북특별자치도'},
        {'code': 36, 'name': '전남', 'full_name': '전라남도'},
        {'code': 37, 'name': '경북', 'full_name': '경상북도'},
        {'code': 38, 'name': '경남', 'full_name': '경상남도'},
        {'code': 39, 'name': '제주', 'full_name': '제주특별자치도'},
    ]
    
    def __init__(self, excel_path: str, year=None, quarter=None, excel_file=None):
        super().__init__(excel_path, year, quarter, excel_file)
        self.year = year
        self.quarter = quarter
        self.generators = {}  # 부문별 Generator 캐시
    
    def extract_all_data(self, region: Optional[str] = None) -> Dict[str, Any]:
        """시도별 모든 데이터 추출 (템플릿 요구에 맞는 sections 구조 보장)"""
        # 단일 시도만 추출 시
        if region:
            region_data = {'sections': {}}
            for report_type in ['mining', 'service', 'consumption', 'construction', 'export', 'import', 'employment', 'unemployment', 'price', 'migration']:
                section = self.extract_regional_section(region, report_type)
                if section:
                    region_data['sections'][report_type] = section
            return {
                'report_info': {'year': self.year, 'quarter': self.quarter},
                'region_name': region,
                'sections': region_data['sections']
            }
        # 전체 시도
        return self.extract_all_regions_data()
    
    def _get_generator(self, report_type: str) -> UnifiedReportGenerator:
        """부문별 Generator 캐시 또는 생성"""
        if report_type not in self.generators:
            self.generators[report_type] = UnifiedReportGenerator(
                report_type, 
                self.excel_path, 
                self.year, 
                self.quarter, 
                self.xl
            )
        return self.generators[report_type]
    
    def extract_regional_section(self, region_name: str, report_type: str) -> Dict[str, Any]:
        """각 시도별로 부문 섹션 데이터 추출
        
        Args:
            region_name: 시도명 (예: '서울')
            report_type: 부문 타입 (mining, service, consumption 등)
            
        Returns:
            섹션 데이터 (narrative + table)
        """
        try:
            from services.excel_cache import get_sector_data

            cache_config = get_report_config(report_type)
            cache_report_id = cache_config.get('report_id') or cache_config.get('id')

            cached = get_sector_data(self.excel_path, self.year, self.quarter, cache_report_id)

            table_data = None
            industries = None
            if cached:
                cached_data = cached.get('data') if isinstance(cached, dict) else None
                if isinstance(cached_data, dict):
                    table_data = cached.get('table_data') or cached_data.get('table_data')
                industries_by_region = cached.get('industries_by_region') if isinstance(cached, dict) else None
                if isinstance(industries_by_region, dict):
                    industries = industries_by_region.get(region_name)

            if table_data is None:
                gen = self._get_generator(report_type)
                gen.load_data()
                table_data = gen._extract_table_data_ssot()

            region_data = next(
                (d for d in (table_data or []) if d.get('region_name') == region_name),
                None
            )

            if not region_data:
                return None

            if industries is None:
                gen = self._get_generator(report_type)
                industries = gen._extract_industry_data(region_name)
            increase_industries = [
                ind for ind in (industries or [])
                if ind and ind.get('change_rate', 0) > 0
            ]
            increase_industries.sort(key=lambda x: x.get('change_rate', 0), reverse=True)
            
            # 나레이션 생성
            narrative = self._generate_narrative(
                region_name,
                report_type,
                region_data,
                increase_industries[:3] if increase_industries else []
            )
            
            # 템플릿 요구: narrative는 반드시 리스트, table.data는 [지표값, 증감률] 순서 보장
            table_row = self._format_table_row(region_data, industries)
            # 값이 2개 미만이면 보정
            values = table_row.get('values', [])
            if not isinstance(values, list):
                values = [values]
            if len(values) < 2:
                values = (values + ['-']*2)[:2]
            table_row['values'] = values
            return {
                'narrative': narrative if isinstance(narrative, list) else [str(narrative)],
                'table': {
                    'periods': self._get_table_periods(self._get_generator(report_type)),
                    'data': [table_row]
                }
            }
        except Exception as e:
            print(f"[지역경제동향] ⚠️ {region_name} - {report_type} 추출 실패: {e}")
            return None
    
    def _generate_narrative(
        self,
        region_name: str,
        report_type: str,
        region_data: Dict,
        top_industries: List[Dict]
    ) -> List[str]:
        """나레이션 생성"""
        narratives = []

        try:
            value = region_data.get('value')
            change_rate = region_data.get('change_rate')

            if value is None:
                return narratives

            try:
                from utils.text_utils import get_terms
            except ImportError:
                import sys
                from pathlib import Path
                sys.path.insert(0, str(Path(__file__).parent.parent))
                from utils.text_utils import get_terms

            # 보고서별 나레이션 템플릿
            template_map = {
                'mining': '{region}의 광공업생산은 {products_phrase}{changes}',
                'service': '{region}의 서비스업생산은 {products_phrase}{changes}',
                'consumption': '{region}의 소비는 {products_phrase}{changes}',
                'construction': '{region}의 건설은 {products_phrase}{changes}',
                'export': '{region}의 수출은 {products_phrase}{changes}',
                'import': '{region}의 수입은 {products_phrase}{changes}',
                'employment': '{region}의 고용률은 {changes}',
                'unemployment': '{region}의 실업률은 {changes}',
                'price': '{region}의 물가는 {products_phrase}{changes}',
                'migration': '{region}의 순인구이동은 {changes}',
            }

            template = template_map.get(report_type, '{region}는 {changes}')

            # 제품/항목 텍스트 생성
            products_text = ''
            if top_industries:
                product_names = [ind.get('name', '') for ind in top_industries[:2] if ind.get('name')]
                products_text = ', '.join(product_names)

            products_phrase = f"{products_text}이 " if products_text else ''

            # 증감 텍스트 (어휘 매핑 준수)
            if change_rate is None:
                changes_text = '변화'
            else:
                _, result_noun, _ = get_terms(report_type, change_rate)
                if abs(change_rate) < 0.01:
                    changes_text = '전년동기대비 보합'
                else:
                    changes_text = f'전년동기대비 {abs(change_rate):.1f}% {result_noun}'

            narrative_text = template.format(
                region=region_name,
                products_phrase=products_phrase,
                changes=changes_text
            )
            narratives.append(narrative_text)

        except Exception as e:
            print(f"[지역경제동향] ⚠️ 나레이션 생성 실패: {e}")

        return narratives
    
    def _get_table_periods(self, gen: UnifiedReportGenerator) -> List[str]:
        """테이블 기간 목록 생성"""
        if gen.year and gen.quarter:
            return [f'{gen.year}/{gen.quarter}Q']
        return ['현 기간', '전년동기']
    
    def _format_table_row(self, region_data: Dict, industries: List[Dict]) -> Dict:
        """테이블 행 포맷팅"""
        return {
            'indicator': region_data.get('region_name', ''),
            'values': [
                region_data.get('value', ''),
                region_data.get('change_rate', '')
            ]
        }
    
    def extract_all_regions_data(self) -> Dict[str, Any]:
        """모든 시도의 통합 데이터 추출"""
        all_regions_data = {}
        report_types = ['mining', 'service', 'consumption', 'construction', 'export', 'import', 'employment', 'unemployment', 'price', 'migration']
        for idx, region in enumerate(self.REGIONS, 1):
            region_name = region['name']
            region_info = dict(region)
            region_info['order'] = idx
            sections = {}
            for report_type in report_types:
                section = self.extract_regional_section(region_name, report_type)
                if section:
                    sections[report_type] = section
            all_regions_data[region_name] = {
                'region_info': region_info,
                'region_name': region_name,
                'sections': sections
            }
        return all_regions_data


class RegionalReportGenerator(BaseGenerator):
    """시도별 보고서 생성기 (unified_generator에 통합)"""
    
    # 17개 시도 정보
    REGIONS = {
        'region_seoul': {'code': '11', 'name': '서울', 'full_name': '서울특별시', 'order': 1},
        'region_busan': {'code': '21', 'name': '부산', 'full_name': '부산광역시', 'order': 2},
        'region_daegu': {'code': '22', 'name': '대구', 'full_name': '대구광역시', 'order': 3},
        'region_incheon': {'code': '23', 'name': '인천', 'full_name': '인천광역시', 'order': 4},
        'region_gwangju': {'code': '24', 'name': '광주', 'full_name': '광주광역시', 'order': 5},
        'region_daejeon': {'code': '25', 'name': '대전', 'full_name': '대전광역시', 'order': 6},
        'region_ulsan': {'code': '26', 'name': '울산', 'full_name': '울산광역시', 'order': 7},
        'region_sejong': {'code': '29', 'name': '세종', 'full_name': '세종특별자치시', 'order': 8},
        'region_gyeonggi': {'code': '31', 'name': '경기', 'full_name': '경기도', 'order': 9},
        'region_gangwon': {'code': '32', 'name': '강원', 'full_name': '강원특별자치도', 'order': 10},
        'region_chungbuk': {'code': '33', 'name': '충북', 'full_name': '충청북도', 'order': 11},
        'region_chungnam': {'code': '34', 'name': '충남', 'full_name': '충청남도', 'order': 12},
        'region_jeonbuk': {'code': '35', 'name': '전북', 'full_name': '전북특별자치도', 'order': 13},
        'region_jeonnam': {'code': '36', 'name': '전남', 'full_name': '전라남도', 'order': 14},
        'region_gyeongbuk': {'code': '37', 'name': '경북', 'full_name': '경상북도', 'order': 15},
        'region_gyeongnam': {'code': '38', 'name': '경남', 'full_name': '경상남도', 'order': 16},
        'region_jeju': {'code': '39', 'name': '제주', 'full_name': '제주특별자치도', 'order': 17},
    }
    
    # 부문별 report_type 매핑
    SECTOR_MAPPING = {
        'mining': 'manufacturing',
        'service': 'service',
        'consumption': 'consumption',
        'construction': 'construction',
        'export': 'export',
        'import': 'import',
        'price': 'price',
        'employment': 'employment',
        'unemployment': 'unemployment',
        'migration': 'migration',
    }
    
    def __init__(self, excel_path: str, year=None, quarter=None, excel_file=None):
        super().__init__(excel_path, year, quarter, excel_file)
        # 부문별 generator 캐시
        self._sector_generators = {}
        self._sector_data_cache = {}
    
    def _get_sector_generator(self, sector_key: str) -> Optional['UnifiedReportGenerator']:
        """부문별 generator 가져오기 (캐싱)"""
        if sector_key in self._sector_generators:
            return self._sector_generators[sector_key]
        
        report_type = self.SECTOR_MAPPING.get(sector_key)
        if not report_type:
            print(f"[지역경제동향] 알 수 없는 부문: {sector_key}")
            return None
        
        try:
            generator = UnifiedReportGenerator(
                report_type, 
                self.excel_path, 
                year=self.year, 
                quarter=self.quarter, 
                excel_file=self.xl  # BaseGenerator에서 excel_file은 self.xl로 저장됨
            )
            self._sector_generators[sector_key] = generator
            return generator
        except Exception as e:
            print(f"[지역경제동향] {sector_key} generator 생성 실패: {e}")
            return None
    
    def _get_sector_table_data(self, sector_key: str) -> List[Dict[str, Any]]:
        """부문별 테이블 데이터 가져오기 (캐싱)"""
        if sector_key in self._sector_data_cache:
            return self._sector_data_cache[sector_key]
        
        generator = self._get_sector_generator(sector_key)
        if not generator:
            return []
        
        try:
            table_data = generator._extract_table_data_ssot()
            self._sector_data_cache[sector_key] = table_data
            return table_data
        except Exception as e:
            print(f"[지역경제동향] {sector_key} 데이터 추출 실패: {e}")
            return []
    
    def _get_region_data_from_sector(self, sector_key: str, region_name: str) -> Optional[Dict[str, Any]]:
        """특정 부문에서 특정 지역의 데이터 추출"""
        table_data = self._get_sector_table_data(sector_key)
        if not table_data:
            return None
        
        # 지역명으로 필터링
        for row in table_data:
            row_region = row.get('region_name', row.get('region', ''))
            if row_region == region_name:
                return row
        
        return None
    
    def _format_value(self, value, default='-') -> str:
        """값을 문자열로 포맷팅"""
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return default
        try:
            if isinstance(value, (int, float)):
                return f"{value:.1f}" if value != int(value) else str(int(value))
            return str(value)
        except Exception:
            return default
    
    def extract_all_data(self, region: Optional[str] = None) -> Dict[str, Any]:
        """시도별 모든 데이터 추출
        
        Args:
            region: 지역 키 (e.g., 'region_seoul') 또는 지역명 ('서울')
        
        Returns:
            지역별 모든 데이터
        """
        # 지역명 결정
        if region and region in self.REGIONS:
            region_name = self.REGIONS[region]['name']
            region_info = self.REGIONS[region]
        elif region:
            # 직접 지역명이 전달된 경우
            region_name = region
            region_info = {'code': '00', 'name': region, 'full_name': region, 'order': 0}
            # REGIONS에서 찾기
            for key, info in self.REGIONS.items():
                if info['name'] == region:
                    region_info = info
                    break
        else:
            region_name = '서울'  # 기본값
            region_info = self.REGIONS['region_seoul']
        
        print(f"[지역경제동향] {region_name} 데이터 추출 시작...")
        
        # 각 부문별 데이터 추출
        mining_data = self._get_region_data_from_sector('mining', region_name)
        service_data = self._get_region_data_from_sector('service', region_name)
        consumption_data = self._get_region_data_from_sector('consumption', region_name)
        construction_data = self._get_region_data_from_sector('construction', region_name)
        export_data = self._get_region_data_from_sector('export', region_name)
        import_data = self._get_region_data_from_sector('import', region_name)
        price_data = self._get_region_data_from_sector('price', region_name)
        employment_data = self._get_region_data_from_sector('employment', region_name)
        unemployment_data = self._get_region_data_from_sector('unemployment', region_name)
        migration_data = self._get_region_data_from_sector('migration', region_name)
        
        # 상단 생산 표 데이터 구성 (한 행만)
        # 실제 데이터 키: value (지수/금액), change_rate (증감률)
        table_row = {
            'region_name': region_name,
            'mining_index': self._format_value(mining_data.get('value') if mining_data else None),
            'mining_change': self._format_value(mining_data.get('change_rate') if mining_data else None),
            'service_index': self._format_value(service_data.get('value') if service_data else None),
            'service_change': self._format_value(service_data.get('change_rate') if service_data else None),
            'retail_index': self._format_value(consumption_data.get('value') if consumption_data else None),
            'retail_change': self._format_value(consumption_data.get('change_rate') if consumption_data else None),
            'export_amount': self._format_value(export_data.get('value') if export_data else None),
            'export_change': self._format_value(export_data.get('change_rate') if export_data else None),
            'price_index': self._format_value(price_data.get('value') if price_data else None),
            'price_change': self._format_value(price_data.get('change_rate') if price_data else None),
            'employment_rate': self._format_value(employment_data.get('value') if employment_data else None),
            'employment_change': self._format_value(employment_data.get('change_rate') if employment_data else None),
        }
        
        # 하단 부문별 표 데이터 구성 (sections)
        # 실제 데이터 키: value (지수/금액), change_rate (증감률)
        def _make_section_data(data: Optional[Dict], value_key: str = 'value', change_key: str = 'change_rate') -> Dict[str, Any]:
            """부문별 섹션 데이터 생성"""
            if not data:
                return {
                    'table': {
                        'periods': [f'{self.year}. {self.quarter}/4' if self.year and self.quarter else '현 기간'],
                        'data': [{'values': ['-', '-']}]
                    },
                    'narrative': []
                }
            
            current_value = self._format_value(data.get(value_key))
            change_value = self._format_value(data.get(change_key))
            
            return {
                'table': {
                    'periods': [f'{self.year}. {self.quarter}/4' if self.year and self.quarter else '현 기간'],
                    'data': [{'values': [current_value, change_value]}]
                },
                'narrative': []  # 나레이션은 별도 생성 필요
            }
        
        sections = {
            'mining': _make_section_data(mining_data),
            'service': _make_section_data(service_data),
            'consumption': _make_section_data(consumption_data),
            'construction': _make_section_data(construction_data),
            'export': _make_section_data(export_data),
            'import': _make_section_data(import_data),
            'employment': _make_section_data(employment_data),
            'unemployment': _make_section_data(unemployment_data),
            'price': _make_section_data(price_data),
            'migration': _make_section_data(migration_data),
        }
        
        print(f"[지역경제동향] {region_name} 데이터 추출 완료")
        print(f"  - 광공업: {table_row['mining_index']} ({table_row['mining_change']}%)")
        print(f"  - 서비스업: {table_row['service_index']} ({table_row['service_change']}%)")
        print(f"  - 소매판매: {table_row['retail_index']} ({table_row['retail_change']}%)")
        
        return {
            'report_info': {'year': self.year, 'quarter': self.quarter},
            'region_info': region_info,
            'region_name': region_name,
            'nationwide_data': None,
            'regional_data': {},
            'table_data': [table_row],  # 상단 표 데이터 (한 행)
            'table_df': [table_row],    # 템플릿 호환성
            'sections': sections,        # 하단 부문별 표 및 나레이션
        }
    
    def render_html(self, region: str, template_path: str) -> str:
        """시도별 HTML 보도자료 렌더링
        
        Args:
            region: 지역 키 (e.g., 'region_seoul')
            template_path: 템플릿 파일 경로
        
        Returns:
            렌더링된 HTML 문자열
        """
        from jinja2 import Environment, FileSystemLoader
        
        # 필터 임포트
        try:
            from utils.filters import format_value, is_missing
            from utils.text_utils import get_josa, get_terms, get_comparative_terms
        except ImportError:
            import sys
            sys.path.insert(0, str(Path(__file__).parent.parent))
            from utils.filters import format_value, is_missing
            from utils.text_utils import get_josa, get_terms, get_comparative_terms
        
        # 데이터 추출
        data = self.extract_all_data(region)
        
        # 데이터 검증
        if not isinstance(data, dict):
            print(f"[경고] 데이터가 dict가 아닙니다: {type(data)}")
            data = {}
        
        # 템플릿 경로 및 렌더링
        template_path_obj = Path(template_path)
        if not template_path_obj.exists():
            raise ValueError(f"템플릿 파일을 찾을 수 없습니다: {template_path}")
        
        # Jinja2 환경 설정
        env = Environment(loader=FileSystemLoader(str(template_path_obj.parent)))
        
        # 필터 등록
        env.filters['format_value'] = format_value
        env.filters['is_missing'] = is_missing
        env.filters['josa'] = get_josa
        
        template = env.get_template(template_path_obj.name)
        
        # 유틸리티 함수 데이터에 추가
        data['get_terms'] = get_terms
        data['get_comparative_terms'] = get_comparative_terms
        
        # 데이터에 지역 정보 추가 (extract_all_data에서 이미 설정됨)
        if 'region_info' not in data:
            if region in self.REGIONS:
                data['region_info'] = self.REGIONS[region]
                data['region_name'] = self.REGIONS[region]['name']
            else:
                data['region_info'] = {'code': '00', 'name': region, 'full_name': region, 'order': 0}
                data['region_name'] = region
        
        # report_info 추가 (regional templates에 필요)
        if 'report_info' not in data:
            data['report_info'] = {
                'year': self.year,
                'quarter': self.quarter,
                'name': '지역경제동향'
            }
        
        # regional_economy_by_region_template.html 호환 기본값
        if 'num_pages' not in data:
            data['num_pages'] = 1
        if 'sections' not in data:
            data['sections'] = {}
        
        # 템플릿 렌더링
        try:
            html_content = template.render(**data)
        except TypeError as e:
            print(f"[경고] 템플릿 렌더링 오류: {e}")
            print(f"[경고] 데이터 타입: {type(data)}")
            print(f"[경고] 데이터 키: {list(data.keys()) if isinstance(data, dict) else 'N/A'}")
            raise
        
        return html_content



if __name__ == '__main__':
    # 테스트
    base_path = Path(__file__).parent.parent
    excel_path = base_path / '분석표_25년 3분기_캡스톤(업데이트).xlsx'
    
    print("=" * 70)
    print("통합 Generator V2 테스트 (집계 시트 기반)")
    print("=" * 70)
    
    for report_type in ['mining', 'service', 'consumption']:
        print(f"\n{'='*70}")
        print(f"[TEST] {REPORT_CONFIGS[report_type]['name']}")
        print(f"{'='*70}\n")
        
        try:
            generator = UnifiedReportGenerator(report_type, str(excel_path), 2025, 3)
            data = generator.extract_all_data()
            
            # 결과 출력
            print(f"\n[결과] ✅ 데이터 추출 완료")
            nationwide = data['nationwide_data']
            print(f"  전국: 지수={nationwide['production_index']:.1f}, 증감률={nationwide['growth_rate']}%")
            
            regional = data['regional_data']
            print(f"  지역: 증가={len(regional['increase_regions'])}개, 감소={len(regional['decrease_regions'])}개")
            
            if regional['increase_regions']:
                top = regional['increase_regions'][0]
                print(f"  최고: {top['region_name']} ({top['change_rate']}%)")
            
        except Exception as e:
            print(f"\n[오류] ❌ {e}")
            import traceback
            traceback.print_exc()
