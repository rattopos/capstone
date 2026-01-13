"""
엑셀 추출기 단위 테스트
헤더 기반 데이터 추출 기능 테스트
"""

import unittest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

# 프로젝트 루트를 경로에 추가
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.excel_extractor import ExcelExtractor
from src.template_manager import TemplateManager
from src.template_filler import TemplateFiller


class TestExcelExtractorHeaderBased(unittest.TestCase):
    """헤더 기반 데이터 추출 테스트"""
    
    def setUp(self):
        """테스트용 엑셀 파일 생성"""
        self.temp_dir = tempfile.mkdtemp()
        self.excel_path = os.path.join(self.temp_dir, "test_data.xlsx")
        
        # 테스트용 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "광공업생산"
        
        # 헤더 행 (1행)
        headers = ["지역", "2023년", "2024년", "2025년"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
        
        # 데이터 행
        data_rows = [
            ["전국", 1000, 1100, 1200],
            ["서울", 500, 550, 600],
            ["부산", 300, 330, 360],
            ["인천", 200, 220, 240],
        ]
        
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx).value = value
        
        wb.save(self.excel_path)
        
        # ExcelExtractor 초기화
        self.extractor = ExcelExtractor(self.excel_path)
        self.extractor.load_workbook()
    
    def tearDown(self):
        """테스트 후 정리"""
        if self.extractor:
            self.extractor.close()
        # 임시 파일 삭제
        if os.path.exists(self.excel_path):
            os.remove(self.excel_path)
        os.rmdir(self.temp_dir)
    
    def test_find_value_with_row_and_col(self):
        """행 키워드와 열 헤더를 모두 지정한 경우"""
        value = self.extractor.find_value(
            sheet_name="광공업생산",
            row_keyword="서울",
            col_header="2024년"
        )
        self.assertEqual(value, 550)
    
    def test_find_value_with_col_only(self):
        """열 헤더만 지정한 경우 (첫 번째 데이터 행 사용)"""
        value = self.extractor.find_value(
            sheet_name="광공업생산",
            col_header="2024년"
        )
        # 첫 번째 데이터 행은 "전국"이므로 1100
        self.assertEqual(value, 1100)
    
    def test_find_value_with_row_only(self):
        """행 키워드만 지정한 경우 (첫 번째 데이터 컬럼 사용)"""
        value = self.extractor.find_value(
            sheet_name="광공업생산",
            row_keyword="서울"
        )
        # 첫 번째 데이터 컬럼은 "지역"이므로 "서울" 문자열
        self.assertEqual(value, "서울")
    
    def test_find_value_not_found(self):
        """존재하지 않는 키워드/헤더"""
        value = self.extractor.find_value(
            sheet_name="광공업생산",
            row_keyword="대전",
            col_header="2024년"
        )
        self.assertIsNone(value)
    
    def test_find_value_type_validation(self):
        """데이터 타입 검증 테스트"""
        import logging
        
        # 로깅 캡처를 위한 핸들러 설정
        log_capture = []
        handler = logging.Handler()
        handler.emit = lambda record: log_capture.append(record)
        logger = logging.getLogger('src.excel_extractor')
        logger.addHandler(handler)
        logger.setLevel(logging.ERROR)
        
        # 숫자가 아닌 값이 있는 셀에 숫자 타입 기대
        # (실제로는 숫자이지만, 테스트를 위해 문자열이 있는 셀을 찾아야 함)
        # 이 테스트는 실제 데이터 구조에 따라 조정 필요
        
        # 정상적인 숫자 값
        value = self.extractor.find_value(
            sheet_name="광공업생산",
            row_keyword="서울",
            col_header="2024년",
            expected_type="number"
        )
        self.assertEqual(value, 550)
        self.assertEqual(len(log_capture), 0)  # 에러 로그 없음
        
        logger.removeHandler(handler)
    
    def test_find_value_similarity_matching(self):
        """유사도 기반 매칭 테스트"""
        # "서울특별시"로 검색해도 "서울"을 찾을 수 있어야 함
        value = self.extractor.find_value(
            sheet_name="광공업생산",
            row_keyword="서울특별시",
            col_header="2024년",
            similarity_threshold=0.5
        )
        self.assertEqual(value, 550)
    
    def test_legacy_cell_address_still_works(self):
        """기존 절대 좌표 방식도 여전히 작동하는지 확인"""
        value = self.extractor.get_cell_value("광공업생산", "B2")
        self.assertEqual(value, 1000)  # 전국, 2023년


class TestTemplateManagerMarkerParsing(unittest.TestCase):
    """템플릿 마커 파싱 테스트"""
    
    def test_parse_cell_address_marker(self):
        """기존 셀 주소 마커 파싱"""
        manager = TemplateManager()
        manager.template_content = "값: {광공업생산:A1}"
        
        markers = manager.extract_markers()
        self.assertEqual(len(markers), 1)
        self.assertEqual(markers[0]['marker_type'], 'cell_address')
        self.assertEqual(markers[0]['cell_address'], 'A1')
        self.assertIsNone(markers[0]['row_keyword'])
        self.assertIsNone(markers[0]['col_header'])
    
    def test_parse_header_based_marker(self):
        """헤더 기반 마커 파싱"""
        manager = TemplateManager()
        manager.template_content = "값: {광공업생산:서울:2024년}"
        
        markers = manager.extract_markers()
        self.assertEqual(len(markers), 1)
        self.assertEqual(markers[0]['marker_type'], 'header_based')
        self.assertEqual(markers[0]['row_keyword'], '서울')
        self.assertEqual(markers[0]['col_header'], '2024년')
        self.assertIsNone(markers[0]['cell_address'])
    
    def test_parse_header_based_marker_with_operation(self):
        """헤더 기반 마커 + 계산식 파싱"""
        manager = TemplateManager()
        manager.template_content = "합계: {광공업생산:서울:2024년:sum}"
        
        markers = manager.extract_markers()
        self.assertEqual(len(markers), 1)
        self.assertEqual(markers[0]['marker_type'], 'header_based')
        self.assertEqual(markers[0]['row_keyword'], '서울')
        self.assertEqual(markers[0]['col_header'], '2024년')
        self.assertEqual(markers[0]['operation'], 'sum')
    
    def test_parse_cell_address_with_operation(self):
        """셀 주소 마커 + 계산식 파싱"""
        manager = TemplateManager()
        manager.template_content = "합계: {광공업생산:A1:A5:sum}"
        
        markers = manager.extract_markers()
        self.assertEqual(len(markers), 1)
        self.assertEqual(markers[0]['marker_type'], 'cell_address')
        self.assertEqual(markers[0]['cell_address'], 'A1:A5')
        self.assertEqual(markers[0]['operation'], 'sum')
    
    def test_parse_dynamic_key_marker(self):
        """동적 키 마커 파싱 (기존 방식)"""
        manager = TemplateManager()
        manager.template_content = "값: {광공업생산:전국_증감률}"
        
        markers = manager.extract_markers()
        self.assertEqual(len(markers), 1)
        self.assertEqual(markers[0]['marker_type'], 'dynamic_key')
        self.assertEqual(markers[0]['cell_address'], '전국_증감률')
        self.assertIsNone(markers[0]['row_keyword'])
        self.assertIsNone(markers[0]['col_header'])


class TestTemplateFillerIntegration(unittest.TestCase):
    """템플릿 필러 통합 테스트"""
    
    def setUp(self):
        """테스트용 엑셀 파일 및 템플릿 생성"""
        self.temp_dir = tempfile.mkdtemp()
        self.excel_path = os.path.join(self.temp_dir, "test_data.xlsx")
        
        # 테스트용 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "광공업생산"
        
        # 헤더 행
        headers = ["지역", "2023년", "2024년", "2025년"]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx).value = header
        
        # 데이터 행
        data_rows = [
            ["전국", 1000, 1100, 1200],
            ["서울", 500, 550, 600],
        ]
        
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx).value = value
        
        wb.save(self.excel_path)
        
        # ExcelExtractor 초기화
        self.extractor = ExcelExtractor(self.excel_path)
        self.extractor.load_workbook()
    
    def tearDown(self):
        """테스트 후 정리"""
        if self.extractor:
            self.extractor.close()
        if os.path.exists(self.excel_path):
            os.remove(self.excel_path)
        os.rmdir(self.temp_dir)
    
    def test_template_filler_with_header_based_marker(self):
        """헤더 기반 마커를 사용한 템플릿 채우기"""
        template_content = """
        <html>
        <body>
            <p>서울의 2024년 값: {광공업생산:서울:2024년}</p>
        </body>
        </html>
        """
        
        manager = TemplateManager()
        manager.template_content = template_content
        
        filler = TemplateFiller(manager, self.extractor)
        result = filler.fill_template(sheet_name="광공업생산", year=2024, quarter=2)
        
        # 마커가 값으로 치환되었는지 확인
        self.assertNotIn("{광공업생산:서울:2024년}", result)
        self.assertIn("550", result)  # 포맷팅된 값이 포함되어야 함
    
    def test_template_filler_with_legacy_marker(self):
        """기존 셀 주소 마커를 사용한 템플릿 채우기 (하위 호환성)"""
        template_content = """
        <html>
        <body>
            <p>셀 B2의 값: {광공업생산:B2}</p>
        </body>
        </html>
        """
        
        manager = TemplateManager()
        manager.template_content = template_content
        
        filler = TemplateFiller(manager, self.extractor)
        result = filler.fill_template(sheet_name="광공업생산", year=2024, quarter=2)
        
        # 마커가 값으로 치환되었는지 확인
        self.assertNotIn("{광공업생산:B2}", result)
        self.assertIn("1000", result)  # 포맷팅된 값이 포함되어야 함
    
    def test_template_filler_with_mixed_markers(self):
        """혼합된 마커 타입 사용"""
        template_content = """
        <html>
        <body>
            <p>헤더 기반: {광공업생산:서울:2024년}</p>
            <p>셀 주소: {광공업생산:B2}</p>
        </body>
        </html>
        """
        
        manager = TemplateManager()
        manager.template_content = template_content
        
        filler = TemplateFiller(manager, self.extractor)
        result = filler.fill_template(sheet_name="광공업생산", year=2024, quarter=2)
        
        # 두 마커 모두 치환되었는지 확인
        self.assertNotIn("{광공업생산:서울:2024년}", result)
        self.assertNotIn("{광공업생산:B2}", result)
        self.assertIn("550", result)  # 헤더 기반 값
        self.assertIn("1000", result)  # 셀 주소 값


if __name__ == '__main__':
    unittest.main()
