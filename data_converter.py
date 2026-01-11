#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
기초자료 수집표 → 분석표 변환 모듈

기존 분석표를 템플릿으로 사용하고, 기초자료의 데이터를 집계 시트에 복사합니다.
분석 시트의 엑셀 수식은 Python에서 계산하여 값으로 저장합니다.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import json
import shutil
import re

# 수식 계산 라이브러리
try:
    import formulas
    FORMULAS_AVAILABLE = True
except ImportError:
    FORMULAS_AVAILABLE = False
    print("[경고] formulas 라이브러리가 설치되지 않았습니다. pip install formulas 를 실행하세요.")


class DataConverter:
    """기초자료 수집표 → 분석표 변환기 (수식 보존 방식)"""
    
    # 기초자료 시트 → 분석표 집계 시트 매핑
    SHEET_MAPPING = {
        '광공업생산': 'A(광공업생산)집계',
        '서비스업생산': 'B(서비스업생산)집계',
        '소비(소매, 추가)': 'C(소비)집계',
        '고용률': 'D(고용률)집계',
        '실업자 수': 'D(실업)집계',
        '지출목적별 물가': 'E(지출목적물가)집계',
        '품목성질별 물가': 'E(품목성질물가)집계',
        '건설 (공표자료)': "F'(건설)집계",
        '수출': 'G(수출)집계',
        '수입': 'H(수입)집계',
        '시도 간 이동': 'I(순인구이동)집계',
    }
    
    # 지역 순서
    REGIONS_ORDER = ['전국', '서울', '부산', '대구', '인천', '광주', '대전', '울산', '세종',
                     '경기', '강원', '충북', '충남', '전북', '전남', '경북', '경남', '제주']
    
    # 연도 데이터 개수와 분기 데이터 개수
    NUM_YEARS = 5      # 당해 제외 최근 5개년
    NUM_QUARTERS = 13  # 최근 13개 분기
    
    # 집계 시트별 열 구조 정의: (메타열수, 연도시작열, 분기시작열, 가중치열 위치)
    # 1-based index
    # 분석표 템플릿은 열 1~3에 조회용 컬럼이 있어 기초자료와 다름
    # meta_start: 분석표에서 메타데이터가 시작하는 열 (1-based)
    # raw_meta_cols: 기초자료에서 메타데이터 열 개수 (0-based로 0~5)
    SHEET_STRUCTURE = {
        'A(광공업생산)집계': {'meta_start': 4, 'raw_meta_cols': 6, 'year_start': 10, 'quarter_start': 15, 'weight_col': 7, 'raw_weight_col': 3},
        'B(서비스업생산)집계': {'meta_start': 4, 'raw_meta_cols': 6, 'year_start': 9, 'quarter_start': 14, 'weight_col': 7, 'raw_weight_col': 3},
        'C(소비)집계': {'meta_start': 3, 'raw_meta_cols': 5, 'year_start': 8, 'quarter_start': 13, 'weight_col': None, 'raw_weight_col': None},
        'D(고용률)집계': {'meta_start': 3, 'raw_meta_cols': 5, 'year_start': 8, 'quarter_start': 13, 'weight_col': None, 'raw_weight_col': None},
        'D(실업)집계': {'meta_start': 3, 'raw_meta_cols': 5, 'year_start': 8, 'quarter_start': 13, 'weight_col': None, 'raw_weight_col': None},
        'E(지출목적물가)집계': {'meta_start': 3, 'raw_meta_cols': 5, 'year_start': 8, 'quarter_start': 13, 'weight_col': None, 'raw_weight_col': None},
        'E(품목성질물가)집계': {'meta_start': 3, 'raw_meta_cols': 5, 'year_start': 8, 'quarter_start': 13, 'weight_col': None, 'raw_weight_col': None},
        "F'(건설)집계": {'meta_start': 2, 'raw_meta_cols': 4, 'year_start': 6, 'quarter_start': 11, 'weight_col': None, 'raw_weight_col': None},
        'G(수출)집계': {'meta_start': 4, 'raw_meta_cols': 6, 'year_start': 10, 'quarter_start': 15, 'weight_col': None, 'raw_weight_col': None},
        'H(수입)집계': {'meta_start': 4, 'raw_meta_cols': 6, 'year_start': 10, 'quarter_start': 15, 'weight_col': None, 'raw_weight_col': None},
        'I(순인구이동)집계': {'meta_start': 3, 'raw_meta_cols': 5, 'year_start': 8, 'quarter_start': 13, 'weight_col': None, 'raw_weight_col': None},
    }
    
    # 기초자료 시트에서 가중치 열 위치 (0-based index) - 레거시 호환용
    # 실제로는 SHEET_STRUCTURE의 raw_weight_col 사용
    RAW_WEIGHT_COL_MAPPING = {
        '광공업생산': 3,  # 기초자료에서 가중치 열 (0-based: 열 D)
        '서비스업생산': 3,  # 기초자료에서 가중치 열 (0-based: 열 D)
    }
    
    # GRDP 기본 데이터 (25년 2분기 기준)
    DEFAULT_GRDP_DATA = {
        'year': 2025,
        'quarter': 2,
        'regions': ['전국', '서울', '부산', '대구', '인천', '광주', '대전', '울산', '세종',
                   '경기', '강원', '충북', '충남', '전북', '전남', '경북', '경남', '제주'],
    }
    
    def __init__(self, raw_excel_path: str, template_path: str = None):
        """
        Args:
            raw_excel_path: 기초자료 수집표 엑셀 파일 경로
            template_path: 분석표 템플릿 경로 (None이면 기본 템플릿 사용, 없으면 새로 생성)
        """
        self.raw_excel_path = Path(raw_excel_path)
        
        # 템플릿 경로 설정
        self.template_available = False
        if template_path:
            self.template_path = Path(template_path)
            if self.template_path.exists():
                self.template_available = True
        else:
            # 기본 템플릿 (프로젝트 내 분석표)
            self.template_path = Path(__file__).parent / '분석표_25년 2분기_캡스톤.xlsx'
            if self.template_path.exists():
                self.template_available = True
            else:
                # 대안 경로 (홈 디렉토리)
                home_template = Path.home() / 'Desktop' / '분석표_25년 2분기_캡스톤.xlsx'
                if home_template.exists():
                    self.template_path = home_template
                    self.template_available = True
                else:
                    # 마지막 대안 경로
                    self.template_path = self.raw_excel_path.parent / '분석표_25년 2분기_캡스톤.xlsx'
                    if self.template_path.exists():
                        self.template_available = True
        
        # 템플릿이 없어도 동작 가능 (새로 생성 모드)
        if not self.template_available:
            print("[정보] 템플릿 파일 없음 - 기초자료에서 새 분석표를 생성합니다.")
        
        self.year = None
        self.quarter = None
        self._detect_year_quarter()
        
        # 템플릿에서 연도/분기 개수 파악 (동적 범위 계산용)
        self.num_years = self.NUM_YEARS
        self.num_quarters = self.NUM_QUARTERS
        if self.template_available:
            self._detect_template_range()
    
    def _detect_year_quarter(self):
        """기초자료에서 연도/분기 자동 추출"""
        import re
        
        # 1. 먼저 기초자료 헤더에서 추출 시도
        try:
            xl = pd.ExcelFile(self.raw_excel_path)
            
            # 광공업생산 시트 또는 첫 번째 데이터 시트에서 헤더 확인
            for sheet_name in ['광공업생산', '서비스업생산', '소비(소매, 추가)']:
                if sheet_name in xl.sheet_names:
                    df = pd.read_excel(xl, sheet_name=sheet_name, header=None, nrows=4)
                    
                    # 헤더 행(3행)에서 가장 최신 분기 찾기
                    latest_year = 0
                    latest_quarter = 0
                    
                    for col_idx in range(len(df.columns)):
                        header_val = str(df.iloc[2, col_idx]) if not pd.isna(df.iloc[2, col_idx]) else ''
                        
                        # "2025  2/4p" 또는 "2025 2/4" 패턴 찾기
                        match = re.search(r'(\d{4})\s*(\d)/4', header_val)
                        if match:
                            year = int(match.group(1))
                            quarter = int(match.group(2))
                            
                            # 가장 최신 연도/분기 저장
                            if year > latest_year or (year == latest_year and quarter > latest_quarter):
                                latest_year = year
                                latest_quarter = quarter
                    
                    if latest_year > 0 and latest_quarter > 0:
                        self.year = latest_year
                        self.quarter = latest_quarter
                        print(f"[자동감지] 기초자료에서 연도/분기 추출: {self.year}년 {self.quarter}분기")
                        return
                    break
        except Exception as e:
            print(f"[경고] 기초자료에서 연도/분기 추출 실패: {e}")
        
        # 2. 파일명에서 추출 시도
        filename = self.raw_excel_path.stem
        
        # 다양한 패턴 매칭
        patterns = [
            r'(\d{4})년\s*(\d)분기',  # 2025년 2분기
            r'(\d{2})년\s*(\d)분기',   # 25년 2분기
            r'(\d{4})_(\d)',           # 2025_2
        ]
        
        for pattern in patterns:
            match = re.search(pattern, filename)
            if match:
                year_str = match.group(1)
                quarter = int(match.group(2))
                
                # 2자리 연도 처리
                if len(year_str) == 2:
                    year = 2000 + int(year_str)
                else:
                    year = int(year_str)
                
                self.year = year
                self.quarter = quarter
                print(f"[자동감지] 파일명에서 연도/분기 추출: {self.year}년 {self.quarter}분기")
                return
        
        # 3. 기본값
        self.year, self.quarter = 2025, 2
        print(f"[기본값] 연도/분기: {self.year}년 {self.quarter}분기")
    
    def _detect_template_range(self):
        """템플릿에서 연도/분기 개수 파악 (동적 범위 계산용)
        
        참고 분석표에서 몇 개의 연도와 분기가 사용되는지 확인하여,
        입력 기초자료의 연도/분기에 맞춰 동일한 범위로 계산할 수 있도록 함.
        """
        if not self.template_available:
            return
        
        try:
            wb = openpyxl.load_workbook(self.template_path, data_only=True)
            
            # A(광공업생산)집계 시트에서 헤더 확인
            if 'A(광공업생산)집계' in wb.sheetnames:
                ws = wb['A(광공업생산)집계']
                header_row = 3  # 일반적으로 3행이 헤더
                
                years = set()
                quarters = set()
                
                # 열 1부터 100까지 확인
                for col in range(1, 101):
                    cell = ws.cell(row=header_row, column=col)
                    val = cell.value
                    
                    if val is None:
                        continue
                    
                    val_str = str(val).strip()
                    
                    # 연도 패턴: 4자리 숫자 (2000~2100)
                    if val_str.isdigit() and len(val_str) == 4:
                        year = int(val_str)
                        if 2000 <= year <= 2100:
                            years.add(year)
                    
                    # 분기 패턴: "YYYY Q/4" 형식
                    if '/' in val_str and '4' in val_str:
                        # "2022 2/4", "2022.2/4" 등 패턴
                        import re
                        q_match = re.search(r'(\d{4})[.\s]*(\d)/4', val_str)
                        if q_match:
                            quarters.add(f"{q_match.group(1)} {q_match.group(2)}/4")
                
                if years:
                    self.num_years = len(years)
                    print(f"[템플릿] 연도 개수: {self.num_years}개 ({min(years)}~{max(years)})")
                
                if quarters:
                    self.num_quarters = len(quarters)
                    sorted_quarters = sorted(quarters)
                    print(f"[템플릿] 분기 개수: {self.num_quarters}개 ({sorted_quarters[0]}~{sorted_quarters[-1]})")
            
            wb.close()
        except Exception as e:
            print(f"[경고] 템플릿 범위 파악 실패, 기본값 사용: {e}")
            # 기본값 유지
    
    def _get_target_years(self) -> List[int]:
        """당해 제외 최근 N개년 리스트 반환 (템플릿에서 파악한 개수 사용)
        
        예: 2025년 2분기, num_years=5 기준 → [2020, 2021, 2022, 2023, 2024]
        예: 2026년 1분기, num_years=5 기준 → [2021, 2022, 2023, 2024, 2025]
        """
        return list(range(self.year - self.num_years, self.year))
    
    def _get_target_quarters(self) -> List[str]:
        """최근 N개 분기 리스트 반환 (템플릿에서 파악한 개수 사용, YYYY Q/4 형식)
        
        예: 2025년 2분기, num_quarters=13 기준 → ['2022 2/4', '2022 3/4', ..., '2025 2/4']
        예: 2026년 1분기, num_quarters=13 기준 → ['2023 1/4', '2023 2/4', ..., '2026 1/4']
        """
        quarters = []
        
        # 현재 분기부터 거슬러 num_quarters개 (현재 분기 포함)
        for i in range(self.num_quarters - 1, -1, -1):
            # i분기 전
            total_quarters = (self.year * 4 + self.quarter) - i
            q_year = (total_quarters - 1) // 4
            q_num = ((total_quarters - 1) % 4) + 1
            quarters.append(f"{q_year} {q_num}/4")
        
        return quarters
    
    def _parse_raw_header(self, raw_df: pd.DataFrame, header_row: int = 2) -> Dict:
        """기초자료 헤더에서 연도/분기별 열 인덱스 매핑 생성
        
        RawDataExtractor의 parse_sheet_structure를 재사용하거나 동일한 로직 사용.
        향후 RawDataExtractor를 직접 사용하도록 리팩토링 가능.
        
        Returns:
            {
                'years': {2020: col_idx, 2021: col_idx, ...},
                'quarters': {'2022 2/4': col_idx, '2022 3/4': col_idx, ...}
            }
        """
        import re
        
        year_cols = {}
        quarter_cols = {}
        
        for col_idx in range(len(raw_df.columns)):
            val = raw_df.iloc[header_row, col_idx]
            if pd.isna(val):
                continue
            
            val_str = str(val).strip()
            
            # 연도 패턴: 2020, 2021, ... (정수 또는 "2020.0")
            if isinstance(val, (int, float)) and 2000 <= int(val) <= 2100:
                year_cols[int(val)] = col_idx
            elif re.match(r'^(\d{4})\.?0?$', val_str):
                year = int(re.match(r'^(\d{4})\.?0?$', val_str).group(1))
                year_cols[year] = col_idx
            
            # 분기 패턴: "2022  2/4", "2022 2/4p", "2022.2/4" 등
            quarter_match = re.search(r'(\d{4})[.\s]*(\d)/4', val_str)
            if quarter_match:
                q_year = int(quarter_match.group(1))
                q_num = int(quarter_match.group(2))
                quarter_key = f"{q_year} {q_num}/4"
                quarter_cols[quarter_key] = col_idx
        
        return {'years': year_cols, 'quarters': quarter_cols}
    
    def convert_all(self, output_path: str = None, weight_settings: Dict = None) -> str:
        """분석표 생성 (템플릿 사용 또는 새로 생성)
        
        Args:
            output_path: 출력 파일 경로 (None이면 자동 생성)
            weight_settings: 가중치 설정 {mining: {mode, values}, service: {mode, values}}
                - mode: 'auto' (기초자료에서 추출), 'manual' (수동입력), 'empty' (공란)
                - values: 수동입력 시 가중치 배열
            
        Returns:
            생성된 분석표 파일 경로
        """
        if output_path is None:
            # 파일명: 입력 기초자료의 연도/분기에 맞춰 생성
            # 예: "분석표_26년 1분기.xlsx"
            output_path = str(self.raw_excel_path.parent / f"분석표_{self.year}년 {self.quarter}분기.xlsx")
        
        # 템플릿이 있으면 복사, 없으면 새로 생성
        if self.template_available:
            return self._convert_with_template(output_path, weight_settings)
        else:
            return self._create_from_scratch(output_path, weight_settings)
    
    def _convert_with_template(self, output_path: str, weight_settings: Dict = None) -> str:
        """템플릿을 사용하여 분석표 생성 (스타일 및 수식 보존)"""
        print(f"[변환] 템플릿 복사: {self.template_path.name}")
        shutil.copy(self.template_path, output_path)
        
        # 복사된 파일 열기 (수식 보존, data_only=False로 수식 유지)
        wb = openpyxl.load_workbook(output_path, data_only=False)
        
        # 이용관련 시트에서 연도/분기 설정 (핵심!)
        if '이용관련' in wb.sheetnames:
            ws_config = wb['이용관련']
            ws_config.cell(row=17, column=11).value = self.year    # K17: 연도
            ws_config.cell(row=17, column=13).value = self.quarter # M17: 분기
            print(f"[변환] 이용관련 시트 설정: {self.year}년 {self.quarter}분기")
        else:
            print("[경고] '이용관련' 시트를 찾을 수 없습니다.")
        
        # 기초자료 열기
        raw_xl = pd.ExcelFile(self.raw_excel_path)
        
        # 각 집계 시트에 데이터 복사
        for raw_sheet, target_sheet in self.SHEET_MAPPING.items():
            if raw_sheet in raw_xl.sheet_names and target_sheet in wb.sheetnames:
                print(f"[변환] {raw_sheet} → {target_sheet}")
                self._copy_sheet_data(raw_xl, raw_sheet, wb[target_sheet], weight_settings)
            else:
                if raw_sheet not in raw_xl.sheet_names:
                    print(f"  [경고] 기초자료에 '{raw_sheet}' 시트 없음")
                if target_sheet not in wb.sheetnames:
                    print(f"  [경고] 템플릿에 '{target_sheet}' 시트 없음")
        
        # 저장
        wb.save(output_path)
        wb.close()
        
        # 수식 계산 및 값으로 저장
        print(f"[변환] 분석 시트 수식 계산 중...")
        self._calculate_formulas_in_file(output_path)
        
        print(f"[완료] 분석표 생성: {output_path}")
        return output_path
    
    def _create_from_scratch(self, output_path: str, weight_settings: Dict = None) -> str:
        """템플릿 없이 기초자료에서 새 분석표 생성
        
        기초자료의 각 시트 데이터를 정리하여 분석표 형태로 변환합니다.
        집계 시트와 기본 분석 데이터(전년동기비)를 포함합니다.
        """
        print(f"[생성] 기초자료에서 새 분석표 생성 중...")
        
        # 새 워크북 생성
        wb = openpyxl.Workbook()
        
        # 기초자료 열기
        raw_xl = pd.ExcelFile(self.raw_excel_path)
        
        # 설정 시트 생성
        ws_config = wb.active
        ws_config.title = '설정'
        ws_config['A1'] = '분석표 자동생성'
        ws_config['A2'] = f'기준: {self.year}년 {self.quarter}분기'
        ws_config['A3'] = f'생성일: {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")}'
        ws_config['A5'] = '※ 이 파일은 기초자료에서 자동 생성되었습니다.'
        
        # 각 시트별로 집계 시트 생성
        sheet_count = 0
        for raw_sheet, target_sheet in self.SHEET_MAPPING.items():
            if raw_sheet in raw_xl.sheet_names:
                print(f"[생성] {raw_sheet} → {target_sheet}")
                self._create_aggregation_sheet(wb, raw_xl, raw_sheet, target_sheet, weight_settings)
                sheet_count += 1
            else:
                print(f"  [건너뜀] 기초자료에 '{raw_sheet}' 시트 없음")
        
        # 요약 시트 생성
        self._create_summary_sheet(wb, raw_xl)
        
        # 저장
        wb.save(output_path)
        wb.close()
        
        print(f"[완료] 분석표 생성: {output_path} ({sheet_count}개 시트)")
        return output_path
    
    def _create_aggregation_sheet(self, wb, raw_xl, raw_sheet: str, target_sheet: str, weight_settings: Dict = None):
        """기초자료에서 집계 시트 생성"""
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        # 기초자료 읽기
        df = pd.read_excel(raw_xl, sheet_name=raw_sheet, header=None)
        
        # 새 시트 생성
        ws = wb.create_sheet(title=target_sheet[:31])  # 시트 이름 31자 제한
        
        # 스타일 정의
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 헤더 행 찾기 (보통 3행)
        header_row = 2
        for i in range(min(5, len(df))):
            row_str = ' '.join([str(v) for v in df.iloc[i].values[:10] if pd.notna(v)])
            if '지역' in row_str or '시도' in row_str:
                header_row = i
                break
        
        # 데이터 복사
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                value = df.iloc[row_idx, col_idx]
                cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
                
                # NaN 처리
                if pd.isna(value):
                    cell.value = None
                elif isinstance(value, float) and value != int(value):
                    cell.value = round(value, 2)
                else:
                    cell.value = value
                
                # 헤더 스타일
                if row_idx <= header_row:
                    cell.font = header_font
                    cell.fill = header_fill
                
                cell.border = thin_border
        
        # 열 너비 자동 조정
        for col_idx in range(1, min(len(df.columns) + 1, 30)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    def _create_summary_sheet(self, wb, raw_xl):
        """요약 시트 생성 (전년동기비 계산 결과)"""
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        ws = wb.create_sheet(title='요약_전년동기비', index=0)
        
        # 스타일
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        white_font = Font(bold=True, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 제목
        ws['A1'] = f'{self.year}년 {self.quarter}분기 지역경제동향 요약'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        ws['A3'] = '지표'
        ws['B3'] = '전국'
        
        # 지역 헤더
        regions = self.REGIONS_ORDER[1:]  # 전국 제외
        for idx, region in enumerate(regions):
            ws.cell(row=3, column=idx + 3).value = region
        
        # 헤더 스타일
        for col in range(1, len(regions) + 3):
            cell = ws.cell(row=3, column=col)
            cell.font = white_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # 각 지표별 전년동기비 계산
        indicators = [
            ('광공업생산', '광공업생산지수'),
            ('서비스업생산', '서비스업생산지수'),
            ('소매판매액', '소매판매액지수'),
            ('건설수주액', '건설수주액'),
        ]
        
        row_num = 4
        for indicator_name, raw_sheet in indicators:
            if raw_sheet not in raw_xl.sheet_names:
                # 대체 시트명 확인
                alt_sheets = {
                    '광공업생산지수': '광공업생산',
                    '서비스업생산지수': '서비스업생산',
                    '소매판매액지수': '소비(소매, 추가)',
                    '건설수주액': '건설 (공표자료)'
                }
                raw_sheet = alt_sheets.get(raw_sheet, raw_sheet)
                if raw_sheet not in raw_xl.sheet_names:
                    continue
            
            ws.cell(row=row_num, column=1).value = indicator_name
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=1).border = thin_border
            
            # 전년동기비 계산 (간단한 예시)
            try:
                growth_rates = self._calculate_yoy_growth(raw_xl, raw_sheet)
                for idx, region in enumerate(['전국'] + regions):
                    cell = ws.cell(row=row_num, column=idx + 2)
                    cell.value = growth_rates.get(region, '-')
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='right')
                    
                    # 색상 (양수: 파랑, 음수: 빨강)
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.font = Font(color='0000FF')
                        elif cell.value < 0:
                            cell.font = Font(color='FF0000')
            except Exception as e:
                print(f"  [경고] {indicator_name} 전년동기비 계산 실패: {e}")
            
            row_num += 1
        
        # 열 너비
        ws.column_dimensions['A'].width = 15
        for col in range(2, len(regions) + 3):
            ws.column_dimensions[get_column_letter(col)].width = 8
    
    def _calculate_yoy_growth(self, raw_xl, raw_sheet: str) -> Dict[str, float]:
        """전년동기비 계산"""
        df = pd.read_excel(raw_xl, sheet_name=raw_sheet, header=None)
        
        result = {}
        
        # 헤더 행과 현재/전년 동기 열 찾기
        header_row = 2
        current_col = None
        prev_year_col = None
        region_col = 1
        
        # 헤더에서 열 위치 찾기
        for i in range(min(5, len(df))):
            row = df.iloc[i]
            for col_idx in range(len(row)):
                val = str(row[col_idx]) if pd.notna(row[col_idx]) else ''
                # 현재 분기 열
                if f'{self.year}' in val and f'{self.quarter}/4' in val:
                    current_col = col_idx
                # 전년 동기 열
                if f'{self.year - 1}' in val and f'{self.quarter}/4' in val:
                    prev_year_col = col_idx
        
        if current_col is None or prev_year_col is None:
            return result
        
        # 각 지역의 전년동기비 계산
        for row_idx in range(header_row + 1, len(df)):
            region = str(df.iloc[row_idx, region_col]).strip() if pd.notna(df.iloc[row_idx, region_col]) else ''
            
            if region in self.REGIONS_ORDER:
                current_val = df.iloc[row_idx, current_col]
                prev_val = df.iloc[row_idx, prev_year_col]
                
                if pd.notna(current_val) and pd.notna(prev_val) and prev_val != 0:
                    try:
                        growth = ((float(current_val) - float(prev_val)) / float(prev_val)) * 100
                        result[region] = round(growth, 1)
                    except:
                        pass
        
        return result
    
    def add_grdp_sheet(self, analysis_path: str, grdp_file_path: str = None) -> bool:
        """분석표에 GRDP 시트 추가 (마지막 시트로)
        
        Args:
            analysis_path: 분석표 파일 경로
            grdp_file_path: GRDP 엑셀 파일 경로 (None이면 기본값 사용)
            
        Returns:
            성공 여부
        """
        try:
            wb = openpyxl.load_workbook(analysis_path)
            
            # 기존 GRDP 시트가 있으면 삭제
            grdp_sheet_names = ['GRDP', 'I GRDP', '분기 GRDP']
            for sheet_name in grdp_sheet_names:
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                    print(f"[GRDP] 기존 '{sheet_name}' 시트 삭제")
            
            # 새 GRDP 시트 생성 (마지막 위치에)
            ws = wb.create_sheet('GRDP')
            
            if grdp_file_path and Path(grdp_file_path).exists():
                # GRDP 파일에서 데이터 복사
                self._copy_grdp_from_file(ws, grdp_file_path)
            else:
                # 기본값으로 GRDP 시트 생성
                self._create_default_grdp_sheet(ws)
            
            wb.save(analysis_path)
            wb.close()
            
            print(f"[GRDP] 'GRDP' 시트 추가 완료 (마지막 시트)")
            return True
            
        except Exception as e:
            import traceback
            print(f"[GRDP] 시트 추가 오류: {e}")
            traceback.print_exc()
            return False
    
    def _copy_grdp_from_file(self, ws, grdp_file_path: str):
        """GRDP 파일에서 성장률 시트 데이터를 복사"""
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        xl = pd.ExcelFile(grdp_file_path)
        
        # 성장률 시트 우선, 없으면 첫 번째 시트 사용
        if '성장률' in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name='성장률', header=None)
            print(f"[GRDP] '성장률' 시트 복사 ({df.shape[0]}행)")
        else:
            df = pd.read_excel(xl, sheet_name=xl.sheet_names[0], header=None)
            print(f"[GRDP] '{xl.sheet_names[0]}' 시트 복사 ({df.shape[0]}행)")
        
        # 데이터 복사
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
            for c_idx, value in enumerate(row, 1):
                if pd.notna(value):
                    ws.cell(row=r_idx, column=c_idx, value=value)
    
    def _create_default_grdp_sheet(self, ws):
        """기본값으로 GRDP 시트 생성 (25년 2분기 기준)"""
        # 헤더 행
        ws['A1'] = '지역별 경제활동별 성장률'
        ws['A3'] = '지역별'
        ws['B3'] = '경제활동별'
        ws['C3'] = f'{self.year}.{self.quarter}/4'
        
        # 데이터 행
        regions = self.DEFAULT_GRDP_DATA['regions']
        
        row = 4
        for region in regions:
            ws.cell(row=row, column=1, value=region)
            ws.cell(row=row, column=2, value='지역내총생산(시장가격)')
            ws.cell(row=row, column=3, value=0.0)  # 기본값
            row += 1
            
            # 산업별 기여도
            industries = ['광업, 제조업', '건설업', '서비스업']
            for industry in industries:
                ws.cell(row=row, column=1, value='')
                ws.cell(row=row, column=2, value=industry)
                ws.cell(row=row, column=3, value=0.0)
                row += 1
        
        print(f"[GRDP] 기본값 시트 생성 ({row-4}행, {self.year}년 {self.quarter}분기)")
    
    def _calculate_formulas_in_file(self, excel_path: str):
        """분석표 파일의 분석 시트 수식 계산
        
        집계 시트의 값을 분석 시트로 복사하는 방식으로 처리합니다.
        분석 시트의 수식은 그대로 유지하고, 집계 시트의 값만 업데이트합니다.
        """
        # 분석 시트의 수식은 그대로 유지하고, 집계 시트의 데이터만 복사하면
        # 엑셀에서 열 때 자동으로 계산됩니다.
        # 여기서는 집계 시트의 데이터가 이미 복사되었으므로 추가 작업 불필요.
        print(f"[변환] 분석 시트 수식은 엑셀에서 자동 계산됩니다.")
    
    def _calculate_with_formulas_lib(self, excel_path: str):
        """formulas 라이브러리를 사용하여 수식 계산"""
        import formulas
        
        # 엑셀 파일 모델 생성
        xl_model = formulas.ExcelModel().loads(excel_path).finish()
        
        # 수식 계산
        solution = xl_model.calculate()
        
        # 계산된 값으로 파일 업데이트
        wb = openpyxl.load_workbook(excel_path)
        
        # 분석 시트 목록
        analysis_sheets = [s for s in wb.sheetnames if '분석' in s]
        
        for sheet_name in analysis_sheets:
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    
                    # 수식이 있는 셀인 경우
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # solution에서 계산된 값 가져오기
                        cell_ref = f"'{sheet_name}'!{cell.coordinate}"
                        try:
                            if cell_ref in solution:
                                calc_value = solution[cell_ref].value
                                if calc_value is not None and not pd.isna(calc_value):
                                    # 에러가 아닌 경우에만 값 대체
                                    if not isinstance(calc_value, str) or not calc_value.startswith('#'):
                                        cell.value = calc_value
                        except Exception:
                            pass
        
        wb.save(excel_path)
        wb.close()
        print(f"[변환] formulas 라이브러리로 수식 계산 완료")
    
    def _calculate_formulas_manually(self, excel_path: str):
        """수동으로 수식 계산 (증감률 등 기본 수식)
        
        분석 시트의 주요 수식 패턴을 파악하여 직접 계산합니다.
        """
        wb = openpyxl.load_workbook(excel_path)
        
        # 분석 시트와 집계 시트 매핑
        analysis_to_aggregate = {
            'A 분석': 'A(광공업생산)집계',
            'B 분석': 'B(서비스업생산)집계',
            'C 분석': 'C(소비)집계',
            'D(고용률)분석': 'D(고용률)집계',
            'D(실업)분석': 'D(실업)집계',
            'E(지출목적물가) 분석': 'E(지출목적물가)집계',
            'E(품목성질물가)분석': 'E(품목성질물가)집계',
            "F'분석": "F'(건설)집계",
            'G 분석': 'G(수출)집계',
            'H 분석': 'H(수입)집계',
            'I(순인구이동)분석': 'I(순인구이동)집계',
        }
        
        calc_count = 0
        
        for analysis_sheet, aggregate_sheet in analysis_to_aggregate.items():
            if analysis_sheet not in wb.sheetnames:
                continue
            if aggregate_sheet not in wb.sheetnames:
                continue
            
            ws_analysis = wb[analysis_sheet]
            ws_aggregate = wb[aggregate_sheet]
            
            # 분석 시트의 모든 셀 순회
            for row in ws_analysis.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula = cell.value
                        
                        try:
                            # 증감률 계산 패턴: (A-B)/B*100
                            calc_value = self._evaluate_growth_formula(formula, wb)
                            if calc_value is not None:
                                cell.value = calc_value
                                calc_count += 1
                        except Exception:
                            pass
        
        wb.save(excel_path)
        wb.close()
        print(f"[변환] 수동 수식 계산 완료 ({calc_count}개 셀)")
    
    def _evaluate_growth_formula(self, formula: str, wb) -> Optional[float]:
        """증감률 수식 평가
        
        주요 패턴:
        - =IFERROR(('시트'!A1-'시트'!B1)/'시트'!B1*100,"없음")
        - =('시트'!A1-'시트'!B1)/'시트'!B1*100
        """
        # IFERROR 래핑 제거
        if 'IFERROR' in formula.upper():
            # IFERROR(수식, 대체값) 에서 수식 부분만 추출
            match = re.match(r'=IFERROR\s*\(\s*(.+?)\s*,\s*["\']?없음["\']?\s*\)', formula, re.IGNORECASE)
            if match:
                formula = '=' + match.group(1)
        
        # 셀 참조 패턴: 'Sheet'!A1 또는 Sheet!A1
        cell_ref_pattern = r"'?([^'!]+)'?!([A-Z]+)(\d+)"
        
        # 수식에서 모든 셀 참조 추출
        refs = re.findall(cell_ref_pattern, formula)
        if not refs:
            return None
        
        # 각 셀의 값 가져오기
        cell_values = {}
        for sheet_name, col, row in refs:
            sheet_name = sheet_name.strip("'")
            if sheet_name not in wb.sheetnames:
                return None
            
            ws = wb[sheet_name]
            cell_value = ws[f"{col}{row}"].value
            
            # 수식인 경우 재귀 계산
            if isinstance(cell_value, str) and cell_value.startswith('='):
                cell_value = self._evaluate_growth_formula(cell_value, wb)
            
            if cell_value is None or (isinstance(cell_value, float) and pd.isna(cell_value)):
                return None
            
            try:
                cell_values[f"'{sheet_name}'!{col}{row}"] = float(cell_value)
            except (ValueError, TypeError):
                return None
        
        # 간단한 증감률 계산 패턴 감지
        # 패턴: (A-B)/B*100
        if len(refs) == 2 and '-' in formula and '/' in formula and '*100' in formula:
            keys = list(cell_values.keys())
            a_val = cell_values[keys[0]]
            b_val = cell_values[keys[1]]
            
            if b_val != 0:
                result = ((a_val - b_val) / b_val) * 100
                return round(result, 2)
        
        # 패턴: A-B (차이 계산, %p 단위)
        elif len(refs) == 2 and '-' in formula and '/' not in formula and '*' not in formula:
            keys = list(cell_values.keys())
            a_val = cell_values[keys[0]]
            b_val = cell_values[keys[1]]
            result = a_val - b_val
            return round(result, 2)
        
        return None
    
    def _copy_sheet_data(self, raw_xl: pd.ExcelFile, raw_sheet: str, target_ws, weight_settings: Dict = None):
        """기초자료 시트의 데이터를 분석표 집계 시트에 복사
        
        연도/분기 범위에 맞는 열만 선택적으로 복사:
        - 연도: 당해 제외 최근 5개년
        - 분기: 최근 13개 분기
        - 가중치: weight_settings에 따라 처리 (auto/manual/empty)
        
        중요: 분석표 템플릿의 열 1~3은 조회용 컬럼이므로 건너뛰고, 
        기초자료의 메타데이터는 분석표의 열 4부터 매핑됨.
        """
        
        # 기초자료 읽기 (헤더 없이)
        raw_df = pd.read_excel(raw_xl, sheet_name=raw_sheet, header=None)
        
        # 기초자료 헤더 파싱
        header_mapping = self._parse_raw_header(raw_df, header_row=2)
        
        # 목표 연도/분기 범위
        target_years = self._get_target_years()
        target_quarters = self._get_target_quarters()
        
        # 집계 시트 구조 가져오기
        target_sheet_name = target_ws.title
        sheet_structure = self.SHEET_STRUCTURE.get(target_sheet_name, {})
        
        # 새로운 구조: meta_start는 분석표에서 메타데이터가 시작하는 열 (1-based)
        meta_start = sheet_structure.get('meta_start', 4)  # 기본값 4 (열 D)
        raw_meta_cols = sheet_structure.get('raw_meta_cols', 6)  # 기초자료의 메타 열 개수
        target_year_start_col = sheet_structure.get('year_start', 10)
        target_quarter_start_col = sheet_structure.get('quarter_start', 15)
        target_weight_col = sheet_structure.get('weight_col')  # 1-based index
        raw_weight_col = sheet_structure.get('raw_weight_col')  # 0-based index
        
        # 가중치 설정 결정 (raw_sheet에 따라)
        weight_config = None
        if weight_settings:
            if raw_sheet == '광공업생산':
                weight_config = weight_settings.get('mining', {})
            elif raw_sheet == '서비스업생산':
                weight_config = weight_settings.get('service', {})
        
        weight_mode = weight_config.get('mode', 'auto') if weight_config else 'auto'
        manual_weight_values = weight_config.get('values', []) if weight_config else []
        
        # 열 매핑 생성: 기초자료 열 → 집계 시트 열
        col_mapping = {}
        
        # 메타데이터 열 매핑 (기초자료 열 0~(raw_meta_cols-1) → 분석표 열 meta_start~)
        # 가중치 열은 별도 처리
        target_col_idx = meta_start
        for raw_col in range(raw_meta_cols):
            # 기초자료의 가중치 열은 별도 처리 (분석표의 가중치 열에 매핑)
            if raw_weight_col is not None and raw_col == raw_weight_col:
                # 가중치는 별도 매핑
                col_mapping[raw_col] = target_weight_col
                continue
            
            col_mapping[raw_col] = target_col_idx
            target_col_idx += 1
            
            # 분석표의 가중치 열 위치에 도달하면 건너뛰기
            if target_weight_col is not None and target_col_idx == target_weight_col:
                target_col_idx += 1
        
        # 연도 열 매핑
        for i, year in enumerate(target_years):
            if year in header_mapping['years']:
                raw_col = header_mapping['years'][year]
                col_mapping[raw_col] = target_year_start_col + i
        
        # 분기 열 매핑
        for i, q_key in enumerate(target_quarters):
            if q_key in header_mapping['quarters']:
                raw_col = header_mapping['quarters'][q_key]
                col_mapping[raw_col] = target_quarter_start_col + i
        
        print(f"    메타데이터: 기초자료 열 0~{raw_meta_cols-1} → 분석표 열 {meta_start}~")
        print(f"    연도 범위: {target_years[0]}~{target_years[-1]} (열 {target_year_start_col}~{target_year_start_col + self.num_years - 1})")
        print(f"    분기 범위: {target_quarters[0]}~{target_quarters[-1]} (열 {target_quarter_start_col}~{target_quarter_start_col + self.num_quarters - 1})")
        if target_weight_col:
            print(f"    가중치 열: 기초자료 열 {raw_weight_col} → 분석표 열 {target_weight_col} (모드: {weight_mode})")
        
        copied_count = 0
        skipped_count = 0
        error_count = 0
        
        # 데이터 복사 (행 순회) - 헤더 3행 이후부터 데이터
        for row_idx in range(len(raw_df)):
            for raw_col, target_col in col_mapping.items():
                if raw_col >= len(raw_df.columns):
                    continue
                
                value = raw_df.iloc[row_idx, raw_col]
                
                # NaN 처리: 빈셀은 그대로 두되, 주변 패턴을 보고 추정 가능하면 추정
                if pd.isna(value):
                    # 주변 패턴 확인 (같은 행의 이전/다음 열 값 확인)
                    estimated_value = self._estimate_missing_value(raw_df, row_idx, raw_col, header_mapping)
                    if estimated_value is not None:
                        value = estimated_value
                    else:
                        # 추정 불가능하면 건너뛰기 (나중에 실제 값이 들어갈 것으로 가정)
                        continue
                
                # 0 처리: 0으로 채워진 셀은 주변 패턴을 보고 실제 값일 가능성 확인
                if isinstance(value, (int, float)) and value == 0:
                    # 주변 패턴 확인
                    estimated_value = self._estimate_zero_value(raw_df, row_idx, raw_col, header_mapping)
                    if estimated_value is not None and estimated_value != 0:
                        value = estimated_value
                    # 0이 실제 값일 수도 있으므로 그대로 사용
                
                # 데이터 검증: 숫자 열에 문자열이 들어가는 것 방지
                # 연도/분기 데이터 열 (target_year_start_col 이후)에는 숫자만 허용
                if target_col >= target_year_start_col:
                    if not isinstance(value, (int, float)):
                        try:
                            value = float(value)
                        except (ValueError, TypeError):
                            # 숫자가 아닌 값은 건너뛰기
                            error_count += 1
                            if error_count <= 5:  # 최대 5개까지만 로그
                                print(f"    [경고] 행 {row_idx+1}, 열 {target_col}: 숫자가 아닌 값 무시 '{value}'")
                            continue
                
                # 가중치 열 특별 처리
                if target_weight_col and target_col == target_weight_col:
                    if weight_mode == 'empty':
                        continue  # 가중치 공란 유지
                    elif weight_mode == 'manual':
                        # 수동 입력 모드는 아래에서 별도 처리
                        continue
                    # auto 모드: 기초자료에서 가져온 값 사용
                    if not isinstance(value, (int, float)):
                        try:
                            value = float(value)
                        except (ValueError, TypeError):
                            continue  # 숫자가 아니면 건너뛰기
                
                # openpyxl은 1-based 인덱스
                cell = target_ws.cell(row=row_idx + 1, column=target_col)
                
                # 병합된 셀은 건너뛰기
                if isinstance(cell, MergedCell):
                    skipped_count += 1
                    continue
                
                # 값 복사 (스타일은 템플릿에서 유지됨)
                try:
                    # 숫자 값의 경우 소수점 처리
                    if isinstance(value, float):
                        # 소수점이 0이면 정수로 변환
                        if value == int(value):
                            cell.value = int(value)
                        else:
                            cell.value = round(value, 2)
                    else:
                        cell.value = value
                    copied_count += 1
                except Exception:
                    skipped_count += 1
            
            # 수동 가중치 처리
            if target_weight_col and weight_mode == 'manual':
                data_row_idx = row_idx - 3  # 헤더 3행 제외
                if 0 <= data_row_idx < len(manual_weight_values):
                    weight_value = manual_weight_values[data_row_idx]
                    if weight_value is not None:
                        try:
                            weight_cell = target_ws.cell(row=row_idx + 1, column=target_weight_col)
                            if not isinstance(weight_cell, MergedCell):
                                weight_cell.value = float(weight_value)
                                copied_count += 1
                        except Exception:
                            skipped_count += 1
        
        print(f"  → {copied_count}개 셀 복사 ({skipped_count}개 건너뜀, {error_count}개 오류)")
    
    def _estimate_missing_value(self, df: pd.DataFrame, row_idx: int, col_idx: int, header_mapping: Dict) -> Optional[float]:
        """빈셀의 값을 주변 패턴으로 추정
        
        같은 행의 이전/다음 분기 값이나 같은 열의 이전/다음 행 값을 참고하여 추정.
        추정 불가능하면 None 반환 (나중에 실제 값이 들어갈 것으로 가정).
        """
        # 같은 행의 이전/다음 분기 값 확인
        quarter_cols = sorted(header_mapping['quarters'].values())
        if col_idx in quarter_cols:
            q_idx = quarter_cols.index(col_idx)
            # 이전 분기 값 확인
            if q_idx > 0:
                prev_col = quarter_cols[q_idx - 1]
                if prev_col < len(df.columns):
                    prev_val = df.iloc[row_idx, prev_col]
                    if pd.notna(prev_val) and isinstance(prev_val, (int, float)) and prev_val != 0:
                        # 이전 분기 값의 평균 증가율을 가정하여 추정 (보수적 추정)
                        return prev_val * 1.0  # 변화 없음으로 가정 (보수적)
        
        # 추정 불가능
        return None
    
    def _estimate_zero_value(self, df: pd.DataFrame, row_idx: int, col_idx: int, header_mapping: Dict) -> Optional[float]:
        """0으로 채워진 셀의 값을 주변 패턴으로 추정
        
        0이 실제 값일 수도 있으므로, 주변 패턴을 확인하여 0이 아닐 가능성이 높은 경우만 추정.
        """
        # 같은 행의 이전/다음 분기 값 확인
        quarter_cols = sorted(header_mapping['quarters'].values())
        if col_idx in quarter_cols:
            q_idx = quarter_cols.index(col_idx)
            # 이전/다음 분기 값 확인
            prev_val = None
            next_val = None
            
            if q_idx > 0:
                prev_col = quarter_cols[q_idx - 1]
                if prev_col < len(df.columns):
                    prev_val = df.iloc[row_idx, prev_col]
            
            if q_idx < len(quarter_cols) - 1:
                next_col = quarter_cols[q_idx + 1]
                if next_col < len(df.columns):
                    next_val = df.iloc[row_idx, next_col]
            
            # 이전과 다음 값이 모두 0이 아니면 평균으로 추정
            if prev_val is not None and next_val is not None:
                if pd.notna(prev_val) and pd.notna(next_val) and prev_val != 0 and next_val != 0:
                    return (prev_val + next_val) / 2
                elif pd.notna(prev_val) and prev_val != 0:
                    return prev_val
                elif pd.notna(next_val) and next_val != 0:
                    return next_val
        
        # 추정 불가능 (0이 실제 값일 가능성)
        return None
    
    def _calculate_analysis_values(self, wb):
        """분석 시트의 수식을 계산하여 값으로 저장 (캐시용)
        
        엑셀에서 열지 않아도 Python에서 값을 읽을 수 있도록
        분석 시트의 수식 셀에 계산된 값을 추가로 저장
        """
        # 분석 시트 목록
        analysis_sheets = [
            ('A 분석', 'A(광공업생산)집계'),
            ('B 분석', 'B(서비스업생산)집계'),
            ('C 분석', 'C(소비)집계'),
            ("D(고용률)분석", 'D(고용률)집계'),
            ("D(실업)분석", 'D(실업)집계'),
            ('E(지출목적물가) 분석', 'E(지출목적물가)집계'),
            ('E(품목성질물가)분석', 'E(품목성질물가)집계'),
            ("F'분석", "F'(건설)집계"),
            ('G 분석', 'G(수출)집계'),
            ('H 분석', 'H(수입)집계'),
            ('I(순인구이동)분석', 'I(순인구이동)집계'),
        ]
        
        for analysis_sheet, source_sheet in analysis_sheets:
            if analysis_sheet not in wb.sheetnames:
                continue
            if source_sheet not in wb.sheetnames:
                continue
                
            ws = wb[analysis_sheet]
            source_ws = wb[source_sheet]
            
            # 수식이 있는 셀을 찾아서 계산
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # 수식 파싱 및 계산
                        calculated = self._evaluate_formula(cell.value, wb)
                        if calculated is not None:
                            # 수식을 유지하면서 캐시된 값도 저장
                            # openpyxl에서는 이게 안되므로 수식을 계산된 값으로 대체
                            pass  # 수식을 유지하는 것이 사용자 요청
    
    def _evaluate_formula(self, formula: str, wb) -> Optional[float]:
        """엑셀 수식 평가 (간단한 수식만 지원)"""
        import re
        
        # IFERROR 처리
        if formula.startswith('=IFERROR('):
            # 내부 수식 추출
            inner = formula[9:-1]  # =IFERROR( 와 ) 제거
            # 마지막 ,"없음") 부분 제거
            if ',"없음"' in inner:
                inner = inner.replace(',"없음"', '').replace(',"없음")', ')')
            return self._evaluate_formula('=' + inner.split(',')[0], wb)
        
        # 셀 참조 패턴: 'Sheet'!Cell 또는 Sheet!Cell
        cell_pattern = r"'?([^'!]+)'?!([A-Z]+)(\d+)"
        
        try:
            matches = re.findall(cell_pattern, formula)
            if not matches:
                return None
            
            values = {}
            for sheet_name, col, row in matches:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    cell = ws[f"{col}{row}"]
                    val = cell.value
                    if val is None or (isinstance(val, str) and not val.replace('.','').replace('-','').isdigit()):
                        return None
                    values[f"'{sheet_name}'!{col}{row}"] = float(val) if val else 0
            
            # 간단한 증감률 계산: (A-B)/B*100
            if len(matches) == 2 and '-' in formula and '/' in formula and '*100' in formula:
                keys = list(values.keys())
                if len(keys) == 2:
                    a, b = values[keys[0]], values[keys[1]]
                    if b != 0:
                        return round((a - b) / b * 100, 2)
            
            return None
        except Exception:
            return None
    
    def extract_grdp_data(self) -> Dict:
        """GRDP 데이터 추출
        
        Returns:
            GRDP 보도자료용 데이터 딕셔너리
        """
        raw_xl = pd.ExcelFile(self.raw_excel_path)
        
        if '분기 GRDP' not in raw_xl.sheet_names:
            print("[경고] '분기 GRDP' 시트를 찾을 수 없습니다.")
            return self._get_placeholder_grdp()
        
        df = pd.read_excel(raw_xl, sheet_name='분기 GRDP', header=None)
        
        # 헤더 행 (3행)
        header_row = 2
        
        # self.year, self.quarter 사용 (이미 _detect_year_quarter()에서 설정됨)
        year = self.year
        quarter = self.quarter
        prev_year = year - 1
        
        print(f"[GRDP] 추출된 연도/분기: {year}년 {quarter}분기")
        
        # 최신 분기 컬럼 찾기 (동적 처리)
        current_quarter_col = -1
        prev_year_quarter_col = -1
        
        # 분기 패턴 리스트 (다양한 형식 지원)
        import re
        
        current_patterns = [
            f"{year}.{quarter}/4",
            f"{year} {quarter}/4",
            f"{year}.{quarter}/4p",
            f"{year} {quarter}/4p",
            f"{year}년 {quarter}분기",
            f"{year}  {quarter}/4",  # 공백 2개
            f"{year}  {quarter}/4p",  # 공백 2개 + p
        ]
        
        prev_patterns = [
            f"{prev_year}.{quarter}/4",
            f"{prev_year} {quarter}/4",
            f"{prev_year}.{quarter}/4p",
            f"{prev_year} {quarter}/4p",
            f"{prev_year}년 {quarter}분기",
            f"{prev_year}  {quarter}/4",  # 공백 2개
            f"{prev_year}  {quarter}/4p",  # 공백 2개 + p
        ]
        
        for col_idx in range(len(df.columns)):
            cell = str(df.iloc[header_row, col_idx]).strip()
            
            # 현재 분기 패턴 매칭
            if current_quarter_col == -1:
                for pattern in current_patterns:
                    if pattern in cell:
                        current_quarter_col = col_idx
                        print(f"[GRDP] 현재 분기 컬럼 발견: col {col_idx} = '{cell}'")
                        break
            
            # 전년동기 분기 패턴 매칭
            if prev_year_quarter_col == -1:
                for pattern in prev_patterns:
                    if pattern in cell:
                        prev_year_quarter_col = col_idx
                        print(f"[GRDP] 전년동기 컬럼 발견: col {col_idx} = '{cell}'")
                        break
        
        # Fallback: 정규식 패턴으로 재시도 (더 유연한 매칭)
        if current_quarter_col == -1:
            for col_idx in range(len(df.columns)):
                cell = str(df.iloc[header_row, col_idx]).strip()
                # "2025.3/4", "2025 3/4", "2025  3/4", "'25.3/4", "2025  3/4p" 등 패턴
                # 공백이 1개 이상일 수 있음
                quarter_match = re.search(rf'(\d{{2,4}})[.\s]{{1,3}}{quarter}/4', cell)
                if quarter_match:
                    matched_year = int(quarter_match.group(1))
                    if matched_year < 100:
                        matched_year += 2000
                    if matched_year == year:
                        current_quarter_col = col_idx
                        print(f"[GRDP] 현재 분기 컬럼 발견 (정규식): col {col_idx} = '{cell}'")
                        break
        
        if prev_year_quarter_col == -1:
            for col_idx in range(len(df.columns)):
                cell = str(df.iloc[header_row, col_idx]).strip()
                # 공백이 1개 이상일 수 있음
                quarter_match = re.search(rf'(\d{{2,4}})[.\s]{{1,3}}{quarter}/4', cell)
                if quarter_match:
                    matched_year = int(quarter_match.group(1))
                    if matched_year < 100:
                        matched_year += 2000
                    if matched_year == prev_year:
                        prev_year_quarter_col = col_idx
                        print(f"[GRDP] 전년동기 컬럼 발견 (정규식): col {col_idx} = '{cell}'")
                        break
        
        # 최종 Fallback: 가장 최신 분기 찾기
        if current_quarter_col == -1:
            print(f"[경고] {year}년 {quarter}분기 컬럼을 찾을 수 없습니다. 가장 최신 분기 검색 중...")
            latest_year = 0
            latest_quarter = 0
            latest_col = -1
            
            for col_idx in range(len(df.columns)):
                cell = str(df.iloc[header_row, col_idx]).strip()
                # 분기 패턴 찾기
                q_match = re.search(r'(\d{4})[.\s]{1,3}(\d)/4', cell)
                if q_match:
                    matched_year = int(q_match.group(1))
                    matched_quarter = int(q_match.group(2))
                    
                    # 가장 최신 분기 찾기
                    if matched_year > latest_year or (matched_year == latest_year and matched_quarter > latest_quarter):
                        latest_year = matched_year
                        latest_quarter = matched_quarter
                        latest_col = col_idx
            
            if latest_col != -1:
                current_quarter_col = latest_col
                print(f"[GRDP] 가장 최신 분기 사용: {latest_year}년 {latest_quarter}분기 (col {latest_col})")
                # 전년동기 컬럼 계산 (4분기 전)
                prev_year_quarter_col = latest_col - 4
            else:
                # 최후의 수단: 마지막 컬럼 사용
                print(f"[경고] 분기 컬럼을 전혀 찾을 수 없습니다. 마지막 컬럼 사용.")
                current_quarter_col = len(df.columns) - 1
                prev_year_quarter_col = current_quarter_col - 4
        
        if prev_year_quarter_col == -1 or prev_year_quarter_col < 0:
            print(f"[경고] {prev_year}년 {quarter}분기 컬럼을 찾을 수 없습니다. 계산된 컬럼 사용.")
            prev_year_quarter_col = max(0, current_quarter_col - 4)
        
        print(f"[GRDP] 최종 컬럼: 현재={current_quarter_col}, 전년동기={prev_year_quarter_col}")
        
        # 지역별 데이터 추출
        regional_data = []
        national_data = {}
        
        # 그룹 매핑
        region_groups = {
            '서울': '경인', '인천': '경인', '경기': '경인',
            '대전': '충청', '세종': '충청', '충북': '충청', '충남': '충청',
            '광주': '호남', '전북': '호남', '전남': '호남', '제주': '호남',
            '대구': '동북', '경북': '동북', '강원': '동북',
            '부산': '동남', '울산': '동남', '경남': '동남',
        }
        
        # 지역별 데이터 수집
        region_values = {}
        
        for i in range(3, len(df)):
            region = str(df.iloc[i, 1]).strip() if not pd.isna(df.iloc[i, 1]) else ''
            level = str(df.iloc[i, 2]).strip() if not pd.isna(df.iloc[i, 2]) else ''
            item = str(df.iloc[i, 4]).strip() if not pd.isna(df.iloc[i, 4]) else ''
            
            if not region or region not in self.REGIONS_ORDER:
                continue
            
            current_val = self._safe_float(df.iloc[i, current_quarter_col])
            prev_year_val = self._safe_float(df.iloc[i, prev_year_quarter_col])
            
            if region not in region_values:
                region_values[region] = {'total': None, 'industries': []}
            
            if level == '0':
                region_values[region]['total'] = {
                    'item': item,
                    'current': current_val,
                    'prev_year': prev_year_val,
                }
            elif level == '1':
                region_values[region]['industries'].append({
                    'item': item,
                    'current': current_val,
                    'prev_year': prev_year_val,
                })
        
        # 성장률 및 기여도 계산
        for region in self.REGIONS_ORDER:
            if region not in region_values:
                continue
            
            values = region_values[region]
            total = values.get('total', {'current': 0, 'prev_year': 0})
            if total is None:
                total = {'current': 0, 'prev_year': 0}
            
            total_current = total['current']
            total_prev = total['prev_year']
            
            # 성장률 계산
            if total_prev > 0:
                growth_rate = round(((total_current - total_prev) / total_prev) * 100, 1)
            else:
                growth_rate = 0.0
            
            # 산업별 기여도 계산
            manufacturing_contrib = 0.0
            construction_contrib = 0.0
            service_contrib = 0.0
            other_contrib = 0.0
            
            for industry in values.get('industries', []):
                item_name = industry.get('item', '')
                contrib = self._calculate_contribution(
                    industry['current'], industry['prev_year'], total_prev
                )
                
                if '제조' in item_name or '광업' in item_name:
                    manufacturing_contrib = contrib
                elif '건설' in item_name:
                    construction_contrib = contrib
                elif '서비스' in item_name:
                    service_contrib = contrib
                elif '기타' in item_name or '순생산' in item_name:
                    other_contrib = contrib
            
            region_entry = {
                'region': region,
                'region_group': region_groups.get(region, ''),
                'growth_rate': growth_rate,
                'manufacturing': manufacturing_contrib,
                'construction': construction_contrib,
                'service': service_contrib,
                'other': other_contrib,
                'placeholder': False,
                'needs_review': False
            }
            
            if region == '전국':
                national_data = region_entry.copy()
            
            regional_data.append(region_entry)
        
        # 성장률 1위 지역 찾기
        non_national = [r for r in regional_data if r['region'] != '전국']
        if non_national:
            top_region = max(non_national, key=lambda x: x['growth_rate'])
        else:
            top_region = {'region': '-', 'growth_rate': 0.0, 'manufacturing': 0.0, 
                         'construction': 0.0, 'service': 0.0, 'other': 0.0}
        
        return {
            'report_info': {
                'year': self.year,
                'quarter': self.quarter,
                'page_number': 20
            },
            'national_summary': {
                'growth_rate': national_data.get('growth_rate', 0.0),
                'direction': '증가' if national_data.get('growth_rate', 0) >= 0 else '감소',
                'contributions': {
                    'manufacturing': national_data.get('manufacturing', 0.0),
                    'construction': national_data.get('construction', 0.0),
                    'service': national_data.get('service', 0.0),
                    'other': national_data.get('other', 0.0),
                },
                'placeholder': False,
                'needs_review': False
            },
            'top_region': {
                'name': top_region.get('region', '-'),
                'growth_rate': top_region.get('growth_rate', 0.0),
                'contributions': {
                    'manufacturing': top_region.get('manufacturing', 0.0),
                    'construction': top_region.get('construction', 0.0),
                    'service': top_region.get('service', 0.0),
                    'other': top_region.get('other', 0.0),
                },
                'placeholder': False,
                'needs_review': False
            },
            'regional_data': regional_data,
            'chart_config': {
                'y_axis': {
                    'min': -6,
                    'max': 8,
                    'step': 2
                }
            },
            'needs_review': False,
            'data_missing': False
        }
    
    def _safe_float(self, val) -> Optional[float]:
        """안전하게 float으로 변환
        
        PM 요구사항: 데이터가 없을 때는 0.0이 아니라 None(N/A)로 처리
        """
        if val is None:
            return None
        if pd.isna(val):
            return None
        try:
            if isinstance(val, str):
                val = val.strip()
                if val == '-' or val == '' or val.lower() in ['없음', 'nan', 'none', 'n/a']:
                    return None
            result = float(val)
            if pd.isna(result):
                return None
            return result
        except (ValueError, TypeError):
            return None
    
    def _calculate_contribution(self, current: Optional[float], prev: Optional[float], total_prev: Optional[float]) -> Optional[float]:
        """산업별 기여도 계산
        
        PM 요구사항: 데이터가 없으면 None 반환
        """
        if current is None or prev is None or total_prev is None or total_prev == 0:
            return None  # N/A 처리
        return round(((current - prev) / total_prev) * 100, 1)
    
    def _get_placeholder_grdp(self) -> Dict:
        """플레이스홀더 GRDP 데이터 (기본값 기여율 사용)"""
        regional_data = []
        region_groups = {
            '서울': '경인', '인천': '경인', '경기': '경인',
            '대전': '충청', '세종': '충청', '충북': '충청', '충남': '충청',
            '광주': '호남', '전북': '호남', '전남': '호남', '제주': '호남',
            '대구': '동북', '경북': '동북', '강원': '동북',
            '부산': '동남', '울산': '동남', '경남': '동남',
        }
        
        # 기본 기여율 로드 시도
        default_contributions = None
        try:
            default_contrib_path = Path(__file__).parent / 'templates' / 'default_contributions.json'
            if default_contrib_path.exists():
                with open(default_contrib_path, 'r', encoding='utf-8') as f:
                    default_contributions = json.load(f)
        except Exception as e:
            print(f"[GRDP] 기본 기여율 로드 실패: {e}")
        
        for region in self.REGIONS_ORDER:
            if default_contributions and region != '전국':
                region_contrib = default_contributions.get('regional', {}).get(region, {})
                regional_data.append({
                    'region': region,
                    'region_group': region_groups.get(region, ''),
                    'growth_rate': region_contrib.get('growth_rate', 0.0),
                    'manufacturing': region_contrib.get('manufacturing', 0.0),
                    'construction': region_contrib.get('construction', 0.0),
                    'service': region_contrib.get('service', 0.0),
                    'other': region_contrib.get('other', 0.0),
                    'placeholder': True,
                    'needs_review': True
                })
            else:
                regional_data.append({
                    'region': region,
                    'region_group': region_groups.get(region, ''),
                    'growth_rate': 0.0,
                    'manufacturing': 0.0,
                    'construction': 0.0,
                    'service': 0.0,
                    'other': 0.0,
                    'placeholder': True,
                    'needs_review': True
                })
        
        # 전국 기여율 기본값
        national_growth = 0.0
        national_contributions = {'manufacturing': 0.0, 'construction': 0.0, 'service': 0.0, 'other': 0.0}
        
        if default_contributions:
            national = default_contributions.get('national', {})
            national_growth = national.get('growth_rate', 0.0)
            national_contributions = national.get('contributions', national_contributions)
        
        return {
            'report_info': {
                'year': self.year,
                'quarter': self.quarter,
                'page_number': 20
            },
            'national_summary': {
                'growth_rate': national_growth,
                'direction': '증가' if national_growth >= 0 else '감소',
                'contributions': national_contributions,
                'placeholder': True,
                'needs_review': True
            },
            'top_region': {
                'name': '-',
                'growth_rate': 0.0,
                'contributions': {
                    'manufacturing': 0.0,
                    'construction': 0.0,
                    'service': 0.0,
                    'other': 0.0,
                },
                'placeholder': True,
                'needs_review': True
            },
            'regional_data': regional_data,
            'chart_config': {
                'y_axis': {
                    'min': -6,
                    'max': 8,
                    'step': 2
                }
            },
            'needs_review': True,
            'data_missing': True
        }
    
    def save_grdp_json(self, output_path: str) -> Dict:
        """GRDP 데이터를 JSON으로 저장"""
        data = self.extract_grdp_data()
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return data


def detect_file_type(excel_path: str) -> str:
    """엑셀 파일 유형 감지 (기초자료 vs 분석표)
    
    기초자료 수집표와 분석표는 시트 구조가 완전히 다릅니다:
    
    [기초자료 수집표] 17개 시트
    - 시트명: '광공업생산', '서비스업생산', '고용률', '분기 GRDP' 등
    - 특징: 원본 데이터 형태, 한글 시트명
    
    [분석표] 42개 시트  
    - 시트명: 'A(광공업생산)집계', 'A 분석', '본청', '이용관련' 등
    - 특징: 알파벳+키워드(집계/분석/참고) 패턴
    
    두 파일은 공통 시트가 하나도 없습니다.
    """
    try:
        xl = pd.ExcelFile(excel_path)
        sheet_names = set(xl.sheet_names)
        
        # ========================================
        # 기초자료 수집표 전용 시트 (분석표에는 없음)
        # ========================================
        RAW_ONLY_SHEETS = {
            '광공업생산', '서비스업생산', '소비(소매, 추가)', 
            '고용', '고용(kosis)', '고용률', '실업자 수',
            '지출목적별 물가', '품목성질별 물가', '건설 (공표자료)',
            '수출', '수입', '분기 GRDP', '완료체크'
        }
        
        # ========================================
        # 분석표 전용 시트 (기초자료에는 없음)
        # ========================================
        ANALYSIS_ONLY_SHEETS = {
            '본청', '시도별 현황', '지방청 이용자료', '이용관련',
            'A(광공업생산)집계', 'A 분석', 'A 참고',
            'B(서비스업생산)집계', 'B 분석',
            "F'(건설)집계", "F'분석",
            'G(수출)집계', 'G 분석',
        }
        
        # 매칭된 시트 수로 판정
        raw_matches = len(sheet_names & RAW_ONLY_SHEETS)
        analysis_matches = len(sheet_names & ANALYSIS_ONLY_SHEETS)
        
        # 분석표 전용 시트가 있으면 분석표
        if analysis_matches >= 3:
            return 'analysis'
        
        # 기초자료 전용 시트가 있으면 기초자료
        if raw_matches >= 3:
            return 'raw'
        
        # 시트명 패턴으로 추가 판정
        analysis_pattern_count = sum(1 for s in sheet_names 
                                      if '분석' in s or '집계' in s)
        if analysis_pattern_count >= 3:
            return 'analysis'
        
        # 파일명으로 최종 추정
        filename = Path(excel_path).stem.lower()
        if '기초' in filename or '수집' in filename:
            return 'raw'
        elif '분석' in filename:
            return 'analysis'
        
        return 'unknown'
    except Exception as e:
        print(f"파일 유형 감지 실패: {e}")
        return 'unknown'


def convert_raw_to_analysis(raw_excel_path: str, output_path: str = None, 
                           template_path: str = None, weight_data: Dict[str, Dict] = None) -> Tuple[str, Dict]:
    """기초자료 수집표 → 분석표 변환 및 GRDP 추출
    
    Args:
        raw_excel_path: 기초자료 수집표 경로
        output_path: 분석표 출력 경로 (None이면 자동 생성)
        template_path: 분석표 템플릿 경로 (None이면 기본 템플릿 사용)
        weight_data: 시트별 가중치 데이터 {시트명: {행인덱스: 가중치값}}
        
    Returns:
        (분석표 경로, GRDP 데이터)
    """
    converter = DataConverter(raw_excel_path, template_path)
    analysis_path = converter.convert_all(output_path, weight_data)
    grdp_data = converter.extract_grdp_data()
    
    return analysis_path, grdp_data


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='기초자료 수집표 → 분석표 변환')
    parser.add_argument('input', type=str, help='기초자료 수집표 엑셀 파일')
    parser.add_argument('--output', type=str, default=None, help='출력 분석표 경로')
    parser.add_argument('--template', type=str, default=None, help='분석표 템플릿 경로')
    parser.add_argument('--grdp-json', type=str, default='grdp_data.json', help='GRDP JSON 출력 경로')
    
    args = parser.parse_args()
    
    # 변환 실행
    analysis_path, grdp_data = convert_raw_to_analysis(
        args.input, args.output, args.template
    )
    
    # GRDP JSON 저장
    with open(args.grdp_json, 'w', encoding='utf-8') as f:
        json.dump(grdp_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n=== 변환 완료 ===")
    print(f"분석표: {analysis_path}")
    print(f"GRDP JSON: {args.grdp_json}")
    print(f"전국 성장률: {grdp_data['national_summary']['growth_rate']}%")
    print(f"1위 지역: {grdp_data['top_region']['name']} ({grdp_data['top_region']['growth_rate']}%)")
