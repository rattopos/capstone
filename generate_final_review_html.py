"""담당자 최종 검토용 HTML 생성 - 기본 템플릿 방식"""

import sys
import os
import logging
from pathlib import Path
sys.path.insert(0, '/Users/topos/Library/CloudStorage/GoogleDrive-ckdwo0605@gmail.com/내 드라이브/capstone')

# 로깅 비활성화
logging.disable(logging.CRITICAL)
os.environ['PYTHONWARNINGS'] = 'ignore'

from templates.unified_generator import UnifiedReportGenerator
from config.reports import SECTOR_REPORTS
from utils.text_utils import get_terms
import openpyxl

excel_path = "/Users/topos/Library/CloudStorage/GoogleDrive-ckdwo0605@gmail.com/내 드라이브/capstone/분석표_25년 3분기_캡스톤(업데이트).xlsx"
year, quarter = 2025, 3
output_dir = Path("/Users/topos/Library/CloudStorage/GoogleDrive-ckdwo0605@gmail.com/내 드라이브/capstone/exports/final_review")
output_dir.mkdir(parents=True, exist_ok=True)

def get_industry_name_mapping(sector_id: str) -> dict:
    """엑셀에서 산업 코드와 산업명의 매핑 딕셔너리 생성"""
    mapping = {}
    config = next((s for s in SECTOR_REPORTS if s['id'] == sector_id), None)
    
    if not config or 'aggregation_structure' not in config:
        return mapping
    
    agg_sheet_name = config['aggregation_structure'].get('sheet')
    if not agg_sheet_name:
        return mapping
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if agg_sheet_name not in wb.sheetnames:
            return mapping
        
        sheet = wb[agg_sheet_name]
        # 첫 5행을 헤더로 간주하고 시작
        for row_idx in range(4, sheet.max_row + 1):
            code_cell = sheet.cell(row_idx, 1)
            name_cell = sheet.cell(row_idx, 8)  # 산업이름
            
            if code_cell.value and name_cell.value:
                code = str(code_cell.value).strip()
                name = str(name_cell.value).strip()
                mapping[code] = name
        
        wb.close()
    except Exception:
        pass
    
    return mapping

print("\n" + "="*70)
print(" 📄 담당자 최종 검토용 HTML 생성 중...")
print("="*70 + "\n")

sectors = [
    'manufacturing',
    'service', 
    'consumption',
    'construction',
    'export',
    'import',
    'price',
    'employment',
    'migration'
]

success_count = 0
failed = []

for sector_id in sectors:
    try:
        config = next(s for s in SECTOR_REPORTS if s['id'] == sector_id)
        sector_name = config['name']
        
        print(f"📊 {sector_name} 생성 중...", end=" ", flush=True)
        
        gen = UnifiedReportGenerator(sector_id, excel_path, year, quarter)
        data = gen.extract_all_data()
        
        # 전국 데이터 찾기
        table_data = data.get('table_data', [])
        nationwide = next((row for row in table_data if row.get('region_name') == '전국'), None)
        
        if not nationwide:
            print(f"❌ 전국 데이터 없음")
            failed.append(f"{sector_name}: 전국 데이터 없음")
            continue
        
        # 생성기에서 직접 업종 데이터 추출 (엑셀에서)
        industries = gen._extract_industry_data('전국')
        
        # 산업 이름 매핑 적용 (코드 -> 이름)
        code_to_name = get_industry_name_mapping(sector_id)
        if code_to_name:
            for ind in industries:
                if 'name' in ind and ind['name'] in code_to_name:
                    ind['name'] = code_to_name[ind['name']]
        
        # 통계 지수 제외 ("총지수" 등 제외하고 실제 업종만 선택)
        filtered_industries = [
            ind for ind in industries
            if ind.get('name') and '총' not in ind.get('name', '') and '합' not in ind.get('name', '')
        ]
        
        if filtered_industries:
            # 변화도 기준으로 정렬
            sorted_industries = sorted(
                filtered_industries,
                key=lambda x: abs(x.get('change_rate', 0) or 0),
                reverse=True
            )[:15]  # 상위 15개
        else:
            sorted_industries = []
        
        # 기본 HTML 템플릿 생성
        industries_html = ""
        for idx, industry in enumerate(sorted_industries, 1):
            name = industry.get('name', 'N/A')
            value = industry.get('value', 'N/A')
            rate = industry.get('change_rate', 'N/A')
            rate_class = 'positive' if isinstance(rate, (int, float)) and rate >= 0 else 'negative'
            
            industries_html += f"""
            <tr>
                <td>{idx}</td>
                <td>{name}</td>
                <td>{value if value != 'N/A' else 'N/A'}</td>
                <td class="{rate_class}">{rate if rate != 'N/A' else 'N/A'}</td>
            </tr>"""
        
        # 전국 값 처리
        current_val = nationwide.get('current_index', 'N/A')
        change_val = nationwide.get('change_rate', 'N/A')
        change_class = 'positive' if isinstance(change_val, (int, float)) and change_val >= 0 else 'negative'
        
        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{sector_name} - {year}년 {quarter}분기 전국 데이터</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        .container {{ 
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 10px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}
        .header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
        }}
        .header p {{
            font-size: 1.1em;
            opacity: 0.9;
        }}
        .content {{ padding: 40px; }}
        .summary {{
            background: #f8f9fa;
            border-left: 4px solid #667eea;
            padding: 30px;
            margin-bottom: 40px;
            border-radius: 5px;
        }}
        .summary h2 {{
            color: #667eea;
            margin-bottom: 20px;
            font-size: 1.5em;
        }}
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-top: 20px;
        }}
        .summary-item {{
            background: white;
            padding: 20px;
            border-radius: 5px;
            border: 1px solid #e0e0e0;
            text-align: center;
        }}
        .summary-label {{
            color: #666;
            font-size: 0.9em;
            text-transform: uppercase;
            letter-spacing: 1px;
            margin-bottom: 10px;
        }}
        .summary-value {{
            font-size: 2em;
            font-weight: bold;
            color: #667eea;
        }}
        .positive {{ color: #27ae60 !important; }}
        .negative {{ color: #e74c3c !important; }}
        h2 {{
            color: #667eea;
            margin-top: 30px;
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 2px solid #667eea;
        }}
        .data-table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            font-size: 0.95em;
        }}
        .data-table th {{
            background: #667eea;
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: 600;
        }}
        .data-table td {{
            padding: 12px 15px;
            border-bottom: 1px solid #e0e0e0;
        }}
        .data-table tr:hover {{
            background: #f5f5f5;
        }}
        .data-table tr:nth-child(even) {{
            background: #f9f9f9;
        }}
        .footer {{
            background: #f5f5f5;
            padding: 20px 40px;
            text-align: center;
            color: #999;
            font-size: 0.9em;
            border-top: 1px solid #e0e0e0;
        }}
        .narrative {{
            background: #f8f9fa;
            padding: 25px;
            border-left: 4px solid #667eea;
            margin-bottom: 30px;
            border-radius: 5px;
            line-height: 1.8;
            color: #333;
        }}
        .narrative p {{
            margin-bottom: 15px;
        }}
        .narrative strong {{
            color: #667eea;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{sector_name}</h1>
            <p>{year}년 {quarter}분기 지역경제동향 전국 데이터</p>
        </div>
        
        <div class="content">
            <div class="summary">
                <h2>📊 주요 지표</h2>
                <div class="summary-grid">
                    <div class="summary-item">
                        <div class="summary-label">지역</div>
                        <div class="summary-value">{nationwide.get('region_name', 'N/A')}</div>
                    </div>
                    <div class="summary-item">
                        <div class="summary-label">현재 지수</div>
                        <div class="summary-value">{current_val if current_val != 'N/A' else '미제공'}</div>
                    </div>
                    <div class="summary-item">
                        <div class="summary-label">증감률</div>
                        <div class="summary-value {change_class}">
                            {change_val if change_val != 'N/A' else '미제공'}%
                        </div>
                    </div>
                </div>
            </div>
            
            <div class="narrative">
                <p><strong>📈 {year}년 {quarter}분기 {sector_name} 동향</strong></p>
                <p>{sector_name}의 전국 지수는 <strong>{current_val if current_val != 'N/A' else '미제공'}</strong>으로 나타났으며, 전기 대비 <strong>{change_val if change_val != 'N/A' else '미제공'}%</strong> {get_terms(sector_id, change_val if isinstance(change_val, (int, float)) else 0)[1]}했습니다. 이는 국내 경제 상황의 변화를 반영하고 있습니다.</p>
                <p>상위 업종/지표를 살펴보면 다양한 산업 분야에서 차별화된 변화가 나타나고 있습니다. 아래의 상세 표에서 각 업종별 지수와 증감률을 확인할 수 있으며, 이를 통해 구체적인 시장 동향을 파악할 수 있습니다.</p>
            </div>
            
            <h2>🏭 업종/지표별 동향 (상위 15개)</h2>
            <table class="data-table">
                <thead>
                    <tr>
                        <th style="width: 50px;">#</th>
                        <th>업종명/지표</th>
                        <th style="width: 150px;">현재 지수</th>
                        <th style="width: 150px;">증감률 (%)</th>
                    </tr>
                </thead>
                <tbody>
INDUSTRIES_PLACEHOLDER
                </tbody>
            </table>
        </div>
        
        <div class="footer">
            <p>📅 생성일시: {year}년 {quarter}분기 | 📁 데이터 출처: 분석표_25년 3분기</p>
        </div>
    </div>
</body>
</html>"""
        # industries_html을 템플릿에 삽입
        html_content = html_content.replace("INDUSTRIES_PLACEHOLDER", industries_html)
        
        output_file = output_dir / f"{sector_name}_전국_{year}년{quarter}분기.html"
        output_file.write_text(html_content, encoding='utf-8')
        
        print(f"✅")
        success_count += 1
        
    except Exception as e:
        print(f"❌")
        failed.append(f"{sector_name}: {str(e)[:50]}")

print("\n" + "="*70)
print(f" ✅ 생성 완료: {success_count}/9개 부문")
print(f" 📁 출력 위치: {output_dir}")
print("="*70)

if failed:
    print("\n⚠️  실패한 부문:")
    for f in failed:
        print(f"  - {f}")

print()
