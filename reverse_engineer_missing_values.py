"""
결측치 역산 스크립트
정답 이미지와 현재 생성 결과를 비교하여 결측치 값을 역산하고 새로운 엑셀 파일을 생성합니다.
"""

import re
import math
import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.template_manager import TemplateManager
from src.excel_extractor import ExcelExtractor
from src.template_filler import TemplateFiller
from src.period_detector import PeriodDetector
from src.schema_loader import SchemaLoader


class MissingValueReverseEngineer:
    """결측치 역산 클래스"""
    
    def __init__(self, excel_path: str, template_path: str, correct_answer_image_path: str = None):
        """
        초기화
        
        Args:
            excel_path: 엑셀 파일 경로
            template_path: 템플릿 파일 경로
            correct_answer_image_path: 정답 이미지 경로 (선택사항, OCR 미구현 시 수동 입력)
        """
        self.excel_path = Path(excel_path)
        self.template_path = Path(template_path)
        self.correct_answer_image_path = correct_answer_image_path
        self.excel_extractor = ExcelExtractor(str(excel_path))
        self.excel_extractor.load_workbook()
        self.template_manager = TemplateManager(str(template_path))
        self.template_manager.load_template()
        self.schema_loader = SchemaLoader()
        self.template_filler = TemplateFiller(self.template_manager, self.excel_extractor, self.schema_loader)
        self.period_detector = PeriodDetector(self.excel_extractor)
        
        # 워크북 로드 (수정용)
        self.workbook = load_workbook(self.excel_path)
        
    def get_current_values(self, year: int = 2025, quarter: int = 2) -> Dict[str, Any]:
        """
        현재 엑셀 파일로 생성된 HTML에서 값들을 추출합니다.
        
        Args:
            year: 연도
            quarter: 분기
            
        Returns:
            마커별 현재 값 딕셔너리
        """
        html_content = self.template_filler.fill_template(year=year, quarter=quarter)
        markers = self.template_manager.extract_markers()
        
        current_values = {}
        # 마커가 HTML에서 치환된 값을 찾기
        # 마커 패턴: {시트명:주소} 또는 {시트명:주소:연산}
        for marker in markers:
            marker_str = marker['full_match']
            # HTML에서 마커가 치환된 값 찾기 (마커 주변의 텍스트)
            # 마커 앞뒤의 패턴을 찾아서 값을 추출
            # 여러 패턴 시도
            patterns = [
                re.escape(marker_str) + r'[^>]*>([^<]+)<',
                r'<[^>]*>' + re.escape(marker_str) + r'</[^>]*>',
            ]
            
            for pattern in patterns:
                match = re.search(pattern, html_content)
                if match:
                    value_str = match.group(1).strip() if match.lastindex else ''
                    if value_str and value_str != marker_str:
                        current_values[marker_str] = value_str
                        break
            
            # 마커를 직접 찾을 수 없는 경우, fill_template의 결과를 직접 사용
            # 이 경우 마커가 이미 치환되었을 가능성이 높음
            if marker_str not in current_values:
                # 마커 주변의 숫자나 텍스트 찾기
                # 마커 위치 찾기
                marker_pos = html_content.find(marker_str)
                if marker_pos == -1:
                    # 이미 치환되었을 가능성 - 마커 주변의 값 찾기 시도
                    continue
        
        return current_values
    
    def extract_value_from_text(self, text: str) -> Optional[float]:
        """
        텍스트에서 숫자 값 추출 (천 단위 구분 제거, 퍼센트 처리)
        
        Args:
            text: 텍스트 (예: "1,234.5", "5.5%", "123")
            
        Returns:
            숫자 값 (None if not found)
        """
        if not text or text.strip() == 'N/A':
            return None
        
        # 천 단위 구분 제거
        cleaned = text.replace(',', '').replace(' ', '').strip()
        
        # 퍼센트 제거
        if cleaned.endswith('%'):
            cleaned = cleaned[:-1]
        
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return None
    
    def calculate_missing_value(
        self, 
        target_growth_rate: float, 
        known_value: Optional[float], 
        is_current_missing: bool
    ) -> Optional[float]:
        """
        증감률과 하나의 값을 알고 있을 때, 다른 값을 역산합니다.
        
        증감률 공식: growth_rate = ((current / prev) - 1) * 100
        
        Args:
            target_growth_rate: 목표 증감률 (%)
            known_value: 알려진 값 (current 또는 prev)
            is_current_missing: True면 current가 결측, False면 prev가 결측
            
        Returns:
            역산된 값 (None if cannot calculate)
        """
        if known_value is None or math.isnan(known_value) or math.isinf(known_value):
            return None
        
        if math.isnan(target_growth_rate) or math.isinf(target_growth_rate):
            return None
        
        # 증감률을 소수로 변환
        growth_ratio = target_growth_rate / 100.0
        
        if is_current_missing:
            # current = prev * (1 + growth_ratio)
            return known_value * (1 + growth_ratio)
        else:
            # prev = current / (1 + growth_ratio)
            if abs(1 + growth_ratio) < 1e-10:  # 0으로 나누기 방지
                return None
            return known_value / (1 + growth_ratio)
    
    def find_cell_for_marker(
        self, 
        marker: Dict[str, str], 
        sheet_name: str, 
        year: int, 
        quarter: int
    ) -> List[Tuple[str, int, int, bool]]:
        """
        마커에 해당하는 엑셀 셀 위치를 찾습니다.
        
        Args:
            marker: 마커 정보
            sheet_name: 시트 이름
            year: 연도
            quarter: 분기
            
        Returns:
            [(sheet_name, row, col, is_current)] 리스트
            is_current: True면 현재 분기, False면 전년 동분기
        """
        cell_positions = []
        cell_address = marker.get('cell_address', '')
        operation = marker.get('operation', '')
        
        sheet = self.excel_extractor.get_sheet(sheet_name)
        
        # 헤더 기반 키인 경우 (동적 마커) - 직접 처리하지 않고 템플릿 필러 사용
        if not re.match(r'^[A-Z]+\d+', cell_address.upper()):
            # 동적 마커는 복잡하므로 나중에 처리
            return []
        
        # 셀 주소가 범위인 경우 (예: A1:A5)
        if ':' in cell_address:
            # 범위의 첫 번째와 마지막 셀 파싱
            start_cell, end_cell = cell_address.split(':')
            start_row, start_col = self.excel_extractor.parse_cell_address(start_cell)
            end_row, end_col = self.excel_extractor.parse_cell_address(end_cell)
            
            # 증감률 계산인 경우
            if operation in ['증감률', 'growth_rate', '증가율']:
                # template_filler의 _get_quarter_columns 메서드 사용
                try:
                    current_col, prev_col = self.template_filler._get_quarter_columns(year, quarter, sheet_name)
                    
                    if prev_col and current_col:
                        for row in range(start_row, end_row + 1):
                            cell_positions.append((sheet_name, row, prev_col, False))
                            cell_positions.append((sheet_name, row, current_col, True))
                except Exception:
                    pass
        else:
            # 단일 셀 주소
            row, col = self.excel_extractor.parse_cell_address(cell_address)
            cell_positions.append((sheet_name, row, col, True))
        
        return cell_positions
    
    def reverse_engineer_missing_values(
        self,
        correct_values: Dict[str, str],
        year: int = 2025,
        quarter: int = 2
    ) -> Dict[Tuple[str, int, int], float]:
        """
        정답 값과 현재 값을 비교하여 결측치를 역산합니다.
        
        Args:
            correct_values: {마커: 정답값} 딕셔너리 (예: {"{광공업생산:A1:A2:증감률}": "5.5"})
            year: 연도
            quarter: 분기
            
        Returns:
            {(sheet_name, row, col): 역산된 값} 딕셔너리
        """
        markers = self.template_manager.extract_markers()
        
        missing_value_map = {}
        
        for marker in markers:
            marker_str = marker['full_match']
            sheet_name = marker['sheet_name']
            
            if marker_str not in correct_values:
                continue
            
            correct_value_str = correct_values[marker_str]
            correct_value = self.extract_value_from_text(correct_value_str)
            
            if correct_value is None:
                continue
            
            # 증감률 마커인 경우
            operation = marker.get('operation', '')
            if operation in ['증감률', 'growth_rate', '증가율']:
                if correct_value is not None:
                    # 해당 마커의 셀 위치 찾기
                    cell_positions = self.find_cell_for_marker(marker, sheet_name, year, quarter)
                    
                    for sheet_name_pos, row, col, is_current in cell_positions:
                        sheet_obj = self.workbook[sheet_name_pos]
                        cell_value = sheet_obj.cell(row=row, column=col).value
                        
                        # 결측치 확인 (None, 빈 문자열, '-', 또는 1.0)
                        is_missing = False
                        if cell_value is None:
                            is_missing = True
                        elif isinstance(cell_value, str):
                            if not cell_value.strip() or cell_value.strip() == '-':
                                is_missing = True
                        elif isinstance(cell_value, (int, float)):
                            if cell_value == 1.0:
                                is_missing = True
                        
                        if is_missing:
                            # 다른 값(prev 또는 current) 찾기
                            for other_sheet, other_row, other_col, other_is_current in cell_positions:
                                if other_is_current != is_current:
                                    other_sheet_obj = self.workbook[other_sheet]
                                    other_value = other_sheet_obj.cell(row=other_row, column=other_col).value
                                    
                                    if other_value is not None and not (isinstance(other_value, str) and (not other_value.strip() or other_value.strip() == '-')):
                                        try:
                                            other_num = float(other_value)
                                            if other_num != 1.0:
                                                # 역산
                                                calculated_value = self.calculate_missing_value(
                                                    correct_value,  # 목표 증감률
                                                    other_num,      # 알려진 값
                                                    is_current      # 현재가 결측인지
                                                )
                                                if calculated_value is not None:
                                                    key = (sheet_name_pos, row, col)
                                                    missing_value_map[key] = calculated_value
                                                break
                                        except (ValueError, TypeError):
                                            continue
            
            # 단순 값 마커인 경우 (증감률 계산이 아닌 경우)
            # 이 경우는 정답 값으로 직접 대체
            elif correct_value is not None:
                cell_positions = self.find_cell_for_marker(marker, sheet_name, year, quarter)
                for sheet_name_pos, row, col, _ in cell_positions:
                    sheet_obj = self.workbook[sheet_name_pos]
                    cell_value = sheet_obj.cell(row=row, column=col).value
                    
                    # 결측치 확인
                    is_missing = False
                    if cell_value is None:
                        is_missing = True
                    elif isinstance(cell_value, str):
                        if not cell_value.strip() or cell_value.strip() == '-':
                            is_missing = True
                    elif isinstance(cell_value, (int, float)):
                        if cell_value == 1.0:
                            is_missing = True
                    
                    if is_missing:
                        key = (sheet_name_pos, row, col)
                        missing_value_map[key] = correct_value
        
        return missing_value_map
    
    def save_modified_excel(self, output_path: str, missing_value_map: Dict[Tuple[str, int, int], float]):
        """
        역산된 결측치 값으로 엑셀 파일을 수정하여 저장합니다.
        
        Args:
            output_path: 출력 파일 경로
            missing_value_map: {(sheet_name, row, col): 값} 딕셔너리
        """
        for (sheet_name, row, col), value in missing_value_map.items():
            if sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                sheet.cell(row=row, column=col).value = value
        
        self.workbook.save(output_path)
        print(f"수정된 엑셀 파일 저장: {output_path}")
        print(f"총 {len(missing_value_map)}개의 셀을 수정했습니다.")


def main():
    """메인 함수"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='결측치 역산 스크립트 - 정답 이미지와 비교하여 결측치 값을 역산합니다.'
    )
    
    parser.add_argument('--excel', '-e', required=True, help='엑셀 파일 경로')
    parser.add_argument('--template', '-t', required=True, help='템플릿 파일 경로')
    parser.add_argument('--correct-values', '-c', required=True, help='정답 값 JSON 파일 경로 (또는 직접 입력)')
    parser.add_argument('--output', '-o', required=True, help='출력 엑셀 파일 경로')
    parser.add_argument('--year', '-y', type=int, default=2025, help='연도 (기본값: 2025)')
    parser.add_argument('--quarter', '-q', type=int, default=2, help='분기 (기본값: 2)')
    
    args = parser.parse_args()
    
    # 정답 값 로드
    correct_values_path = Path(args.correct_values)
    if correct_values_path.exists() and correct_values_path.suffix == '.json':
        with open(correct_values_path, 'r', encoding='utf-8') as f:
            correct_values = json.load(f)
    else:
        # JSON 파일이 아니면 직접 파싱 시도 (형식: 마커=값)
        correct_values = {}
        with open(correct_values_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if '=' in line:
                    marker, value = line.split('=', 1)
                    correct_values[marker.strip()] = value.strip()
    
    # 역산 엔지니어 초기화
    engineer = MissingValueReverseEngineer(
        excel_path=args.excel,
        template_path=args.template
    )
    
    # 역산 수행
    print("결측치 역산 중...")
    missing_value_map = engineer.reverse_engineer_missing_values(
        correct_values=correct_values,
        year=args.year,
        quarter=args.quarter
    )
    
    # 수정된 엑셀 파일 저장
    engineer.save_modified_excel(args.output, missing_value_map)
    
    print("\n역산된 값들:")
    for (sheet_name, row, col), value in missing_value_map.items():
        col_letter = get_column_letter(col)
        print(f"  {sheet_name}!{col_letter}{row}: {value}")


if __name__ == '__main__':
    main()

